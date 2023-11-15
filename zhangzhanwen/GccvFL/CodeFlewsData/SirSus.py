import datetime
# import ijson
import os
import math
import util
import openpyxl
import random
import pandas as pd

import mbfl.mbfl_for
import mbfl.command
import sbfl.sbfl_for
import sbfl.command

from ChmbflSus import Tools
from CodeFlewsData.data_codeflaws import MutationRule


mbfl_for_list = [
    "mbfl.mbfl_for.Tarantula",
    "mbfl.mbfl_for.Op2",
    "mbfl.mbfl_for.Jaccard",
    "mbfl.mbfl_for.Ochiai",
    "mbfl.mbfl_for.Dstar",
    "mbfl.mbfl_for.GP13",
    "mbfl.mbfl_for.Naish1",
    "mbfl.mbfl_for.Barinel",
    # "mbfl.mbfl_for.muse",
    # "mbfl.mbfl_for.nsus",
]

sbfl_for_list = [
    "sbfl.sbfl_for.Tarantula",
    "sbfl.sbfl_for.Op2",
    "sbfl.sbfl_for.Jaccard",
    "sbfl.sbfl_for.Ochiai",
    "sbfl.sbfl_for.Dstar",
    "sbfl.sbfl_for.GP13",
    "sbfl.sbfl_for.Naish1",
    "sbfl.sbfl_for.Barinel",
]

continuecreat = False


# data_dirpath = '/home/cyxy/sir_result' if Tools().get_host_ip() == '202.4.130.29' else './sir'
data_dirpath = {
    '202.4.130.29': '/home/cyxy/sir_result',
    '222.199.230.148': '/home/cyxy/sir_result',
    '192.168.50.130': './sir/sir_result',
    '222.199.230.216': '/home/caiyuxiaoyang/sir_result',
    '192.168.50.26': 'F:/Prgorams/Python/access/sir/sir_result',
    '192.168.1.105': 'F:/Prgorams/Python/access/sir/sir_result',
}[Tools().get_host_ip()]


class Data(object):
    def __init__(self, path):
        self._data = 'empty'
        self.path = path

    # 访问器 getter方法
    @property
    def data(self):
        try:
            if self._data == 'empty':
                self._data = SirTools().sir_rules(self.path, covread=True, firstread=True, oneline=True)
        except Exception as e:
            self._data = False
            print(self.path, 'less', e)
        return self._data

class SirTools:

    def sir_rules(self, path, firstread=True, highread=True, covread=True, oneline=False):
        st = datetime.datetime.now()

        name = '_'.join(path.split('/')[3:])
        # print('%s 开始读取 %s' % (datetime.datetime.now(), path))
        # print(os.listdir(path))

        # or_list
        with open(os.path.join(path, 'output', 'res_vector.in'), 'r') as f:
            or_list = list(map(lambda x:int(x), f.readlines()[0].strip()))

        # line_len
        line_len = -1
        linedefectpath = os.path.join(path.replace('sir_result', 'Artificial'), 'defect_root', 'source')
        if not os.path.exists(linedefectpath):
            linedefectpath = os.path.join(path.replace('sir_result', 'dataset/Artificial'), 'defect_root', 'source')

        for file in os.listdir(linedefectpath):
            if '.c' not in file:
                continue
            with open(os.path.join(linedefectpath, file), 'r') as f:
                line_len = len(f.readlines())
        if line_len == -1:
            print('          ', 'line_len error')
            return False

        # sbfl
        cov = []
        if covread:
            with open(os.path.join(path, 'output', 'res_cov_matrix.in'), 'r') as f:
                for test_cov in f.readlines():
                    cov_t = []
                    try:
                        for line, covres in enumerate(map(lambda x:int(x), test_cov.strip())):
                            if covres == 1:
                                cov_t.append(line+1)
                        cov.append(cov_t)
                    except:
                        continue

        # Fault_Record
        with open(os.path.join(path, 'output', 'Fault_Record.txt'), 'r') as f:
            Fault_Record = []
            for line in f.readlines():
                if not line.strip():
                    continue
                Fault_Record.append(int(line.split(':')[0].split(' ')[1]))


        # All First Order Mutant
        fom_list = []
        if firstread:
            with open(os.path.join(path, 'output', 'all_execute_mutant.txt'), 'r') as f:
                fommessagelist = f.readlines()
            with open(os.path.join(path, 'output', 'res_original_version.in'), 'r') as f:
                fomoutlist = f.readlines()
            with open(os.path.join(path, 'output', 'res_fault_version.in'), 'r') as f:
                fomkilllist = f.readlines()
            if not len(fommessagelist) == len(fomoutlist) or not len(fomoutlist) == len(fomkilllist):
                print(path, 'output', 'num error')
                return False
            for i in range(len(fommessagelist)):
                line = fommessagelist[i].strip('\n')
                out_list = list(map(lambda x:int(x), fomoutlist[i].strip()))
                if not len(or_list) == len(out_list):
                    print(path, 'out_list', 'len error')
                    return False
                kill_list = list(map(lambda x:int(x), fomkilllist[i].strip()))
                if not len(or_list) == len(kill_list):
                    print(path, 'kill_list', 'len error')
                    return False
                for key in ['Mutant_index: ', 'Source_line: ', 'index: ', 'original: ', '   mutated: ']:
                    line = line.replace(key, '<:>')
                fom = line.split('<:>')
                fom_list.append({
                    'message': [[int(fom[2]), '', '', fom[4],  fom[5], int(fom[3])]],
                    "out_list": out_list,
                    "kill_list": kill_list,
                    'time': 0,
                })
        # print('fomnum', len(fom_list))

        # All Second Order Mutant
        hom_out_list = []
        if highread:
            with open(os.path.join(path, '_entire_randomization_higher_order_mutant_set', 'all_execute_mutant.txt'), 'r') as f:
                fommessagelist = f.readlines()
            with open(os.path.join(path, '_entire_randomization_higher_order_mutant_set', 'res_original_version.in'), 'r') as f:
                fomoutlist = f.readlines()
            with open(os.path.join(path, '_entire_randomization_higher_order_mutant_set', 'res_fault_version.in'), 'r') as f:
                fomkilllist = f.readlines()
            if not len(fommessagelist) == len(fomoutlist) or not len(fomoutlist) == len(fomkilllist):
                print(path, 'output', 'num error')
                return False
            for i in range(len(fommessagelist)):
                out_list = list(map(lambda x:int(x), fomoutlist[i].strip()))
                if not len(or_list) == len(out_list):
                    print(path, 'out_list', 'len error')
                    return False
                kill_list = list(map(lambda x:int(x), fomkilllist[i].strip()))
                if not len(or_list) == len(kill_list):
                    print(path, 'kill_list', 'len error')
                    return False
                fom = []
                line = fommessagelist[i]
                for mutant_message in line.strip('\n').split('   |||')[1:]:
                    for key in ['   line: ', '   index: ', '   original: ', '   mutated: ']:
                        mutant_message = mutant_message.replace(key, '<:>')
                    fom.append([
                        int(mutant_message.split('<:>')[1]),
                        '',
                        '',
                        mutant_message.split('<:>')[3],
                        mutant_message.split('<:>')[4],
                        int(mutant_message.split('<:>')[2]),
                    ])
                hom_out_list.append({
                    'message': fom,
                    "out_list": out_list,
                    "kill_list": kill_list,
                    'time': 0,
                })
        # print('homnum', len(hom_out_list) )

        # All OneLine Second Order Mutant
        oneline_out_list = []
        if oneline:
            onelinepath = path.replace('sir_result', 'sir_oneline_result')
            with open(os.path.join(onelinepath, '_entire_randomization_higher_order_mutant_set', 'all_execute_mutant.txt'), 'r') as f:
                fommessagelist = f.readlines()
            with open(os.path.join(onelinepath, '_entire_randomization_higher_order_mutant_set', 'res_original_version.in'), 'r') as f:
                fomoutlist = f.readlines()
            with open(os.path.join(onelinepath, '_entire_randomization_higher_order_mutant_set', 'res_fault_version.in'), 'r') as f:
                fomkilllist = f.readlines()
            if not len(fommessagelist) == len(fomoutlist) or not len(fomoutlist) == len(fomkilllist):
                print(onelinepath, 'output', 'num error')
                return False
            for i in range(len(fommessagelist)):
                out_list = list(map(lambda x:int(x), fomoutlist[i].strip()))
                if not len(or_list) == len(out_list):
                    print(path, 'out_list', 'len error')
                    return False
                kill_list = list(map(lambda x:int(x), fomkilllist[i].strip()))
                if not len(or_list) == len(kill_list):
                    print(path, 'kill_list', 'len error')
                    return False
                fom = []
                line = fommessagelist[i]
                try:
                    for mutant_message in line.strip('\n').split('   |||')[1:]:
                        for key in ['   line: ', '   index: ', '   original: ', '   mutated: ']:
                            mutant_message = mutant_message.replace(key, '<:>')
                        fom.append([
                            int(mutant_message.split('<:>')[1]),
                            '',
                            '',
                            mutant_message.split('<:>')[3],
                            mutant_message.split('<:>')[4],
                            int(mutant_message.split('<:>')[2]),
                        ])
                except:
                    continue
                oneline_out_list.append({
                    'message': fom,
                    "out_list": out_list,
                    "kill_list": kill_list,
                    'time': 0,
                })
        # print('homnum', len(hom_out_list) )


        et = datetime.datetime.now()
        print('%s      用时 %s, 读取完成 %s' % (et, et-st, path))

        return {name:
            {
                'or_list': or_list,
                'oneline_out_list': oneline_out_list,
                'linelen': line_len,
                'hom_out_list': hom_out_list,
                'fom_list': fom_list,
                'Fault_Record': Fault_Record,
                'sbfl': {
                    'cov': cov,
                    'time': 0
                },
            }
        }


    # 获取sir路径
    def version_path(self):
        ipProjects = {
            '202.4.130.29': ['tcas', 'sed', 'printtokens2', 'printtokens'],    # 4
            '222.199.230.148': ['sed'],    # 5
            '222.199.230.216': ['printtokens2'],    # 6
        }
        passlist = [
            ['printtokens', 'five_faults', 'v2'],
            ['printtokens', 'five_faults', 'v3'],
            ['printtokens', 'four_faults', 'v2'],
            ['printtokens', 'four_faults', 'v3'],
            ['printtokens', 'three_faults', 'v2'],
            ['printtokens', 'three_faults', 'v5'],
            ['printtokens', 'two_faults', 'v2'],
            ['printtokens', 'two_faults', 'v3'],
            ['printtokens', 'two_faults', 'v4'],
            ['printtokens', 'two_faults', 'v5'],

            ['printtokens2', 'five_faults', 'v4'],
            ['printtokens2', 'five_faults', 'v5'],
            ['printtokens2', 'four_faults', 'v5'],
            ['printtokens2', 'one_fault', 'v6'],
            ['printtokens2', 'two_faults', 'v5'],

            ['tcas', 'one_fault', 'v8'],
            ['tcas', 'one_fault', 'v9'],
            ['tcas', 'one_fault', 'v7'],
            ['sed', 'three_faults', 'v2'],
            ['sed', 'three_faults', 'v5'],
            # ['printtokens', 'five_faults', 'v1'],
        ]

        pathlist = []
        for project in os.listdir(data_dirpath):
            if '.' in project:
                continue
            if project not in ipProjects[Tools().get_host_ip()]:
                continue


            project_path = os.path.join(data_dirpath, project)
            for faultnum in os.listdir(project_path):
                if '.' in faultnum:
                    continue
                # if faultnum == 'one_fault':
                #     continue
                faultnum_path = os.path.join(data_dirpath, project, faultnum)
                for version in os.listdir(faultnum_path):
                    if '.' in version:
                        continue
                    addpath = [project, faultnum, version]
                    if addpath in passlist:
                        continue
                    pathlist.append([project, faultnum, version])
                    # print([project, faultnum, version])
        if Tools().get_host_ip() == '222.199.230.148':
            pathlist.reverse()
        return pathlist



class Get_sus:
    # 计算last2first方法怀疑度
    def Last2First(self, data, times=0.5):
        if not data:
            return False
        try:

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_out_list = data['hom_list']

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # ------------------------------------------------
        # 生成som组合
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            random_fomloc = random.sample(range(len(fom_list)), len(fom_list))
            while len(random_fomloc) > 1 and sus_data['homnum'] < times*sus_data['fomnum']:
                locfirst = random_fomloc.pop(0)
                # print(len(random_fomloc), sus_data['homnum'])
                fom_first = fom_list[locfirst]
                for i in range(len(random_fomloc)-1, -1, -1):
                    # fom_last = fom_list[i]
                    loclast = random_fomloc[i]
                    fom_last = fom_list[loclast]
                    if fom_first['message'][0][0] == fom_last['message'][0][0]:
                        continue
                    som = sorted([fom_first['message'][0], fom_last['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom = hom_out_list[loc]
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    random_fomloc.pop(i)
                    break
            if times == 0.5:
                break

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算differentOperator方法怀疑度
    def DifferentOperator(self, data, times=0.5):
        if not data:
            return False
        try:
            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_out_list = data['hom_list']

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            random_fomlist = random.sample(fom_list, len(fom_list))
            while len(random_fomlist) > 1 and sus_data['homnum'] < times*sus_data['fomnum']:
                fom_1 = random_fomlist.pop(0)
                message_1 = fom_1['message'][0]
                for i, fom in enumerate(random_fomlist):
                    message = fom['message'][0]
                    if message_1[0] == message[0] or message_1[3] == message[3]:
                        continue
                    som = sorted([message_1, message], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    random_fomlist.pop(i)
                    hom = hom_out_list[loc]
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, som)] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(som) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    break

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算randomMix方法怀疑度
    def RandomMix(self, data, times=0.5):
        if not data:
            return False
        try:

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_out_list = data['hom_list']

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # ------------------------------------------------
        # 生成som组合
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            random_fom_list = random.sample(fom_list, len(fom_list))
            while sus_data['homnum'] < times*sus_data['fomnum'] and len(random_fom_list) > 0:
                loc1 = random.sample(range(len(random_fom_list)), 1)[0]
                fom1 = random_fom_list.pop(loc1)
                for loc2 in random.sample(range(len(random_fom_list)), len(random_fom_list)):
                    fom2 = random_fom_list[loc2]
                    if fom1['message'][0][0] == fom2['message'][0][0]:
                        continue
                    som_message = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x:x[0])
                    try:
                        loc = hom_list_simple.index(som_message)
                        hom = hom_out_list[loc]
                        random_fom_list.pop(loc2)
                    except:
                        continue
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, som_message)] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(som_message) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    break

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算fom的times倍数的 random
    def Random(self, data, times=1000):
        if not data:
            return False
        try:

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_out_list = data['hom_list']

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            fomnum = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]

        out_dic = dict()
        kill_dic = dict()
        for hom in random.sample(hom_out_list,  min(math.ceil(fomnum*times), len(hom_out_list))):
            message = hom['message']
            if not MutationRule(message[0], message[1]):
                # 同一行进行了叠加变异
                continue
            out_list = hom['out_list']
            kill_list = hom['kill_list']
            sus_data['homnum'] += 1
            precisions[Tools().precision_calculate(Fault_Record, message)] += 1
            varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(message) else x[1],
                                enumerate(varietys)))
            for fom in message:
                line = fom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(out_list)
                kill_dic[line].append(kill_list)

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)

                sus_data[word][mbfl_for] = dict()

                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算mbfl方法怀疑度
    def Mbfl(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            # hom_out_list = data['hom_out_list']
            # sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
        except Exception as e:
            print('数据缺失 %s' % e)
            return False


        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        for mbfl_for in mbfl_for_list:
            sus_data[mbfl_for] = dict()
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_list.append([line, sus])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            def charge(ranks, Fault_Record, touple):
                for line in Fault_Record:
                    if line not in touple['linedata'].keys():
                        return 0
                if sum(map(lambda x:x[1], ranks)) == len(ranks):
                    return 0
                # print(len(touple['linedata'].keys()), ranks)
                return 0
            # if mbfl_for == "mbfl.mbfl_for.Tarantula":
            #     charge(ranks, Fault_Record, touple)


            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[mbfl_for]['rank_best'] = rank[0]
            sus_data[mbfl_for]['rank_average'] = rank[1]
            sus_data[mbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[mbfl_for]['exam_best'] = exam[0]
            sus_data[mbfl_for]['exam_average'] = exam[1]
            sus_data[mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 计算sbfl方法怀疑度
    def Sbfl(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            Fault_Record = data["Fault_Record"]
            # fom_list = data['fom_list']
            # fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            sbfl_data = data["sbfl"]
            if "linelen" in data:
                line_len = data['linelen']
            else:
                src_path = data['path']
                try:
                    path_list = src_path.split('/')
                    while len(path_list) > 0:
                        if not path_list[0] == 'access':
                            path_list.pop(0)
                        else:
                            path_list.pop(0)
                            break
                    nowpath = os.getcwd()
                    for part in path_list:
                        nowpath = os.path.join(nowpath, part)
                    line_len = len(util.File().read_line(nowpath))
                except:
                    line_len = 0


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = 0
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
        except Exception as e:
            print('数据缺失 %s' % e)
            return False



        # ------------------------------------------------
        # 计算sbfl怀疑度
        touple = sbfl.command.touple(or_list, sbfl_data['cov'])
        for sbfl_for in sbfl_for_list:
            sus_data[sbfl_for] = dict()
            sus_list = []
            for line, touple_list in touple.items():
                sus = eval(sbfl_for)(touple_list)
                sus_list.append([line, sus])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[sbfl_for]['rank_best'] = rank[0]
            sus_data[sbfl_for]['rank_average'] = rank[1]
            sus_data[sbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[sbfl_for]['exam_best'] = exam[0]
            sus_data[sbfl_for]['exam_average'] = exam[1]
            sus_data[sbfl_for]['exam_worst'] = exam[2]
        sus_data['totaltime'] = sbfl_data['time']

        return sus_data

    # 计算sbfl方法怀疑度
    def Mcbfl(self, data_json):
        mcbfl_for_list = [
            ["sbfl.sbfl_for.Tarantula", "mbfl.mbfl_for.Tarantula"],
            ["sbfl.sbfl_for.Op2", "mbfl.mbfl_for.Op2"],
            ["sbfl.sbfl_for.Jaccard", "mbfl.mbfl_for.Jaccard"],
            ["sbfl.sbfl_for.Ochiai", "mbfl.mbfl_for.Ochiai"],
            ["sbfl.sbfl_for.Dstar", "mbfl.mbfl_for.Dstar"],
            ["sbfl.sbfl_for.GP13", "mbfl.mbfl_for.GP13"],
            ["sbfl.sbfl_for.Naish1", "mbfl.mbfl_for.Naish1"],
            ["sbfl.sbfl_for.Barinel", "mbfl.mbfl_for.Barinel"],
        ]


        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
        except Exception as e:
            print('数据缺失 %s' % e)
            return False


        # ------------------------------------------------
        # 获取mbfl元组
        mbfl.mbfl_for.type_mbfl = 'max'
        totaltime = 0
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            totaltime += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # 获取sbfl元组
        totaltime += sbfl_data['time']
        touple_sbfl = sbfl.command.touple(or_list, sbfl_data['cov'])

        sus_data['totaltime'] = totaltime

        # 计算怀疑度
        for sbfl_for, mbfl_for in mcbfl_for_list:
            sus_data[mbfl_for] = dict()
            sus_dic = dict()

            # 计算mbfl怀疑度
            for line in touple_mbfl['linedata']:
                touple_list = touple_mbfl['tf'], \
                              touple_mbfl['tp'], \
                              touple_mbfl['linedata'][line]['f'], \
                              touple_mbfl['linedata'][line]['p'], \
                              touple_mbfl['f2p'], \
                              touple_mbfl['p2f'], \
                              touple_mbfl['linedata'][line]['kf'], \
                              touple_mbfl['linedata'][line]['kp'], \
                              touple_mbfl['linedata'][line]['nf'], \
                              touple_mbfl['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_dic[line] = sus + -1

            # 计算sbfl怀疑度
            for line, touple_list in touple_sbfl.items():
                sus = eval(sbfl_for)(touple_list)
                if line in sus_dic:
                    sus_dic[line] += sus + 1
                else:
                    sus_dic[line] = sus + -1

            sus_list = []
            for line, sus in sus_dic.items():
                sus_list.append([line, sus/2])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[mbfl_for]['rank_best'] = rank[0]
            sus_data[mbfl_for]['rank_average'] = rank[1]
            sus_data[mbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[mbfl_for]['exam_best'] = exam[0]
            sus_data[mbfl_for]['exam_average'] = exam[1]
            sus_data[mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 仅使用高阶的NS-Delta
    def DeltaNS(self, data_json, times=1000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            if 'oneline_out_list' in data.keys():
                oneline_out_list = data["oneline_out_list"]
            else:
                oneline_out_list = False
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            # sus_data['homnum'] = len(hom_out_list)
            sus_data['Desrcibe'] = ''
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        # ------------------------------------------------
        # 计算一阶怀疑度
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # print(Fault_Record)
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_lines_dic = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_lines_dic:
                fom_lines_dic[line] = []
            fom_lines_dic[line].append(fom)

        hom_lines_dic = dict()
        tnum = 0
        if oneline_out_list:
            for hom in oneline_out_list:
                if not NsRule(hom['message'][0], hom['message'][1]):
                    continue
                line = hom['message'][0][0]
                if line not in hom_lines_dic:
                    hom_lines_dic[line] = []
                hom_lines_dic[line].append(hom)
                tnum += 1
        else:
            for line, fom_list_indic in fom_lines_dic.items():
                hom_lines_dic[line] = []
                for i1, fom1 in enumerate(fom_list_indic):
                    for i2, fom2 in enumerate(fom_list_indic):
                        if i1 >= i2:
                            continue
                        if not NsRule(fom1['message'][0], fom2['message'][0]):
                            continue
                        som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                        try:
                            loc = hom_list_simple.index(som)
                        except:
                            continue
                        hom_lines_dic[line].append(hom_out_list[loc])
                        tnum += 1

        # ------------------------------------------------
        powerlist = list(map(lambda x: [x[0], min(len(x[1])*math.ceil(fomnum*times/tnum), len(x[1]))],
                             hom_lines_dic.items()))
        powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        out_dic = dict()
        kill_dic = dict()
        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            hom_lines = hom_lines_dic[line]
            for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                sus_data['homnum'] += 1


        touple_ns = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            mbfl.mbfl_for.type_mbfl = 'frequency'
            sus_data[word] = dict()
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list_T = dict()
                for line in touple_mbfl['linedata']:
                    touple_list = touple_mbfl['tf'], \
                                  touple_mbfl['tp'], \
                                  touple_mbfl['linedata'][line]['f'], \
                                  touple_mbfl['linedata'][line]['p'], \
                                  touple_mbfl['f2p'], \
                                  touple_mbfl['p2f'], \
                                  touple_mbfl['linedata'][line]['kf'], \
                                  touple_mbfl['linedata'][line]['kp'], \
                                  touple_mbfl['linedata'][line]['nf'], \
                                  touple_mbfl['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] = sus

                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple_ns['linedata']:


                    touple_list = touple_ns['tf'], \
                                  touple_ns['tp'], \
                                  touple_ns['linedata'][line]['f'], \
                                  touple_ns['linedata'][line]['p'], \
                                  touple_ns['f2p'], \
                                  touple_ns['p2f'], \
                                  touple_ns['linedata'][line]['kf'], \
                                  touple_ns['linedata'][line]['kp'], \
                                  touple_ns['linedata'][line]['nf'], \
                                  touple_ns['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    if line not in sus_list_T:
                        sus_list_T[line] = sus
                    else:
                        sus_list_T[line] += sus
                    sus_line_ave = sum(sus_list_T[line])/len(sus_list_T[line])
                    if word == 'max':
                        sus_list.append([line, max(sus)-sus_line_ave])
                    elif word == 'ave':
                        sus_list.append([line, sum(sus)/len(sus)-sus_line_ave])
                    elif word == 'frequency':
                        sus_list.append([line, list(map(lambda x:x-sus_line_ave, sus))])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 计算mbfl方法怀疑度
    def DeltaMbfl(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False


        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'frequency'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        for mbfl_for in mbfl_for_list:
            sus_data[mbfl_for] = dict()
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_list.append([line, max(sus)-sum(sus)/len(sus)])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[mbfl_for]['rank_best'] = rank[0]
            sus_data[mbfl_for]['rank_average'] = rank[1]
            sus_data[mbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[mbfl_for]['exam_best'] = exam[0]
            sus_data[mbfl_for]['exam_average'] = exam[1]
            sus_data[mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 使用高阶NS-Delta 和 mbfl-Delta
    def DeltaNsMbfl(self, data, times=1000):
        if not data:
            return False
        try:

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_out_list = data['hom_list']
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        def WordNone(sus_list_format):
            frequency_dict = dict()
            for sus in sus_list_format:
                if sus not in frequency_dict:
                    frequency_dict[sus] = 0
                frequency_dict[sus] += 1
            return max(sorted(list(map(lambda x: [frequency_dict[x], x],
                                       list(frequency_dict.keys()))),
                              key=lambda x: x[1], reverse=True))[1]

        # ------------------------------------------------
        # 计算一阶怀疑度
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            # sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # print(Fault_Record)
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_lines_dic = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_lines_dic:
                fom_lines_dic[line] = []
            fom_lines_dic[line].append(fom)

        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_lines_dic.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        # ------------------------------------------------
        guide = ''
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    # sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2 + fomnum*times/2/len(sus_list)], enumerate(sus_list)))
            # powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times], enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])

        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*fomnum*times/tnum],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        out_dic = dict()
        kill_dic = dict()
        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            hom_lines = hom_lines_dic[line]
            for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                sus_data['homnum'] += 1

        resnum = math.ceil(resnum)
        while resnum > 0 and not times == 1000:
            for line, neednum in powerlist_sorted:
                if line not in hom_lines_dic:
                    continue
                hom_lines = hom_lines_dic[line]
                for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    resnum -= 1
                    sus_data['homnum'] += 1
                    if resnum == 0:
                        break
                if resnum == 0:
                    break

        touple_ns = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        mbfl.mbfl_for.type_mbfl = 'frequency'
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list_T = dict()
                sus_list_m = dict()
                for line in touple_mbfl['linedata']:
                    touple_list = touple_mbfl['tf'], \
                                  touple_mbfl['tp'], \
                                  touple_mbfl['linedata'][line]['f'], \
                                  touple_mbfl['linedata'][line]['p'], \
                                  touple_mbfl['f2p'], \
                                  touple_mbfl['p2f'], \
                                  touple_mbfl['linedata'][line]['kf'], \
                                  touple_mbfl['linedata'][line]['kp'], \
                                  touple_mbfl['linedata'][line]['nf'], \
                                  touple_mbfl['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] = sus

                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple_ns['linedata']:
                    touple_list = touple_ns['tf'], \
                                  touple_ns['tp'], \
                                  touple_ns['linedata'][line]['f'], \
                                  touple_ns['linedata'][line]['p'], \
                                  touple_ns['f2p'], \
                                  touple_ns['p2f'], \
                                  touple_ns['linedata'][line]['kf'], \
                                  touple_ns['linedata'][line]['kp'], \
                                  touple_ns['linedata'][line]['nf'], \
                                  touple_ns['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] += sus

                for line in sus_list_T.keys():
                    sus_line_ave = sum(sus_list_T[line])/len(sus_list_T[line])
                    if word == 'max':
                        sus_list.append([line, max(sus_list_T[line])-sus_line_ave])
                    elif word == 'ave':
                        sus_list.append([line, sum(sus_list_T[line])/len(sus_list_T[line])-sus_line_ave])
                    elif word == 'frequency':
                        sus_list.append([line, list(map(lambda x:x-sus_line_ave, sus_list_T[line]))])
                    elif word == 'none':
                        sus_list.append([line, WordNone(sus_list_T[line])-sus_line_ave])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data








class Sirsus:
    def __init__(self):
        self.continuecreat = True
        # self.continuecreat = False
        # self.times()
        # self.timestest()
        self.times()
        # self.newStatic()
        # self.reWriteDif()


    # 不需要倍数，需要重复次数
    def times(self):
        clasterfunction = [
            ['Sbfl', Get_sus().Sbfl],
            ['Mbfl', Get_sus().Mbfl],
            ['Mcbfl', Get_sus().Mcbfl],

            ['DeltaNS', Get_sus().DeltaNS],

            # ['Last2First', Get_sus().Last2First],
            # ['DifferentOperator', Get_sus().DifferentOperator],
            # ['RandomMix', Get_sus().RandomMix],
            # ['Random', Get_sus().Random],

            # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
            # ['ED.MBFL', Get_sus().ErrorDistribution, [True]],

            # ['NS.RANDOM', Get_sus().NS, [0]],
            # ['NS.SBFL', Get_sus().NS, [2]],
            # ['NS.MBFL', Get_sus().NS, [1]],

            # ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
            # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
            # ['MutCluster', Get_sus().MutCluster, [True, True]],
            # ['MutCluster.in', Get_sus().MutCluster, [True, False]],

            # ['NS.mbfl^MC', Get_sus().NSpMC],
            # ['NS.mbfl^ED', Get_sus().NSpED],
            # ['MC.mseer^ED', Get_sus().MCpED],
            # ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
        ]

        ftitle = [
            'Method',
            'times',
            'repeat',
            'Statement',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        repeattimes = 5
        timerange = list(range(1, 1+1))

        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 5

        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/DN')
        if not os.path.exists(docpath):
            os.makedirs(docpath)

        versionnum = len(SirTools().version_path())
        st = datetime.datetime.now()

        for file_i, [project, faultnum, version] in enumerate(SirTools().version_path()):

            nt = datetime.datetime.now()
            file_resttime = (nt-st)/file_i*(versionnum-file_i) if file_i > 0 else '???'
            print("{},    {}/{}, {}-{}-{}, 以用时:{}".format(
                nt,
                file_i,
                versionnum,
                project,
                faultnum,
                version,
                nt-st)
            )

            file = '%s-%s-%s' % (project, faultnum, version)
            path = os.path.join(data_dirpath, project, faultnum, version)

            filepath_do = os.path.join(docpath, '%s.xlsx' % (file))

            # 表格初始化
            passlist = []
            wb = False
            if self.continuecreat:
                if os.path.exists(filepath_do):
                    df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
                    for i in df.index.values:
                        row_data_read = df.loc[i, df.columns.values].to_dict()
                        passlist.append([
                            row_data_read['Method'],
                            row_data_read['times'],
                            row_data_read['repeat'],
                            row_data_read['Statement'],
                        ])
                    wb = openpyxl.load_workbook(filepath_do)
            if not wb:
                wb = openpyxl.Workbook()
                del wb['Sheet']
                for ws_name in ['exam-best', 'exam-average', 'exam-worst']:
                    ws = wb.create_sheet(ws_name)
                    ws.cell(1, 1, file)
                    for j, title in enumerate(ftitle):
                        ws.cell(2, j+1, title)
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                        for k in range(maxfaultnum):
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+maxfaultnum+k, 'rank%s' % str(k+1))

            sir_data = Data(path)

            nnum = len(passlist)
            tnum = len(clasterfunction)*len(timerange)*repeattimes*3
            lt = datetime.datetime.now()
            for parameter in clasterfunction:
                dosave = False
                for times in timerange:
                    for repeat in range(1, repeattimes+1):
                        method = parameter[0]
                        if [method, times, repeat, 'frequency'] in passlist:
                            continue
                        dosave = False
                        function = parameter[1]

                        if not sir_data.data:
                            continue
                        dosave = True

                        print("{},    {}/{}-filerest:{},    {}/{}-{}/{}/{}".format(
                            datetime.datetime.now(),
                            file_i,
                            versionnum,
                            file_resttime,
                            nnum,
                            tnum,
                            method,
                            times,
                            repeat,
                        ))
                        lt = datetime.datetime.now()

                        if parameter[0] in ['Sbfl', 'Mbfl', 'Mcbfl']:
                            sus_data_json = function(sir_data.data)
                        else:
                            if len(parameter) <= 2:
                                sus_data_json = function(sir_data.data, times=times)
                            else:
                                sus_data_json = function(sir_data.data, times=times, seting=parameter[2])


                        for tie in ['best', 'average', 'worst']:
                            ws_name = 'exam-%s' % tie
                            ws = wb[ws_name]
                            for numi, word in enumerate(['max', 'ave', 'frequency']):

                                sheetdata_list = [
                                    method,
                                    times,
                                    repeat,
                                    word,
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        ws.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                if word not in sus_data_json:
                                    wordMessage = sus_data_json
                                else:
                                    wordMessage = sus_data_json[word]

                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    '',
                                ]

                                if 'variety' in sus_data_json:
                                    varietys = sus_data_json['variety']
                                    sum_varietys = sum(varietys)
                                    sheetdata_list += list(map(lambda x: 0 if sum_varietys == 0 else x/sum_varietys,
                                                               varietys))
                                else:
                                    sheetdata_list += [0 for i in range(15)]

                                if 'precision' in sus_data_json:
                                    precisions = sus_data_json['precision']
                                    sum_precisions = sum(precisions)
                                    sheetdata_list += list(map(lambda x: 0 if sum_precisions == 0 else x/sum_precisions,
                                                               precisions))
                                else:
                                    sheetdata_list += [0 for i in range(3)]

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):

                                    if parameter[0] == 'Sbfl':
                                        formula = mbfl_for.replace('mbfl', 'sbfl')
                                    else:
                                        formula = mbfl_for

                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if formula not in wordMessage:
                                            examlist.append('')
                                            ranklist.append('')
                                            continue
                                        if fault_ith > len(wordMessage[formula]['rank_%s' % tie])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(wordMessage[formula]['exam_%s' % tie][fault_ith])
                                            ranklist.append(wordMessage[formula]['rank_%s' % tie][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    ws.cell(nnum + numi + 3, sheetdata_i+1, sheetdata)
                        nnum += 3

                        nt = datetime.datetime.now()
                        print("{},    cost:{}".format(
                            datetime.datetime.now(),
                            '???' if nnum == 0 else nt-lt,
                        ))


                # 存储文件
                if dosave:
                    wb.save(filepath_do)
                    print('文件保存 %s ' % filepath_do)
                # return

            # os.remove(filepath_undo)
            # wb.save(filepath_do)
            # print('文件保存 %s ' % filepath_do)
        return


    # 不需要倍数，需要重复次数
    def newStatic(self):

        ftitle = [
            '版本',
            '代码行数',
            '测试用例个数',
            '错误数量',
            '生成一阶变异体数量',
        ]

        wb = openpyxl.Workbook()
        ws = wb['Sheet']
        for j, title in enumerate(ftitle):
            ws.cell(1, j+1, title)

        line = 2

        for file_i, [project, faultnum, version] in enumerate(SirTools().version_path()):

            path = os.path.join(data_dirpath, project, faultnum, version)
            name = '_'.join(path.split('/')[3:])
            sir_data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
            if not sir_data:
                continue
            writeData = [
                '-'.join([project, faultnum, version]),
                sir_data[name]['linelen'],
                len(sir_data[name]['or_list']),
                len(sir_data[name]['Fault_Record']),
                len(sir_data[name]['fom_list']),
            ]
            for j, value in enumerate(writeData):
                ws.cell(line, j+1, value)
            line += 1
        wb.save('需要统计的数据集信息.xlsx')
        return


    # 不需要倍数，需要重复次数
    def timestest(self):
        clasterfunction = [
            ['Sbfl', Get_sus().Sbfl],
            ['Mbfl', Get_sus().Mbfl],
            ['Mcbfl', Get_sus().Mcbfl],

            ['DeltaNS', Get_sus().DeltaNS],

            # ['Last2First', Get_sus().Last2First],
            # ['DifferentOperator', Get_sus().DifferentOperator],
            # ['RandomMix', Get_sus().RandomMix],
            # ['Random', Get_sus().Random],

            # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
            # ['ED.MBFL', Get_sus().ErrorDistribution, [True]],

            # ['NS.RANDOM', Get_sus().NS, [0]],
            # ['NS.SBFL', Get_sus().NS, [2]],
            # ['NS.MBFL', Get_sus().NS, [1]],

            # ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
            # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
            # ['MutCluster', Get_sus().MutCluster, [True, True]],
            # ['MutCluster.in', Get_sus().MutCluster, [True, False]],

            # ['NS.mbfl^MC', Get_sus().NSpMC],
            # ['NS.mbfl^ED', Get_sus().NSpED],
            # ['MC.mseer^ED', Get_sus().MCpED],
            # ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
        ]


        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/DN')
        if not os.path.exists(docpath):
            os.makedirs(docpath)


        for filename in os.listdir(docpath):
            filepath_do = os.path.join(docpath, filename)

            # 表格初始化
            passlist = []
            df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
            for i in df.index.values:
                row_data_read = df.loc[i, df.columns.values].to_dict()
                passlist.append(row_data_read['Method'])

            less = []
            for parameter in clasterfunction:
                method = parameter[0]
                if method not in passlist:
                    less.append(method)
            if len(less) > 0:
                print(filepath_do)
                print(less)
        return

    # 不需要倍数，需要重复次数
    def yojan(self):
        reduction_list = [
            'testanalysis',
            'samping',
            'some',
            'wsome',
            # 'fomanalysis',
        ]

        clasterfunction = [
            # ['Last2First', Get_sus().Last2First],
            # ['DifferentOperator', Get_sus().DifferentOperator],
            # ['Random', Get_sus().Random],

            # ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
            ['ED.MBFL', Get_sus().ErrorDistribution, [True]],

            ['NS.RANDOM', Get_sus().NS, [0]],
            ['NS.SBFL', Get_sus().NS, [2]],
            # ['NS.MBFL', Get_sus().NS, [1]],

            # ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
            # ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
            ['MutCluster', Get_sus().MutCluster, [True, True]],
            # ['MutCluster.in', Get_sus().MutCluster, [True, False]],

            # ['NS.mbfl^MC', Get_sus().NSpMC],
            # ['NS.mbfl^ED', Get_sus().NSpED],
            # ['MC.mseer^ED', Get_sus().MCpED],
            # ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
        ]

        # 创建ftitle
        ftitle = [
            'Method',
            'Reduction',
            'times',
            'repeat',
            'Statement',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += ['None Accurate', 'Pare Accurate', 'Accurate', ]


        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/yojan-need')
        if not os.path.exists(docpath):
            os.makedirs(docpath)
        maxfaultnum = 5
        repeattimes = 5
        timeslist = [1.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9]

        timeCreat = Timecreat(len(SirTools().version_path()))

        for file_i, [project, faultnum, version] in enumerate(SirTools().version_path()):

            # 数据初始化
            file = '%s-%s-%s' % (project, faultnum, version)
            path = os.path.join(data_dirpath, project, faultnum, version)
            # filepath_undo = os.path.join(docpath, '%s-undo.xlsx' % (file))
            filepath_do = os.path.join(docpath, '%s.xlsx' % (file))
            methodsNum = len(clasterfunction)*len(reduction_list)*len(timeslist)
            timeCreat.newFile(file, methodsNum)

            # 断点续传
            wb = False
            passlist = []
            if self.continuecreat:
                if os.path.exists(filepath_do):
                    df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
                    for i in df.index.values:
                        row_data_read = df.loc[i, df.columns.values].to_dict()
                        if not row_data_read['Statement'] == 'max':
                            continue
                        if not row_data_read['Reduction'] == 'testanalysis':
                            continue
                        passlist.append([
                            row_data_read['Method'],
                            row_data_read['times'],
                            row_data_read['repeat'],
                        ])
                    wb = openpyxl.load_workbook(filepath_do)
                    for ws_name in ['exam-best', 'exam-average', 'exam-worst']:
                        ws = wb[ws_name]
                        ws.nowLine = len(passlist)*len(reduction_list)*3+3

            # 表格初始化
            if not wb:
                wb = openpyxl.Workbook()
                del wb['Sheet']
                for ws_name in ['exam-best', 'exam-average', 'exam-worst']:
                    ws = wb.create_sheet(ws_name)
                    ws.cell(1, 1, file)
                    for j, title in enumerate(ftitle):
                        ws.cell(2, j+1, title)
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                        for k in range(maxfaultnum):
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                            ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+maxfaultnum+k, 'rank%s' % str(k+1))
                    ws.nowLine = 3

            # 数据读取
            try:
                sir_data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
            except Exception as e:
                print(path, 'less', e)
                continue

            for parameter in clasterfunction:
                for times in timeslist:
                    for repeat in range(1, repeattimes+1):
                        method = parameter[0]
                        function = parameter[1]

                        timeCreat.newMethod('%s-%s-%s'%(method, times, repeat))

                        # 断点续传
                        if [method, times, repeat] in passlist:
                            continue

                        reducefunction = list(map(lambda key:eval('Reduction(%s).%s' % (str(times), key)),
                                                  reduction_list))
                        if len(parameter) <= 2:
                            sus_data_json = function(sir_data, times=1, reducefunction=reducefunction)
                        else:
                            sus_data_json = function(sir_data, times=1, seting=parameter[2], reducefunction=reducefunction)

                        for tie in ['best', 'average', 'worst']:
                            ws_name = 'exam-%s' % tie
                            ws = wb[ws_name]
                            for reductionKey in reduction_list:
                                for word in ['max', 'ave', 'frequency']:
                                    sheetdata_list = [
                                        method,
                                        reductionKey,
                                        times,
                                        repeat,
                                        word,
                                    ]
                                    if not sus_data_json:
                                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                            ws.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                        continue
                                    susData = sus_data_json[reductionKey]
                                    sheetdata_list += [
                                        len(susData['Fault_Record']),
                                        susData['totaltime'],
                                        susData['fomnum'],
                                        susData['homnum'],
                                        '',
                                    ]

                                    varietys = susData['variety']
                                    sum_varietys = sum(varietys)
                                    sheetdata_list += list(map(lambda x: 0 if sum_varietys == 0 else x/sum_varietys,
                                                               varietys))

                                    precisions = susData['precision']
                                    sum_precisions = sum(precisions)
                                    sheetdata_list += list(map(lambda x: 0 if sum_precisions == 0 else x/sum_precisions,
                                                               precisions))

                                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                        examlist = []
                                        ranklist = []
                                        for fault_ith in range(maxfaultnum):
                                            if fault_ith > len(susData[word][mbfl_for]['rank_%s' % tie])-1:
                                                examlist.append('')
                                                ranklist.append('')
                                            else:
                                                examlist.append(susData[word][mbfl_for]['exam_%s' % tie][fault_ith])
                                                ranklist.append(susData[word][mbfl_for]['rank_%s' % tie][fault_ith])
                                        sheetdata_list += examlist
                                        sheetdata_list += ranklist
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        ws.cell(ws.nowLine, sheetdata_i+1, sheetdata)
                                    ws.nowLine += 1
                        timeCreat.doneMethod()

                    # 存储文件
                    wb.save(filepath_do)
                    print('文件保存 %s ' % filepath_do)
                    # return

            # os.remove(filepath_undo)
            wb.save(filepath_do)
            print('文件保存 %s ' % filepath_do)
        return


    # 不需要倍数，需要重复次数
    def reWriteDif(self):
        clasterfunction = [
            ['DifferentOperator', Get_sus().DifferentOperator],
        ]

        ftitle = [
            'Method',
            'times',
            'repeat',
            'Statement',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]

        for i in range(15):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 5
        repeattimes = 5

        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile-sir/K')
        if not os.path.exists(docpath):
            os.makedirs(docpath)

        for file_i, file in enumerate(os.listdir(docpath)):
            project, faultnum, version = file.split('.')[0].split('-')

            print("{},    {}/{}, {}-{}-{}".format(
                datetime.datetime.now(),
                file_i,
                len(os.listdir(docpath)),
                project,
                faultnum,
                version,)
            )

            path = os.path.join(data_dirpath, project, faultnum, version)
            filepath_do = os.path.join(docpath, file)

            # 数据检测
            df = pd.read_excel(filepath_do, sheet_name='exam-best', header=[1])
            reCreat = []
            for i in df.index.values:
                row_data_read = df.loc[i, df.columns.values].to_dict()
                if not 'DifferentOperator' == row_data_read['Method']:
                    continue
                if not row_data_read['Homnum'] == 0:
                    continue
                reCreat.append(i)
            if len(reCreat) == 0:
                continue
            print(filepath_do, 'diferror')
            continue

            wb = openpyxl.load_workbook(filepath_do)
            # a = wb['exam-best'].cell(1, 1)

            # 数据读取
            try:
                sir_data = SirTools().sir_rules(path, covread=True, firstread=True, oneline=True)
            except Exception as e:
                print(path, 'less', e)
                continue

            times = 1
            parameter = ['DifferentOperator', Get_sus().DifferentOperator]
            for repeat in range(1, repeattimes+1):
                method = parameter[0]
                function = parameter[1]

                print("{},    {}/{}-{}".format(
                    datetime.datetime.now(),
                    file_i,
                    len(os.listdir(docpath)),
                    repeat,
                ))

                if len(parameter) <= 2:
                    sus_data_json = function(sir_data, times=times)
                else:
                    sus_data_json = function(sir_data, times=times, seting=parameter[2])

                for tie in ['best', 'average', 'worst']:
                    ws_name = 'exam-%s' % tie
                    ws = wb[ws_name]
                    for numi, word in enumerate(['max', 'ave', 'frequency']):
                        line = reCreat[(repeat-1)*3 + numi]+3
                        sheetdata_list = [
                            method,
                            times,
                            repeat,
                            word,
                        ]
                        if not sus_data_json:
                            for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                ws.cell(file_i + 3, sheetdata_i+1, sheetdata)
                            continue
                        sheetdata_list += [
                            len(sus_data_json['Fault_Record']),
                            sus_data_json['totaltime'],
                            sus_data_json['fomnum'],
                            sus_data_json['homnum'],
                            '',
                        ]

                        varietys = sus_data_json['variety']
                        sum_varietys = sum(varietys)
                        sheetdata_list += list(map(lambda x: 0 if sum_varietys == 0 else x/sum_varietys,
                                                   varietys))

                        precisions = sus_data_json['precision']
                        sum_precisions = sum(precisions)
                        sheetdata_list += list(map(lambda x: 0 if sum_precisions == 0 else x/sum_precisions,
                                                   precisions))

                        for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                            examlist = []
                            ranklist = []
                            for fault_ith in range(maxfaultnum):
                                if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % tie])-1:
                                    examlist.append('')
                                    ranklist.append('')
                                else:
                                    examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % tie][fault_ith])
                                    ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % tie][fault_ith])
                            sheetdata_list += examlist
                            sheetdata_list += ranklist
                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                            ws.cell(line, sheetdata_i+1, sheetdata)

            wb.save(filepath_do)
            print('文件保存 %s ' % filepath_do)
        return



if __name__ == '__main__':
    Sirsus()

