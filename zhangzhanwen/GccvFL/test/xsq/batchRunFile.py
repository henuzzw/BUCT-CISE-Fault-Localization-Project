import os
import subprocess
from time import sleep

import common
import batchRunFileSetting
import pandas as pd
import openpyxl
from pyecharts import options as opts
from pyecharts.charts import Line, Page
# from tqdm import tqdm
# 这个程序是个脚本，用来批量运行Grace，要在项目的根目录处执行

import json
import os
import pickle
import subprocess
import common
import batchRunFileSetting
from tqdm import tqdm

# 导入必要的库

# 这个程序是个脚本，用来批量处理Grace的运行结果的，要在项目的根目录处执行

# 所有故障定位方法结果的集合
programFL = {}
questionList = []# ["1","2","3","4","5","6","7","8","9","10"]


# init函数中要用到的，这个是读取question下tag所有的信息 用字典存起来
def errorLineNumber(path='', tag={}):
    fileList = common.dirList(path)
    # print(fileList)

    for i, item in enumerate(fileList):
        if len(item) == 0:
            return
        with open(common.jointPath([path, item]), 'r') as f:
            lines = f.readlines()
        tag[lines[0].rstrip('\n')] = [int(x) for x in lines[2].rstrip('\n').split(',')]
    return


# 初始化函数，主要是为了收集数据集中的错误行信息，有个问题需要注意一下，定位到的错误行和数据集中的错误行是否是对应的
def init():
    quesList = common.dirList(batchRunFileSetting.dataDir)
    # programFL = {}
    mydataset = ["2812","2825","2830","2866","1927","2864"]

    for ques in quesList:  # 对于每个question
        if ques not in mydataset:
            continue
        quesPath = common.jointPath([batchRunFileSetting.dataDir, ques, 'Tag_c'])
        questionList.append(ques)
        errorLineNumber(quesPath, programFL)


# 以每个代码文件为单位，收集top指标和EXAM
def forProgram(topInfo, errLine, sumLineNumber):
    # print(topInfo,errLine,sumLineNumber)
    topN10 = [0 for i in range(13)]
    ansTopn = 0
    EXAM = 0
    EXAMArr=[]
    sumLineNumber = 0
    for index, topn in enumerate(topInfo):
        sumLineNumber += len(topn)


    for index, topn in enumerate(topInfo):  # 对于怀疑排序，进行预处理前缀和
        topnLen = len(topn)
        ansTopn += topnLen  # 采用最差topn，计算个数

        # if ansTopn > 10:  # 只统计top前十
        #     break

        for index, topLine in enumerate(topn):  # 对于当前top的行号，如果是错误行，就定位了topIndex中
            if topLine in errLine:
                # print(ansTopn)
                if ansTopn <=10:
                    topN10[ansTopn] += 1
                EXAM += ansTopn / sumLineNumber  # 计算EXAM，稍后会除以总数，取平均
                EXAMArr.append(ansTopn / sumLineNumber)
    for i in range(1, 13):
        topN10[i] += topN10[i - 1]  # 计算topN个数
    return topN10, topN10[1], topN10[3], topN10[5], topN10[10], EXAM , EXAMArr
# 以question为单位，收集top指标、EXAM等信息，需要用到forProgram函数，
allFileEXAM = []
def forQuestion(sheet, question):
    # 保留原始信息，文件名、怀疑排名、怀疑都
    # print(sheet)
    # for i,row in enumerate(sheet.rows):
    #     for j,data in enumerate(row):
    #         print(i,j,data.value,end="  ")
    #     print()

    programNameList = sheet['name']
    # print(sheet['name'])
    topKNList = sheet['suspicion_rank']
    suspicionList = sheet['suspicion']
    question['programNameList'] = programNameList

    question['suspicionList'] = suspicionList  # 这好像是个二维数组 【程序名】【怀疑分数】
    question['topNList'] = topKNList  # 这个好像是三维数组【程序】【topn数组】【行号】
    # print("question = ",question)

    sumTopArr = [0 for i in range(13)]
    # question["EXAMARR"]
    top1 = []
    top3 = []
    top5 = []
    top10 = []
    EXAM = []
    sumTop1 = 0
    sumTop3 = 0
    sumTop5 = 0
    sumTop10 = 0
    sumEXAM = 0
    errLineNum = 0
    sumEXAM = 0
    FileEXAM = []

    # print(programNameList)
    for index, programName in enumerate(programNameList):
        if '1741137' in programName:
            continue
        if '270283' in programName:
            continue
        # question['topKN'] = [0 for i in range(12)]
        question[programName] = {}
        question[programName]['trueAns'] = programFL[programName]

        question[programName]['suspicionRank'] = topKNList[index]
        question[programName]['1'] = {}
        question[programName]['3'] = {}
        question[programName]['5'] = {}
        question[programName]['10'] = {}
        question[programName]['EXAM'] = {}
        # print(suspicionList[index],len(suspicionList[index]))
        question[programName]['topNList'], question[programName]['1'], question[programName]['3'], \
        question[programName]['5'], question[programName]['10'], question[programName][
            'EXAM'],question[programName][
            'EXAMArr'] = forProgram(topKNList[index], programFL[programName], len(suspicionList[index]))

        # 统计question下的所有指标
        top1.append(question[programName]['1'])  # 收集所有 top1
        top3.append(question[programName]['3'])
        top5.append(question[programName]['5'])
        top10.append(question[programName]['10'])
        EXAM.extend(question[programName]['EXAMArr'])
        sumTop1 += question[programName]['1']
        sumTop3 += question[programName]['3']
        sumTop5 += question[programName]['5']
        sumTop10 += question[programName]['10']
        sumEXAM += question[programName]['EXAM']
        if 0 == len(question[programName]['EXAMArr']):
            FileEXAM.append(1)
        else:
            FileEXAM.append(question[programName]['EXAM']/len(question[programName]['EXAMArr']))

        errLineNum += len(programFL[programName])

        # print('(question[programName][topKNList] = ',(question[programName]['topNList']))
        # print('sumTopArr = ',sumTopArr)
        for i in range(len(question[programName]['topNList'])):
            sumTopArr[i] += question[programName]['topNList'][i]
        # sumEXAM += question[programName]['EXAM']
    question['top1'] = top1
    question['top3'] = top3
    question['top5'] = top5
    question['top10'] = top10
    question['sumTop1'] = sumTop1
    question['sumTop3'] = sumTop3
    question['sumTop5'] = sumTop5
    question['sumTop10'] = sumTop10
    question['sumEXAM'] = sumEXAM
    question['FileEXAM'] = FileEXAM
    question['EXAM'] = EXAM
    question['errLineNum'] = errLineNum
    allFileEXAM.extend(FileEXAM)


# 用于生成可视化数据，也就是echarts画个折线图
def drawLine(end, name="折线统计图", xName="", xLine=[1, 2, 3]):
    line_chart = Line(init_opts=opts.InitOpts(width="100%"))

    # 添加数据
    line_chart.add_xaxis(xaxis_data=xLine)
    line_chart.add_yaxis(series_name="top1", y_axis=end['top1'])
    line_chart.add_yaxis(series_name="top3", y_axis=end['top3'])
    line_chart.add_yaxis(series_name="top5", y_axis=end['top5'])
    line_chart.add_yaxis(series_name="top10", y_axis=end['top10'])
    # line_chart2 = Line()
    # line_chart2.add_xaxis(xaxis_data=end['amount']['questName'])
    # line_chart.add_yaxis(series_name="EXAM", y_axis=end['amount']['EXAM'])

    # 设置全局配置
    line_chart.set_global_opts(
        title_opts=opts.TitleOpts(title=name),
        xaxis_opts=opts.AxisOpts(name="xName"),
        yaxis_opts=opts.AxisOpts(name="topN"),
        tooltip_opts=opts.TooltipOpts(trigger="axis"),
    )

    # 渲染图表并保存到文件
    line_chart.render("A"+name + "line_chart.html")


def forBaseLine(myxlsx, baseLine, baseLineName):
    sumTopArr = [0 for i in range(13)]
    top1 = [0 for i in range(len(questionList))]
    top3 = [0 for i in range(len(questionList))]
    top5 = [0 for i in range(len(questionList))]
    top10 = [0 for i in range(len(questionList))]
    EXAM = [ ]
    sumTop1 = 0
    sumTop3 = 0
    sumTop5 = 0
    sumTop10 = 0
    errLineNum = 0
    sumEXAM = 0
    FileEXAM = []
    # 🤔 我们需要整合一下，对于每个方法，我们用一个数组，来存储top1，3，5，10，exam。我们用一个变量sumtop1，3，5，10，sumExam

    for i, question in enumerate(questionList):  # myxlsx中，有若干个question，开始收集每个questtion的信息


        # print(question,myxlsx.keys())
        if question not in list(myxlsx.keys()):
            continue
        baseLine[question] = {}
        forQuestion(myxlsx[question], baseLine[question])
        top1[i] += baseLine[question]['sumTop1']
        top3[i] += baseLine[question]['sumTop3']
        top5[i] += baseLine[question]['sumTop5']
        top10[i] += baseLine[question]['sumTop10']
        sumTop1 += baseLine[question]['sumTop1']
        sumTop3 += baseLine[question]['sumTop3']
        sumTop5 += baseLine[question]['sumTop5']
        sumTop10 += baseLine[question]['sumTop10']
        errLineNum += baseLine[question]['errLineNum']
        sumEXAM += baseLine[question]['sumEXAM']
        EXAM.append(baseLine[question]['sumEXAM'])
        FileEXAM.extend(baseLine[question]['FileEXAM'])
    baseLine['questName'] = questionList
    baseLine['top1'] = top1
    baseLine['top3'] = top3
    baseLine['top5'] = top5
    baseLine['top10'] = top10
    baseLine['sumTop1'] = sumTop1
    baseLine['sumTop3'] = sumTop3
    baseLine['sumTop5'] = sumTop5
    baseLine['sumTop10'] = sumTop10
    baseLine['sumEXAM'] = sumEXAM
    baseLine['EXAM'] = EXAM
    baseLine['errLineNum'] = errLineNum
    baseLine['FileEXAM'] = FileEXAM
    # print('drawLine', baseLineName)
    drawLine(baseLine, baseLineName, "baseLine", questionList)


# 转换成字典
def toDictionary(myxlsx):
    res = {}
    print(myxlsx.sheetnames)
    for ques in questionList:

        print(ques,(ques not in myxlsx.sheetnames))
        if ques not in myxlsx.sheetnames:
            continue
        res[ques] = {}
        table = myxlsx[ques]
        colName = ['name', 'suspicion_rank', 'suspicion']
        for i, row in enumerate(table.rows):
            # print("row = ", row)
            for j, data in enumerate(row):
                # print(data.value)
                if j >=3 :
                    continue
                if i == 0:
                    res[ques][colName[j]] = []
                    # colName[j] = data.value
                elif j == 0:
                    res[ques][colName[j]].append(data.value)
                else:
                    res[ques][colName[j]].append(eval(data.value))

    # print("res = ",res)
    # sleep(20)
    return res


def cate(path, cate):
    baseLineFiles = common.dirList(path)  # 获取SBFL目录下所有的文件名，一个文件名意味着一种baseline
    sumTopArr = [0 for i in range(13)]
    top1 = [0 for i in range(len(baseLineFiles))]
    top3 = [0 for i in range(len(baseLineFiles))]
    top5 = [0 for i in range(len(baseLineFiles))]
    top10 = [0 for i in range(len(baseLineFiles))]
    EXAM = [0 for i in range(len(baseLineFiles))]
    sumTop1 = 0
    sumTop3 = 0
    sumTop5 = 0
    sumTop10 = 0
    errLineNum = 0
    sumEXAM = 0
    # 针对每个baseLine
    for i, baseLineFile in enumerate(baseLineFiles):

        baseLinePath = common.jointPath([path, baseLineFile])  # 拼接baseline的路径
        print("OPen  " ,baseLinePath,path,baseLineFiles)
        myxlsx = openpyxl.load_workbook(baseLinePath)
        myxlsx = toDictionary(myxlsx)

        cate[baseLineFile] = {}
        forBaseLine(myxlsx, cate[baseLineFile], baseLineFile)  # 针对每个baseLine，获取结果
        top1[i] += cate[baseLineFile]['sumTop1']
        top3[i] += cate[baseLineFile]['sumTop3']
        top5[i] += cate[baseLineFile]['sumTop5']
        top10[i] += cate[baseLineFile]['sumTop10']
        sumTop1 += cate[baseLineFile]['sumTop1']
        sumTop3 += cate[baseLineFile]['sumTop3']
        sumTop5 += cate[baseLineFile]['sumTop5']
        sumTop10 += cate[baseLineFile]['sumTop10']
        errLineNum += cate[baseLineFile]['errLineNum']
        sumEXAM += cate[baseLineFile]['sumEXAM']
        EXAM.append(cate[baseLineFile]['sumEXAM'])
        # 将每个方法的结果写进文件中
        with open(batchRunFileSetting.resultDir + "/" + baseLineFile + 'result.json', 'w') as f:
            json.dump(cate[baseLineFile], f)
        # print(batchRunFileSetting.resultDir +cateDir+'result.json')
        # sleep(5)

    cate['top1'] = top1
    cate['top3'] = top3
    cate['top5'] = top5
    cate['top10'] = top10
    cate['sumTop1'] = sumTop1
    cate['sumTop3'] = sumTop3
    cate['sumTop5'] = sumTop5
    cate['sumTop10'] = sumTop10
    cate['sumEXAM'] = sumEXAM
    cate['EXAM'] = EXAM
    # cate['errLineNum'] = errLineNum
    cate['questName'] = questionList


def start():
    fileDirList = common.dirList(batchRunFileSetting.preFileDir)
    # print(batchRunFileSetting.resultDir, fileDirList)
    getResultEpc10 = {}
    for i, cateDir in enumerate(fileDirList):  # 遍历每个Grace文件夹
        getResultEpc10[cateDir] = {}
        cate(common.jointPath([batchRunFileSetting.preFileDir, cateDir]), getResultEpc10[cateDir])  # 处理每一个类别

    return getResultEpc10


# def dealEnd(resEpc10):
#     top1 = []
#     top3 = []
#     top5 = []
#     top10 = []
#     EXAM = []
#     sumTop1 = 0
#     sumTop3 = 0
#     sumTop5 = 0
#     sumTop10 = 0
#
#     questName = []
#     sumEXAM = 0.0
#     ansNum = 0
#     for quest in resEpc10:
#         questName.append(quest)
#         topAnsLen = 0
#         for resIndex in resEpc10[quest]:
#             if str(resIndex) == "-1":  # 索引为-1表示里面是一些统计指标而不是版本
#                 continue
#             if topAnsLen < len(resEpc10[quest][resIndex]['top']):  # 找到最长的那个来统计top指标，防止溢出
#                 topAnsLen = len(resEpc10[quest][resIndex]['top'])
#         topN = [0 for i in range((topAnsLen) + 3)]
#         Exam = 0.0  # 当前问题下的所有的Exam求和，为总Exam做铺垫
#
#         AnsNum = 0  # 当前quest下错误行总数
#         for resIndex in resEpc10[quest]:  # 对于每个问题下的每个版本
#             if str(resIndex) == "-1":  # 索引为-1表示里面是一些统计指标而不是版本
#                 continue
#             for i, val in enumerate(resEpc10[quest][resIndex]['top']):  # 对于每个版本下的top，进行累计
#                 topN[i] += val  # topN i  表示topi 的指标  用于统计当前quest下的所有top指标
#
#             Exam += resEpc10[quest][resIndex]['sumRank']
#             AnsNum += len(resEpc10[quest][resIndex]['ans'])
#
#         sumEXAM += Exam
#         resEpc10[quest]['top'] = topN
#         resEpc10[quest]["EXAM"] = Exam / (AnsNum)  # 错误语句的总分数 / 总的错误语句数量
#         # resEpc10[quest][]
#         resEpc10[quest]['ansNum'] = AnsNum  # 总错误语句数量
#         top1.append(topN[0])
#         top3.append(topN[2])
#         top5.append(topN[4])
#         top10.append(topN[9])
#         EXAM.append(Exam / AnsNum)
#         sumTop1 += topN[0]
#         sumTop3 += topN[2]
#         sumTop5 += topN[4]
#         sumTop10 += topN[9]
#         ansNum += AnsNum
#
#     resEpc10['amount'] = {
#         "top1": top1,
#         "top3": top3,
#         "top5": top5,
#         "top10": top10,
#         "EXAM": EXAM,
#         "sumTop1": sumTop1,
#         "sumTop3": sumTop3,
#         "sumTop5": sumTop5,
#         "sumTop10": sumTop10,
#         "questName": questName,
#         "sumEXAM": sumEXAM / ansNum,
#         "ansNum": ansNum
#     }
#     # print(ansNum)
#     return resEpc10


if __name__ == "__main__":
    init()
    p = subprocess.Popen("mkdir  " + batchRunFileSetting.resultDir, shell=True)

    # print(programFL)
    end = start()
    # end = dealEnd(end)
    # print(len(end))
    with open(batchRunFileSetting.resultDir + '/result.json', 'w') as f:
        json.dump(end, f)

    subprocess.Popen("mv VFL*   " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv SBFL*  " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv Vsus*  " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv VSB*  " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv AVFL*   " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv ASBFL*  " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv AVsus*  " + batchRunFileSetting.resultDir, shell=True)
    subprocess.Popen("mv AVSB*  " + batchRunFileSetting.resultDir, shell=True)
