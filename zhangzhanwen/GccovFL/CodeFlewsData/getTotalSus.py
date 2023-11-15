import openpyxl
import os
import datetime

from MethodSus import Get_sus
from MethodSus import for_list
from MethodSus import Tools
import TestYojian

data_dirpath = './report/CHMBFL/mutinfo-new'
save_dirpath = 'report/CHMBFL/susfile-single' if 'single' in data_dirpath else 'report/CHMBFL/susfile'


class Wbcreat:
    def __init__(self, flName, maxfaultnum):
        self.maxfaultnum = maxfaultnum
        self.method = flName
        # 表格初始化
        ftitle = ['Id', 'Version', 'Fault_Record', 'Fomnum', 'Homnum', 'Formula', 'TotalTest', 'RealTest', ]
        for k in range(self.maxfaultnum):
            ftitle.append('exam%s' % str(k+1))
        for k in range(self.maxfaultnum):
            ftitle.append('rank%s' % str(k+1))
        self.WbsAble = {
            'max': False,
            'ave': False,
            'frequency': False,
            'none': False,
        }
        self.Wbs = {
            'max': openpyxl.Workbook(),
            'ave': openpyxl.Workbook(),
            'frequency': openpyxl.Workbook(),
            'none': openpyxl.Workbook(),
        }
        for wbName, wb in self.Wbs.items():
            ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
            del wb['Sheet']
            for i in range(3):
                for j, title in enumerate(ftitle):
                    ws[i].cell(1, j+1, title)
                ws[i].nowline = 2

    def add_data(self, susDataJson, filename, fileid):
        for susType, wb in self.Wbs.items():
            # word = susType
            wsList = {
                'best': wb['exam-best'],
                'average': wb['exam-average'],
                'worst': wb['exam-worst'],
            }
            for tieType, ws in wsList.items():
                sheetdataList = [
                    fileid,    # str(file_i+1),
                    filename,    # file[5:-5],
                    len(susDataJson['Fault_Record']),
                    susDataJson['fomnum'],
                    susDataJson['homnum'],
                ]
                for formulaIndex, formulaItem in enumerate(for_list):
                    if (susType, formulaItem) in susDataJson:
                        resdata = susDataJson[(susType, formulaItem)]
                    elif ('', formulaItem) in susDataJson:
                        resdata = susDataJson[('', formulaItem)]
                    else:
                        continue
                    formuladataList = [formulaItem]
                    if formulaItem in susDataJson['Desrcibe']:
                        formuladataList.append(susDataJson['Desrcibe'][formulaItem]['total'])
                        formuladataList.append(susDataJson['Desrcibe'][formulaItem]['real'])
                    else:
                        formuladataList.append(0)
                        formuladataList.append(0)

                    self.WbsAble[susType] = True
                    examlist = ['' for i in range(self.maxfaultnum)]
                    ranklist = ['' for i in range(self.maxfaultnum)]
                    for faultIndex in range(len(resdata['rank_best'])):
                        examlist[faultIndex] = resdata['exam_%s' % tieType][faultIndex]
                        ranklist[faultIndex] = resdata['rank_%s' % tieType][faultIndex]
                    formuladataList += examlist
                    formuladataList += ranklist

                    for sheetdataIndex, sheetdataItem in enumerate(sheetdataList + formuladataList):
                        ws.cell(ws.nowline, sheetdataIndex+1, str(sheetdataItem))
                    ws.nowline += 1

    def save_file(self, relativePath, adding=''):
        # for wb_i, wb in enumerate(Wbs):
        for susType, wb in self.Wbs.items():
            if not self.WbsAble[susType]:
                continue
            docpath = os.path.join(os.getcwd(), save_dirpath, relativePath, self.method, susType)
            if not os.path.exists(docpath):
                os.makedirs(docpath)
            filepath = os.path.join(docpath, '%s_%s_%s.xlsx' % (self.method, susType, adding))
            wb.save(filepath)
            print('文件保存 %s ' % filepath)


class Research:
    def __init__(self):
        self.timesable_repeatable()
        self.timesable_repeatless()
        self.timesless_repeatless()
        self.timesless_repeatable()

    def timesable_repeatable(self):
        yojianFucntion = [
            ['CTCR', TestYojian.CTCR_yojian],
            ['IETCR', TestYojian.IETCR_yojian],
            ['RANDOM', TestYojian.RANDOM_yojian],
        ]
        flFunction = [
            ['DNMBFL', Get_sus().DeltaNsMbfl],
        ]

        maxfaultnum = 3
        timeslist = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
        repeatlist = range(5)

        for times in timeslist:
            strtimes = str(int(times*10%10))
            for repeat in repeatlist:
                for flName, flFun in flFunction:
                    for yojianName, yojianFun in yojianFucntion:
                        parm = {
                            'yojian': yojianFun,
                            'yojiantimes': times,
                            'times': 1,
                        }
                        # 表格初始化
                        Wbs = Wbcreat(flName, maxfaultnum)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):
                            print('%s baseline:%s, yojina:%s, times:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), flName, yojianName, times, repeat, file, file_i))
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            sus_data_json = flFun(Tools().json_rules(path), parm)
                            Wbs.add_data(sus_data_json, file[5:-5], str(file_i+1))

                        # 存储文件
                        Wbs.save_file('Yojian', '%s_%s_%sth' % (yojianName, strtimes, repeat))
        return

    def timesable_repeatless(self):
        yojianFucntion = [
            ['CTCR', TestYojian.CTCR_yojian],
            ['IETCR', TestYojian.IETCR_yojian],
            ['RANDOM', TestYojian.RANDOM_yojian],
        ]
        flFunction = [
            ['MBFL', Get_sus().Mbfl],
            ['MCBFL', Get_sus().Mcbfl],
        ]

        maxfaultnum = 3
        timeslist = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
        repeatlist = range(1)

        for times in timeslist:
            strtimes = str(int(times*10%10))
            for repeat in repeatlist:
                for flName, flFun in flFunction:
                    for yojianName, yojianFun in yojianFucntion:
                        parm = {
                            'yojian': yojianFun,
                            'yojiantimes': times,
                            'times': 1,
                        }
                        # 表格初始化
                        Wbs = Wbcreat(flName, maxfaultnum)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):
                            print('%s baseline:%s, yojina:%s, times:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), flName, yojianName, times, repeat, file, file_i))
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            sus_data_json = flFun(Tools().json_rules(path), parm)
                            Wbs.add_data(sus_data_json, file[5:-5], str(file_i+1))

                        # 存储文件
                        Wbs.save_file('Yojian', '%s_%s_%sth' % (yojianName, strtimes, repeat))
        return

    def timesless_repeatless(self):
        yojianFucntion = [
            ['FRMES', TestYojian.FRMES_yojian],
        ]
        flFunction = [
            ['MBFL', Get_sus().Mbfl],
            ['MCBFL', Get_sus().Mcbfl],
        ]

        maxfaultnum = 3
        timeslist = [1.0]
        repeatlist = range(1)

        for times in timeslist:
            strtimes = str(int(times*10%10))
            for repeat in repeatlist:
                for flName, flFun in flFunction:
                    for yojianName, yojianFun in yojianFucntion:
                        parm = {
                            'yojian': yojianFun,
                            'yojiantimes': times,
                            'times': 1,
                        }
                        # 表格初始化
                        Wbs = Wbcreat(flName, maxfaultnum)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):
                            print('%s baseline:%s, yojina:%s, times:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), flName, yojianName, times, repeat, file, file_i))
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            sus_data_json = flFun(Tools().json_rules(path), parm)
                            Wbs.add_data(sus_data_json, file[5:-5], str(file_i+1))

                        # 存储文件
                        Wbs.save_file('Yojian', '%s_%s_%sth' % (yojianName, strtimes, repeat))
        return

    def timesless_repeatable(self):
        yojianFucntion = [
            ['FRMES', TestYojian.FRMES_yojian],
        ]
        flFunction = [
            ['DNMBFL', Get_sus().DeltaNsMbfl],
        ]

        maxfaultnum = 3
        timeslist = [1.0]
        repeatlist = range(5)

        for times in timeslist:
            strtimes = str(int(times*10%10))
            for repeat in repeatlist:
                for flName, flFun in flFunction:
                    for yojianName, yojianFun in yojianFucntion:
                        parm = {
                            'yojian': yojianFun,
                            'yojiantimes': times,
                            'times': 1,
                        }
                        # 表格初始化
                        Wbs = Wbcreat(flName, maxfaultnum)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):
                            print('%s baseline:%s, yojina:%s, times:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), flName, yojianName, times, repeat, file, file_i))
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            sus_data_json = flFun(Tools().json_rules(path), parm)
                            Wbs.add_data(sus_data_json, file[5:-5], str(file_i+1))

                        # 存储文件
                        Wbs.save_file('Yojian', '%s_%s_%sth' % (yojianName, strtimes, repeat))

        return


if __name__ == '__main__':
    Research()