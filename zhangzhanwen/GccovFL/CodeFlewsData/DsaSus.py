import datetime
import json
# import ijson
import os
import math
import numpy as np
import socket

import mbfl.mbfl_for
import mbfl.command
import sbfl.sbfl_for
import sbfl.command
import util
import openpyxl
import random
# from CHMBFL_Flow import Cluster_Fom
from mbfl.mutpolyn import mutation_trick

mbfl_for_list = [
    "mbfl.mbfl_for.Tarantula",
    "mbfl.mbfl_for.Op2",
    "mbfl.mbfl_for.Jaccard",
    "mbfl.mbfl_for.Ochiai",
    "mbfl.mbfl_for.Dstar",
    "mbfl.mbfl_for.GP13",
    "mbfl.mbfl_for.Naish1",
    "mbfl.mbfl_for.Barinel",
]

sbfl_for_list = [
    "sbfl.sbfl_for.Tarantula",
    "sbfl.sbfl_for.Op2",
    "sbfl.sbfl_for.Jaccard",
    "sbfl.sbfl_for.Ochiai",
    "sbfl.sbfl_for.Dstar",
    "sbfl.sbfl_for.GP13",
    "sbfl.sbfl_for.Naish1",
    "sbfl.sbfl_for.Barinel",
]



clasterfunction = ['KMeans', 'AgglomerativeClustering']


class Tools:
    def json_rules(self, path):
        # 读取小于3g的文件
        # if os.stat(path).st_size > 3*1024*1024*1024:
        #     print('内存过大')
        #     return False

        try:
            with open(path) as f_obj:
                data_json = json.load(f_obj)
        except:
            print('读取异常 %s' % path)
            return False
        return data_json


    def get_host_ip(self):
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM)
        s.connect(('8.8.8.8',80))
        ip=s.getsockname()[0]
        print(ip)
        # try:
        #     pass
        # finally:
        #     s.close()
        #     return
        return ip

class Get_sus:
    # 计算last2first方法怀疑度
    def get_sus_last2first(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        hom_list_simple = list(map(lambda x: x['message'], hom_list_all))

        # ------------------------------------------------
        # 生成som组合
        out_dic = dict()
        kill_dic = dict()
        random_fomloc = random.sample(range(len(fom_list)), len(fom_list))
        while len(random_fomloc) > 1:
            locfirst = random_fomloc.pop(0)
            fom_first = fom_list[locfirst]
            for i in range(len(random_fomloc)-1, -1, -1):
                # fom_last = fom_list[i]
                loclast = random_fomloc[i]
                fom_last = fom_list[loclast]
                if fom_first['message'][0][0] == fom_last['message'][0][0]:
                    continue
                som = sorted([fom_first['message'][0], fom_last['message'][0]], key=lambda x: x[0])
                try:
                    loc = hom_list_simple.index(som)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                random_fomloc.pop(i)
                break


        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 计算differentOperator方法怀疑度
    def get_sus_differentoperator(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        hom_list_simple = list(map(lambda x: x['message'], hom_list_all))

        def find_operator(linebefore, lineafter):
            for i in range(2):
                for key in mutation_trick[i]:
                    if key in linebefore:
                        for value in mutation_trick[i][key]:
                            if value in lineafter:
                                return key, value

        # ------------------------------------------------
        # 生成som组合
        out_dic = dict()
        kill_dic = dict()
        while len(fom_list) > 1:
            fom_1 = fom_list.pop(0)
            message_1 = fom_1['message'][0]
            operator_1_key = find_operator(message_1[1], message_1[2])

            for fom2_loc in random.sample(range(len(fom_list)), len(fom_list)):
                fom_2 = fom_list[fom2_loc]
                message = fom_2['message'][0]
                if message_1[0] == message[0]:
                    continue
                operator_2_key, operator_2_value = find_operator(message[1], message[2])
                if operator_1_key == operator_2_key or operator_1_key == operator_2_value:
                    continue
                som = sorted([fom_1['message'][0], fom_2['message'][0]], key=lambda x: x[0])
                try:
                    loc = hom_list_simple.index(som)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                fom_list.pop(fom2_loc)
                break


        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 计算randomMix方法怀疑度
    def get_sus_randommix(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        hom_list_simple = list(map(lambda x: x['message'], hom_list_all))

        # ------------------------------------------------
        # 生成som组合
        # print('生成som组合')
        out_dic = dict()
        kill_dic = dict()
        while len(fom_list) > 1:
            loc1 = random.sample(range(len(fom_list)), 1)[0]
            fom_1 = fom_list.pop(loc1)
            message_1 = fom_1['message'][0]
            for loc2 in random.sample(range(len(fom_list)), len(fom_list)):
                fom_2 = fom_list[loc2]
                message_2 = fom_2['message'][0]
                if message_1[0] == message_2[0]:
                    continue
                som = sorted([fom_1['message'][0], fom_2['message'][0]], key=lambda x: x[0])
                try:
                    loc = hom_list_simple.index(som)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                fom_list.pop(loc2)
                break


        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        return sus_data



    # 计算sbfl方法怀疑度
    def get_sus_mcbfl(self, data_json):
        mcbfl_for_list = [
            "Tarantula",
            "Op2",
            "Jaccard",
            "Ochiai",
            "Dstar",
            "GP13",
            "Naish1",
            "Barinel",
        ]
        sbfl_for_path = 'sbfl.sbfl_for.'
        mbfl_for_path = 'mbfl.mbfl_for.'


        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0


        # ------------------------------------------------
        # 获取mbfl元组
        mbfl.mbfl_for.type_mbfl = 'max'
        totaltime = 0
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            totaltime += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # 获取sbfl元组
        sus_list_sbfl = []
        totaltime += sbfl_data['time']
        touple_sbfl = sbfl.command.touple(or_list, sbfl_data['cov'])

        sus_data['totaltime'] = totaltime

        # 计算怀疑度
        for method in mcbfl_for_list:
            sus_data[method] = dict()
            sbfl_for = sbfl_for_path+method
            mbfl_for = mbfl_for_path+method
            sus_dic = dict()

            # 计算mbfl怀疑度
            for line in touple_mbfl['linedata']:
                touple_list = touple_mbfl['tf'], \
                              touple_mbfl['tp'], \
                              touple_mbfl['linedata'][line]['f'], \
                              touple_mbfl['linedata'][line]['p'], \
                              touple_mbfl['f2p'], \
                              touple_mbfl['p2f'], \
                              touple_mbfl['linedata'][line]['kf'], \
                              touple_mbfl['linedata'][line]['kp'], \
                              touple_mbfl['linedata'][line]['nf'], \
                              touple_mbfl['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_dic[line] = sus + -1

            # 计算sbfl怀疑度
            for line, touple_list in touple_sbfl.items():
                sus = eval(sbfl_for)(touple_list)
                sus_list_sbfl.append([line, sus])
                if line in sus_dic:
                    sus_dic[line] += sus + 1
                else:
                    sus_dic[line] = sus + -1

            sus_list = []
            for line, sus in sus_dic.items():
                sus_list.append([line, sus/2])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[method]['rank_best'] = rank[0]
            sus_data[method]['rank_average'] = rank[1]
            sus_data[method]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[method]['exam_best'] = exam[0]
            sus_data[method]['exam_average'] = exam[1]
            sus_data[method]['exam_worst'] = exam[2]

        return sus_data

    # 计算DSA fom对应times倍数som 怀疑度
    # 使用 path中文件的数据
    def get_sus_DSA_random(self, data_json, times):
        if not data_json:
            return False

        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        if len(hom_list_all)/(fomnum*times) < 0.9:
            print('数量不足')
            return False



        out_dic = dict()
        kill_dic = dict()
        for loc in random.sample(range(len(hom_list_all)), min(len(hom_list_all), fomnum*times)):
            hom = hom_list_all[loc]
            message = hom['message']
            out_list = hom['out_list']
            kill_list = hom['kill_list']
            sus_data['totaltime'] += hom['time']
            for fom in message:
                line = fom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(out_list)
                kill_dic[line].append(kill_list)
        del hom_list_all


        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)

                sus_data[word][mbfl_for] = dict()

                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # sbfl指导生成hmbfl
    def get_sus_SBFLguide_hmbfl(self, data_json, times):
        # print('%s 读取数据' % datetime.datetime.now())
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            # nowpath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True


        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0


        # 计算一阶变异体怀疑度
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))




        # ------------------------------------------------
        # 计算sbfl怀疑度
        touple = sbfl.command.touple(or_list, sbfl_data['cov'])
        sus_list_sbfl = []
        for line, touple_list in touple.items():
            sus = sbfl.sbfl_for.GP13(touple_list)
            sus_list_sbfl.append([line, sus, -1, -1])

        # print('%s 获取组队数据' % datetime.datetime.now())

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_sbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_sbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: x['message'], hom_list_all))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        out_dic = dict()
        kill_dic = dict()
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    som_pair_list.append([fom0['message'][0], fom1['message'][0]])
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum >= 1 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                if sum(map(lambda x: getSus(fom_data_list, x), hom['message'])) > 0:
                    continue
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # mbfl指导生成hmbfl
    def get_sus_MBFLguide_hmbfl(self, data_json, times):
        # print('%s 读取数据' % datetime.datetime.now())
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True


        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0

        # 计算一阶变异体怀疑度
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))


        # print('%s 计算一阶怀疑度' % datetime.datetime.now())
        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # print('%s 获取组队数据' % datetime.datetime.now())

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)


        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: x['message'], hom_list_all))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        out_dic = dict()
        kill_dic = dict()
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    som_pair_list.append([fom0['message'][0], fom1['message'][0]])
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum >= 1 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                if sum(map(lambda x: getSus(fom_data_list, x), hom['message'])) > 0:
                    continue
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        # sus_data['totaltime'] += sbfl_data['time']
        return sus_data


    # mbfl指导生成hmbfl
    def get_sus_MBFLguide_hmbfl_old(self, data_json, times):
        # print('%s 读取数据' % datetime.datetime.now())
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fomnum = data['fomnum']
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            # nowpath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        def get_som_out(som_message_self, hom_list_all_self):
            som_dict = {}
            for hom_all in hom_list_all_self:
                if hom_all["message"][0] == som_message_self[0] and hom_all["message"][1] == som_message_self[1]:
                    # if "out_or" not in hom_all:
                    #     continue
                    # if len(hom_all["out_or"]) == 0:
                    #     continue
                    for fom_message in som_message:
                        som_dict[fom_message[0]] = [hom_all["out_list"], hom_all["kill_list"]]
                    return som_dict
            return False

        # print('%s 计算一阶怀疑度' % datetime.datetime.now())
        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
        sus_list = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list.append([line, sus, -1, -1])

        # print('%s 获取组队数据' % datetime.datetime.now())

        # 获取组队数据
        line_pairs_list = Tools().guide_line_pair(sus_list, fomnum*times)
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def getsimplehomlist(hom_list_all):
            hom_list_simple = []
            for hom in hom_list_all:
                hom_list_simple.append(hom['message'])

            return hom_list_simple
        hom_list_simple = getsimplehomlist(hom_list_all)

        # print('%s 获取高阶元组' % datetime.datetime.now())

        # 获取元组输入字典

        sfterload_hom = []

        out_dic = dict()
        kill_dic = dict()
        som_num = 0
        # print('%s len %s ' % (datetime.datetime.now(), len(line_pairs_list)))
        for pair in line_pairs_list:
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue

            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    som_pair_list.append([fom0['message'][0], fom1['message'][0]])

            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                som_num += 1
                break

            if len(random_som_pair_list) > 0:
                sfterload_hom += random_som_pair_list


            # for som_message in random.sample(som_pair_list, len(som_pair_list)):
            #     try:
            #         loc = hom_list_simple.index(som_message)
            #     except:
            #         continue
            #     hom = hom_list_all[loc]
            #     sus_data['totaltime'] += hom['time']
            #     for fom_in_hom in hom['message']:
            #         line = fom_in_hom[0]
            #         if line not in out_dic:
            #             out_dic[line] = []
            #             kill_dic[line] = []
            #         out_dic[line].append(hom['out_list'])
            #         kill_dic[line].append(hom['kill_list'])
            #     som_num += 1
            #     break
        # print('%s 补充高阶元组 %s- %s' % (datetime.datetime.now(), som_num , len(line_pairs_list)))
        while som_num < len(line_pairs_list) and len(sfterload_hom) > 0:
            som_message = sfterload_hom.pop(0)
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_list_all[loc]
            sus_data['totaltime'] += hom['time']
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])
            som_num += 1

        if som_num < len(line_pairs_list):
            print('变异体数量不足')
            return False

        # print('%s 计算高阶怀疑度' % datetime.datetime.now())

        for word in ['max', 'ave', 'frequency']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        sus_data['totaltime'] += sbfl_data['time']
        return sus_data


def addSbflGuideData():
    data_dirpath = './report/DSA/NewData'
    repeat_time = 10

    # for times in range(1, 16):
    times = 1
    for repeat in range(1, repeat_time+1):
        Wblist = []
        Wslist = []
        # print('读取 repeat:%s times:%s'% (repeat, times))
        for word in ['max', 'ave', 'frequency']:
            docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'high_order', 'sbflguide', word, '%s_repeat_result' % repeat)
            filepath = os.path.join(docpath, '%sth_sbflguide_hmbfl_ranke.xlsx' % times)
            try:
                wb = openpyxl.load_workbook(filepath)
            except:
                print('读取出错', filepath)
                return
            Wblist.append(wb)
            Wslist.append([wb['exam-best'], wb['exam-average'], wb['exam-worst']])

        for i in range(3, 137):
            lackdata = 0
            for ws1, ws2, ws3 in Wslist:
                if ws1.cell(i, 3).value == None:
                    lackdata += 1
                if ws2.cell(i, 3).value == None:
                    lackdata += 1
                if ws3.cell(i, 3).value == None:
                    lackdata += 1
            if lackdata == 0:
                continue

            file = 'data_%s.json' % Wslist[0][0].cell(i, 2).value
            print('%s 数据缺失 repeat:%s times:%s file:%s lack:%s' % (datetime.datetime.now(), repeat, times, file, lackdata))
            path = os.path.join(os.getcwd(), data_dirpath, file)

            for _ in range(10):
                sus_data_json = Get_sus().get_sus_SBFLguide_hmbfl(Tools().json_rules(path), times)
                if not sus_data_json:
                    continue
                for type, ws in enumerate(Wslist):
                    word = ['max', 'ave', 'frequency'][type]
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        sheetlist = [
                            [ws[0], 'best'],
                            [ws[1], 'average'],
                            [ws[2], 'worst'],
                        ]
                        for sheet, name in sheetlist:
                            sheet.cell(i, 3, len(sus_data_json['Fault_Record']))
                            sheet.cell(i, 4, sus_data_json['totaltime'])
                            for fault_ith in range(len(sus_data_json['Fault_Record'])):
                                rank = sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith]
                                exam = sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith]

                                sheet.cell(i, 5+fault_ith + j*6, exam)
                                sheet.cell(i, 5+fault_ith+3 + j*6, rank)
                print('补充成功')
                break



        for wb, word in [[Wblist[0], 'max'], [Wblist[1], 'ave'], [Wblist[2], 'frequency']]:
            docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'high_order', 'sbflguide', word, '%s_repeat_result' % repeat)
            filepath = os.path.join(docpath, '%sth_sbflguide_hmbfl_ranke.xlsx' % times)
            wb.save(filepath)
        # print('存储完成 repeat:%s times:%s'% (repeat, times))
        # Qr_excel().save(filepath, 'exam-best', ws_best)
        # Qr_excel().save(filepath, 'exam-average', ws_average)
        # Qr_excel().save(filepath, 'exam-worst', ws_worst)
        pass

    return


def pastSbflGuideData():
    data_dirpath = './report/DSA/NewData'
    repeat_time = 10

    for times in range(1, 16):
        for word in ['max', 'ave', 'frequency']:
            for sheet in ['exam-best', 'exam-average', 'exam-worst']:
                Wblist = []
                Wslist = []
                for repeat in range(1, repeat_time+1):
                    docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'high_order', 'sbflguide', word, '%s_repeat_result' % repeat)
                    filepath = os.path.join(docpath, '%sth_sbflguide_hmbfl_ranke.xlsx' % times)
                    wb = openpyxl.load_workbook(filepath)
                    Wblist.append([wb, filepath])
                    Wslist.append(wb[sheet])

                for i in range(3, 137):
                    lackdata = False
                    data = []
                    for ws in Wslist:
                        if ws.cell(i, 3).value == None:
                            lackdata = True
                        else:
                            if len(data) == 0:
                                for j in range(3, 53):
                                    if ws.cell(i, j).value == None:
                                        data.append('')
                                    else:
                                        data.append(ws.cell(i, j).value)

                    if lackdata and len(data) > 0:
                        for ws in Wslist:
                            if ws.cell(i, 3).value == None:
                                for j in range(3, 53):
                                    try:
                                        ws.cell(i, j, data[j-3])
                                    except:
                                        continue

                for wb, filepath in Wblist:
                    wb.save(filepath)
                    print(filepath)
                pass

    return


def additionalExperiments():
    version = [
        'v1721', 'v1605', 'v1638', 'v1622', 'v1660',
        'v1664', 'v1722', 'v1556', 'v1737', 'v1615'
    ]
    titlelist1 = [
        '2-HOMs', '2-HOMs_Ochiai_sus',
        'FOM1 stmt. Index', 'FOM1 stmt. Ochiai_sus', 'FOM1 index', 'FOM1_Ochiai_sus',
        'FOM2 stmt. Index', 'FOM2 stmt. Ochiai_sus', 'FOM2 index', 'FOM2_Ochiai_sus',
        'Fault1 stmt. index', 'Fault2 stmt. index', 'Fault3 stmt. index'
    ]
    titlelist2 = [
        '二阶变异体编号', '二阶变异体怀疑度',
        '一阶变异体对应语句行号', '一阶变异体语句怀疑度', '一阶变异体编号', '一阶变异体怀疑度',
        '一阶变异体对应语句行号', '一阶变异体语句怀疑度', '一阶变异体编号', '一阶变异体怀疑度',
        '错误语句行号', '错误语句行号', '错误语句行号',
    ]
    times = 9
    for v in version:
        filepath = './report/DSA/NewData/data_%s.json' % v
        wb = openpyxl.Workbook()
        del wb['Sheet']
        for repeat in range(1, 3):
            # 表格初始化
            ws = wb.create_sheet('repeat%s' % repeat)
            for i, title in enumerate(titlelist1):
                ws.cell(1, i+1, title)
            for i, title in enumerate(titlelist2):
                ws.cell(2, i+1, title)
            som_id = 0

            # 初始化数据
            data_json = Tools().json_rules(filepath)
            if not data_json:
                return False

            try:
                doc = list(data_json.keys())[0]
                data = data_json[doc]
                del data_json

                or_list = data["or_list"]
                src_path = data['path']
                Fault_Record = data['Fault_Record']
                fomnum = data['fomnum']
                fom_list = data['fom_list']
                hom_list_all = data['hom_list_all']
                homnum = data['homnum']
                # or_out = data['or_out']
                # true_out = data['true_out']
                del data

                sus_data = dict()
                sus_data['Fault_Record'] = Fault_Record
            except Exception as e:
                print('数据缺失 %s' % e)
                return False

            def mutTouple(or_list, mut_kill):
                kf, kp, nf, np = 0, 0, 0, 0
                for j in range(len(or_list)):
                    if or_list[j] == 1:
                    # 原始pass
                        if mut_kill[j] == 1:
                            # 变异体n
                            np += 1
                        else:
                            # 变异体k
                            kp += 1
                    else:
                        # 原始fail
                        if mut_kill[j] == 1:
                            # 变异体n
                            kf += 1
                        else:
                            # 变异体k
                            nf += 1
                return [kf, kp, nf, np]

            def mutOchiai(touple):
                akf, akp, anf, anp = touple
                try:
                    return akf/math.sqrt((akf+anf)*(akf+akp))
                except:
                    return -1


            # 计算一阶语句怀疑度
            mbfl.mbfl_for.type_mbfl = 'max'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = fom["out_list"]
                fom_kill_list = fom["kill_list"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
            sus_list = dict()
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list[line] = sus

            # 计算一阶变异体怀疑度
            fom_message_list = []
            fom_data_list = []
            for fom in fom_list:
                fom_message = fom["message"][0]
                line = fom_message[0]
                fom_kill_list = fom["kill_list"]
                sus = mutOchiai(mutTouple(or_list, fom_kill_list))
                fom_message_list.append(fom_message)
                fom_data_list.append([line, sus_list[line], len(fom_message_list), sus])


            # 计算高阶怀疑度
            while som_id < fomnum*times and len(hom_list_all) > 0:
                hom = hom_list_all.pop(random.sample(range(len(hom_list_all)), 1)[0])
                message = hom['message']
                kill_list = hom['kill_list']
                sus = mutOchiai(mutTouple(or_list, kill_list))
                celldata = []
                celldata.append(som_id)
                celldata.append(sus)
                for j, fom in enumerate(message):
                    try:
                        fom_data = fom_data_list[fom_message_list.index(fom)]
                    except:
                        break
                    celldata += fom_data
                if not len(celldata) == 10:
                    continue
                celldata += Fault_Record
                for j in range(len(celldata)):
                    ws.cell(som_id+3, j+1, celldata[j])
                som_id += 1

            if som_id < fomnum*times:
                print('数量不足', v)

        savepath = './report/DSA/susfile/additional/%s.xlsx' % v
        wb.save(savepath)
        print(savepath)
    return




def DSA():

    # DSA高阶变异体 随机生成结果
    # 筛选只用划定的数据集部分
    # 重复max ave frequency三组
    # 生成10份重复文件
    def DSA_highorder_random_res():
        clasterfunction = [
            # ['RandomMix', Get_sus().get_sus_randommix],
            # ['DifferentOperator', Get_sus().get_sus_differentoperator],
            ['Last2First', Get_sus().get_sus_last2first],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'K',
        ]
        maxfaultnum = 3

        repeat_time = 10
        for method, function in clasterfunction:
            for repeat in range(1, repeat_time+1):
                # 表格初始化
                Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                Wss = []
                for wb in Wbs:
                    ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                    del wb['Sheet']
                    for i in range(3):
                        ws[i].cell(1, 1, method)
                        for j, title in enumerate(ftitle):
                            ws[i].cell(2, j+1, title)
                        for j, mbfl_for in enumerate(mbfl_for_list):
                            ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                            for k in range(maxfaultnum):
                                ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                    Wss.append(ws)

                # 获取数据
                data_dirpath = './report/DSA/NewData'
                for file_i, file in enumerate(os.listdir(data_dirpath)):
                    print('%s repeat:%s, baseline:%s read:%s-%s' % (datetime.datetime.now(), repeat, method, file, file_i))

                    # 计算相应方法的怀疑度
                    path = os.path.join(data_dirpath, file)
                    sus_data_json = function(Tools().json_rules(path))


                    for type, ws in enumerate(Wss):
                        word = ['max', 'ave', 'frequency'][type]
                        for j, mbfl_for in enumerate(mbfl_for_list):
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['homnum']/sus_data_json['fomnum'],
                                ]
                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)



                # 存储文件
                for type, wb in enumerate(Wbs):
                    word = ['max', 'ave', 'frequency'][type]
                    docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'high_order', method, '%s_repeat_result' % repeat, word)
                    if not os.path.exists(docpath):
                        os.makedirs(docpath)
                    filepath = os.path.join(docpath, '%s_%sth_%s_ranke.xlsx' % (method, repeat, word))
                    wb.save(filepath)
                    print('文件保存 %s ' % filepath)
        return


    # DSA高阶变异体 随机生成结果
    # 筛选只用划定的数据集部分
    # 重复max ave frequency三组
    # 生成10份重复文件
    def DSA_highorder_guide_res():
        clasterfunction = [
            # ['SBFLguide', Get_sus().get_sus_SBFLguide_hmbfl],
            # ['MBFLguide', Get_sus().get_sus_MBFLguide_hmbfl],
            ['Random', Get_sus().get_sus_DSA_random],
        ]
        repeat_time = 10
        for method, function in clasterfunction:
            for repeat in range(1, repeat_time+1):
                for times in range(1, 16):

                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            ws[i].cell(2, 1, 'id')
                            ws[i].cell(2, 2, 'version')
                            ws[i].cell(2, 3, 'fault_num')
                            ws[i].cell(2, 4, 'time_spend')
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*6+5, mbfl_for.split('.')[2])
                                ws[i].cell(2, j*6+5+0, 'exam1')
                                ws[i].cell(2, j*6+5+1, 'exam2')
                                ws[i].cell(2, j*6+5+2, 'exam3')
                                ws[i].cell(2, j*6+5+3, 'rank1')
                                ws[i].cell(2, j*6+5+4, 'rank2')
                                ws[i].cell(2, j*6+5+5, 'rank3')
                        Wss.append(ws)

                    # 获取数据
                    data_dirpath = './report/DSA/NewData'
                    for i, file in enumerate(os.listdir(data_dirpath)):

                        print('%s %s repeat:%s, times:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, times, file, i))


                        # 计算相应方法的怀疑度
                        path = os.path.join(os.getcwd(), data_dirpath, file)
                        sus_data_json = function(Tools().json_rules(path), times)


                        for type, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency'][type]
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                sheetlist = [
                                    [ws[0], 'best'],
                                    [ws[1], 'average'],
                                    [ws[2], 'worst'],
                                ]
                                for sheet, name in sheetlist:
                                    if not sus_data_json:
                                        sheet.cell(i + 3, 1, str(i+1))
                                        sheet.cell(i + 3, 2, file[5:-5])
                                        continue
                                    else:
                                        sheet.cell(i + 3, 1, str(i+1))
                                        sheet.cell(i + 3, 2, file[5:-5])
                                        sheet.cell(i + 3, 3, len(sus_data_json['Fault_Record']))
                                        sheet.cell(i + 3, 4, sus_data_json['totaltime'])
                                    for fault_ith in range(len(sus_data_json['Fault_Record'])):
                                        rank = sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith]
                                        exam = sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith]

                                        sheet.cell(i + 3, 5+fault_ith + j*6, exam)
                                        sheet.cell(i + 3, 5+fault_ith+3 + j*6, rank)

                    # 存储文件
                    for type, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency'][type]
                        docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'high_order',
                                               method, word, '%s_repeat_result' % repeat)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%sth_%stimes_%s_ranke.xlsx' % (method, repeat, times, word))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return


    # DSA高阶变异体 随机生成结果
    # 筛选只用划定的数据集部分
    # 重复max ave frequency三组
    # 生成10份重复文件
    def DSA_highorder_res():
        clasterfunction = [
            ['Last2First', Get_sus().get_sus_last2first],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'K',
        ]
        maxfaultnum = 3

        for method, function in clasterfunction:
            # 表格初始化
            Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
            Wss = []
            for wb in Wbs:
                ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                del wb['Sheet']
                for i in range(3):
                    ws[i].cell(1, 1, method)
                    for j, title in enumerate(ftitle):
                        ws[i].cell(2, j+1, title)
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                        for k in range(maxfaultnum):
                            ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                            ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                Wss.append(ws)

            # 获取数据
            data_dirpath = './report/DSA/NewData'
            for file_i, file in enumerate(os.listdir(data_dirpath)):
                # if not 'v1648' in file:
                #     continue
                print('%s baseline:%s read:%s-%s' % (datetime.datetime.now(), method, file, file_i))

                # 计算相应方法的怀疑度
                path = os.path.join(os.getcwd(), data_dirpath, file)
                sus_data_json = function(Tools().json_rules(path))

                for type, ws in enumerate(Wss):
                    word = ['max', 'ave', 'frequency'][type]
                    for j, mbfl_for in enumerate(mbfl_for_list):
                        sheetlist = [
                            [ws[0], 'best'],
                            [ws[1], 'average'],
                            [ws[2], 'worst'],
                        ]
                        for sheet, name in sheetlist:
                            sheetdata_list = [
                                str(file_i+1),
                                file[5:-5],
                            ]
                            if not sus_data_json:
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                continue
                            sheetdata_list += [
                                len(sus_data_json['Fault_Record']),
                                sus_data_json['totaltime'],
                                sus_data_json['homnum']/sus_data_json['fomnum'],
                                ]
                            for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                examlist = []
                                ranklist = []
                                for fault_ith in range(maxfaultnum):
                                    if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                        examlist.append('')
                                        ranklist.append('')
                                    else:
                                        examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                        ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                sheetdata_list += examlist
                                sheetdata_list += ranklist
                            for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

            # 存储文件
            for type, wb in enumerate(Wbs):
                word = ['max', 'ave', 'frequency'][type]
                docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'high_order', method, word)
                if not os.path.exists(docpath):
                    os.makedirs(docpath)
                filepath = os.path.join(docpath, '%s_%s_ranke.xlsx' % (method, word))
                wb.save(filepath)
                print('文件保存 %s ' % filepath)
        return


    # DSA sbfl结果
    def DSA_sbfl_res():
        wb_name = 'Sbfl'
        wb = openpyxl.Workbook()
        ws = [[wb.create_sheet('exam-best'), 'best'],
              [wb.create_sheet('exam-average'), 'average'],
              [wb.create_sheet('exam-worst'), 'worst']]
        del wb['Sheet']
        for i in range(3):
            ws[i][0].cell(1, 1, 'hmbfl')
            ws[i][0].cell(2, 1, 'id')
            ws[i][0].cell(2, 2, 'version')
            ws[i][0].cell(2, 3, 'fault_num')
            ws[i][0].cell(2, 4, 'time_spend')
            for j, mbfl_for in enumerate(mbfl_for_list):
                ws[i][0].cell(1, j*6+5, mbfl_for.split('.')[2])
                ws[i][0].cell(2, j*6+5+0, 'exam1')
                ws[i][0].cell(2, j*6+5+1, 'exam2')
                ws[i][0].cell(2, j*6+5+2, 'exam3')
                ws[i][0].cell(2, j*6+5+3, 'rank1')
                ws[i][0].cell(2, j*6+5+4, 'rank2')
                ws[i][0].cell(2, j*6+5+5, 'rank3')

        data_dirpath = './report/DSA/NewData'
        for i, file in enumerate(os.listdir(data_dirpath)):
            # print('\r %s baseline:%s, read:%s' % (datetime.datetime.now(), wb_name, file), end='')
            print('%s baseline:%s, read:%s - %s' % (datetime.datetime.now(), wb_name, file, i))
            if 'random' in file:
                continue
            # if 'v1582' not in file:
            #     continue

            # 计算相应方法的怀疑度
            path = os.path.join(os.getcwd(), data_dirpath, file)
            sus_data_json = Get_sus().get_sus_sbfl(Tools().json_rules(path))

            if sus_data_json == False:
                continue

            for sheet, sheet_name in ws:
                sus_formula_j = 0
                for key, value in sus_data_json.items():
                    if key == 'Fault_Record':
                        sheet.cell(i + 3, 1, str(i+1))
                        sheet.cell(i + 3, 2, file[5:-5])
                        sheet.cell(i + 3, 3, len(value))
                    elif key == 'totaltime':
                        sheet.cell(i + 3, 4, value)
                    else:
                        for fault_ith in range(len(value['exam_best'])):
                            rank = value['rank_%s' % sheet_name][fault_ith]
                            exam = value['exam_%s' % sheet_name][fault_ith]

                            sheet.cell(i + 3, 5+fault_ith + sus_formula_j*6, exam)
                            sheet.cell(i + 3, 5+fault_ith+3 + sus_formula_j*6, rank)
                        sus_formula_j += 1

        docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'baseline')
        if not os.path.exists(docpath):
            os.makedirs(docpath)
        filepath = os.path.join(docpath, '%s_rank.xlsx' % wb_name)
        wb.save(filepath)
        print('文件保存 %s ' % filepath)
        return


    # DSA baseline结果
    # 分别 mbfl 和 mcbfl
    def DSA_mbfl_res():
        baseline_list = [
            ['Mbfl', Get_sus().get_sus_mbfl],
            ['Mcbfl', Get_sus().get_sus_mcbfl],
        ]
        for wb_name, baseline in baseline_list:
            wb = openpyxl.Workbook()
            ws = [[wb.create_sheet('exam-best'), 'best'],
                  [wb.create_sheet('exam-average'), 'average'],
                  [wb.create_sheet('exam-worst'), 'worst']]
            del wb['Sheet']
            for i in range(3):
                ws[i][0].cell(1, 1, 'hmbfl')
                ws[i][0].cell(2, 1, 'id')
                ws[i][0].cell(2, 2, 'version')
                ws[i][0].cell(2, 3, 'fault_num')
                ws[i][0].cell(2, 4, 'time_spend')
                for j, mbfl_for in enumerate(mbfl_for_list):
                    ws[i][0].cell(1, j*6+5, mbfl_for.split('.')[2])
                    ws[i][0].cell(2, j*6+5+0, 'exam1')
                    ws[i][0].cell(2, j*6+5+1, 'exam2')
                    ws[i][0].cell(2, j*6+5+2, 'exam3')
                    ws[i][0].cell(2, j*6+5+3, 'rank1')
                    ws[i][0].cell(2, j*6+5+4, 'rank2')
                    ws[i][0].cell(2, j*6+5+5, 'rank3')

            data_dirpath = './report/DSA/NewData'
            for i, file in enumerate(os.listdir(data_dirpath)):
                # print('\r %s baseline:%s, read:%s' % (datetime.datetime.now(), wb_name, file), end='')
                print('%s baseline:%s, read:%s - %s' % (datetime.datetime.now(), wb_name, file, i))
                if 'random' in file:
                    continue
                # if 'v1582' not in file:
                #     continue

                # 计算相应方法的怀疑度
                path = os.path.join(os.getcwd(), data_dirpath, file)
                sus_data_json = baseline(Tools().json_rules(path))

                if sus_data_json == False:
                    continue

                for sheet, sheet_name in ws:
                    sus_formula_j = 0
                    for key, value in sus_data_json.items():
                        if key == 'Fault_Record':
                            sheet.cell(i + 3, 1, str(i+1))
                            sheet.cell(i + 3, 2, file[5:-5])
                            sheet.cell(i + 3, 3, len(value))
                        elif key == 'totaltime':
                            sheet.cell(i + 3, 4, value)
                        else:
                            for fault_ith in range(len(value['exam_best'])):
                                rank = value['rank_%s' % sheet_name][fault_ith]
                                exam = value['exam_%s' % sheet_name][fault_ith]

                                sheet.cell(i + 3, 5+fault_ith + sus_formula_j*6, exam)
                                sheet.cell(i + 3, 5+fault_ith+3 + sus_formula_j*6, rank)
                            sus_formula_j += 1

            docpath = os.path.join(os.getcwd(), 'report', 'DSA', 'susfile', 'baseline')
            if not os.path.exists(docpath):
                os.makedirs(docpath)
            filepath = os.path.join(docpath, '%s.xlsx' % wb_name)
            wb.save(filepath)
            print('\r 文件保存 %s ' % filepath, end='')
            print('\n')
        return




    # DSA_highorder_guide_res()
    # DSA_mbfl_res()
    # DSA_highorder_res()
    # DSA_highorder_random_res()
    DSA_highorder_guide_res()
    # DSA_highorder_mbflguide_res(word)
    return


if __name__ == '__main__':
    DSA()
    # pastSbflGuideData()
    # additionalExperiments()

