import datetime
import json
# import ijson
import os
import math
# import numpy as np
import socket
import util
import openpyxl
import random
import difflib

import mbfl.mbfl_for
import mbfl.command
import sbfl.sbfl_for
import sbfl.command
from CHMBFL_Flow import Cluster_Fom
from mbfl.mutpolyn import new_mutation_trick
from CodeFlewsData.data_codeflaws import MutationRule


Rewrite = False
# Rewrite = True

# data_dirpath = './report/CHMBFL/mutinfo'
data_dirpath = './report/CHMBFL/mutinfo-new'


mbfl_for_list = [
    "mbfl.mbfl_for.Tarantula",
    "mbfl.mbfl_for.Op2",
    "mbfl.mbfl_for.Jaccard",
    "mbfl.mbfl_for.Ochiai",
    "mbfl.mbfl_for.Dstar",
    "mbfl.mbfl_for.GP13",
    "mbfl.mbfl_for.Naish1",
    "mbfl.mbfl_for.Barinel",
    # "mbfl.mbfl_for.muse",
    # "mbfl.mbfl_for.nsus",
]

sbfl_for_list = [
    "sbfl.sbfl_for.Tarantula",
    "sbfl.sbfl_for.Op2",
    "sbfl.sbfl_for.Jaccard",
    "sbfl.sbfl_for.Ochiai",
    "sbfl.sbfl_for.Dstar",
    "sbfl.sbfl_for.GP13",
    "sbfl.sbfl_for.Naish1",
    "sbfl.sbfl_for.Barinel",
]



class Tools:
    # 计算二阶变异体精确度
    def precision_calculate(self, faultrecord, som):
        if not len(som) == 2:
            print('精确度计算 输入变异体错误')
            return False
        precision = 0
        for fom in som:
            line = fom[0]
            if line in faultrecord:
                precision += 1
        return precision

    # 计算变异算子种类
    def variety_calculate(self, som):
        variety = []
        for fom in som:
            operate = fom[3]
            for i, mutation_trick_dic in enumerate(new_mutation_trick):
                if operate in list(mutation_trick_dic.keys()):
                    variety.append(i)
        return variety

    # 聚类中，使用同一类簇中的变异体进行类RandomMix生成
    def clusterSameClassRandomMix(self, clusterdata, fom_list, times):
        times = times*1.2
        som_classes = []
        tnum = 0
        for label in clusterdata:
            classfomlist = clusterdata[label]
            som_class = []
            for fom_i_loc in classfomlist:
                fom_i = fom_list[fom_i_loc]
                for fom_j_loc in classfomlist:
                    fom_j = fom_list[fom_j_loc]
                    if not MutationRule(fom_i['message'][0], fom_j['message'][0]):
                        continue
                    som = sorted([fom_i['message'][0], fom_j['message'][0]], key=lambda x: x[0])
                    som_class.append(som)
                    tnum += 1
            som_classes.append(som_class)
        try:
            ration = len(fom_list)*times/sum(list(map(lambda x: len(x), som_classes)))
        except:
            ration = 0

        som_list = []
        readnum = 0
        for som_class in som_classes:
            readnum += len(som_class)
            som_list += random.sample(som_class, min(math.ceil(len(som_class)*ration), len(som_class)))
            while len(som_list) > math.ceil(readnum*ration):
                som_list.pop(-1)

        return som_list

    # 聚类中，使用同一类簇中的变异体进行类RandomMix生成
    def clusterDifClassRandomMix(self, clusterdata, fom_list, times):
        times = times*1.2
        som_classes = []
        tnum = 0
        for i, label_i in enumerate(clusterdata):
            for j, label_j in enumerate(clusterdata):
                if i >= j:
                    continue
                # 取上三角类簇
                som_class = []

                classfomlist_i = clusterdata[label_i]
                classfomlist_j = clusterdata[label_j]
                for fom_i_loc in classfomlist_i:
                    fom_i = fom_list[fom_i_loc]
                    for fom_j_loc in classfomlist_j:
                        fom_j = fom_list[fom_j_loc]
                        if not MutationRule(fom_i['message'][0], fom_j['message'][0]):
                            continue
                        som = sorted([fom_i['message'][0], fom_j['message'][0]], key=lambda x: x[0])
                        som_class.append(som)
                        tnum += 1
                som_classes.append(som_class)
        if len(som_classes) == 0:
            return []
        ration = len(fom_list)*times/sum(list(map(lambda x: len(x), som_classes)))

        som_list = []
        readnum = 0
        for som_class in som_classes:
            readnum += len(som_class)
            som_list += random.sample(som_class, min(math.ceil(len(som_class)*ration), len(som_class)))
            while len(som_list) > math.ceil(readnum*ration):
                som_list.pop(-1)

        return som_list

    def json_rules(self, path):
        # 读取小于3g的文件
        # if os.stat(path).st_size > 3*1024*1024*1024:
        #     print('内存过大')
        #     return False

        try:
            with open(path) as f_obj:
                data_json = json.load(f_obj)
        except:
            print('读取异常 %s' % path)
            return False
        return data_json

    # 一阶数量的30倍，需要生成的高阶数量
    def guide_line_pair(self, list_source_line, mutant_num_every_order):
        now_rank = 1
        for source_line_index in range(len(list_source_line)):
            if source_line_index == 0:
                list_source_line[source_line_index][2] = now_rank
                now_rank += 1
            elif list_source_line[source_line_index][1] == list_source_line[source_line_index-1][1]:
                list_source_line[source_line_index][2] = list_source_line[source_line_index-1][2]
                now_rank += 1
            else:
                list_source_line[source_line_index][2] = now_rank
                now_rank += 1

        for source_line_index in range(len(list_source_line)):
            list_source_line[source_line_index][3] = now_rank - list_source_line[source_line_index][2]


        min_order_num = 2
        max_order_num = 2

        hom_line_pair = []
        for mutant_line_num in range(min_order_num, max_order_num+1):    # Select between the min_order_numnd-order to max_order_numth-order mutants.
            list_choice_line_index = []  # list of choice line index.
            now_choice_index = 0  # number of current selections.


            mutation_source_num = mutant_num_every_order * mutant_line_num
            list_mutation_source_num = []
            sum_mutation_source_num = 0
            for source_line in list_source_line:
                list_mutation_source_num.append([source_line[0], source_line[3] * mutant_line_num, 0.0])
                sum_mutation_source_num += (source_line[3] * mutant_line_num)
            multiple = math.ceil(mutation_source_num / sum_mutation_source_num)
            sum_mutation_source_num = 0
            for mutation_source_num_index in range(len(list_mutation_source_num)):
                list_mutation_source_num[mutation_source_num_index][1] *= multiple
                sum_mutation_source_num += list_mutation_source_num[mutation_source_num_index][1]
            # print(list_mutation_source_num)

            for i in range(mutant_num_every_order):
                list_randomization_rate = []
                list_randomization_source_line_index = []
                # ==========Exclude situations where candidate statements are insufficient========================
                for mutation_source_num_index in range(len(list_mutation_source_num)):
                    if list_mutation_source_num[mutation_source_num_index][1] == 0:
                        list_mutation_source_num[mutation_source_num_index][1] += 1
                        sum_mutation_source_num += 1
                # =================================================================================================
                for mutation_source_num_index in range(len(list_mutation_source_num)):
                    list_mutation_source_num[mutation_source_num_index][2] = list_mutation_source_num[mutation_source_num_index][1] / sum_mutation_source_num
                    list_randomization_rate.append(list_mutation_source_num[mutation_source_num_index][2])
                    list_randomization_source_line_index.append(list_mutation_source_num[mutation_source_num_index][0])
                # print(list_mutation_source_num)
                np.random.seed(random.randint(0, 999999))
                p = np.array(list_randomization_rate)


                if mutant_line_num > len(list_randomization_rate):
                    mutant_line_num = len(list_randomization_rate)

                list_choice_line_index = []
                now_choice_index = 0
                while now_choice_index < mutant_line_num:
                    choice_line_index = np.random.choice(list_randomization_source_line_index, p=p.ravel())
                    if choice_line_index in list_choice_line_index:
                        continue
                    list_choice_line_index.append(int(choice_line_index))
                    for mutation_source_num_index in range(len(list_mutation_source_num)):
                        if int(choice_line_index) == int(list_mutation_source_num[mutation_source_num_index][0]):
                            list_mutation_source_num[mutation_source_num_index][1] -= 1
                            sum_mutation_source_num -= 1
                            break
                    now_choice_index += 1
                list_choice_line_index.sort()
                # for choice_line_index in range(len(list_choice_line_index)):
                #     list_choice_line_index[choice_line_index] = str(list_choice_line_index[choice_line_index])
                hom_line_pair.append(list_choice_line_index)
                # f.write(",".join(list_choice_line_index))
                # if i != mutant_num_every_order - 1 or mutant_line_num != max_order_num:
                #     f.write("\n")
        return hom_line_pair

    def get_host_ip(self):
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM)
        s.connect(('8.8.8.8',80))
        ip=s.getsockname()[0]
        print(ip)
        return ip

    def victor_bite(self, testcasespace, testcase):
        victor = []
        for i in range(len(testcasespace)):
            if testcasespace[i] == 1:
                victor.append(testcase[i])

        return victor

class Reduction:
    def testanalysis(self, data_json, times=1):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]

            # 获取所需数据
            or_list = data["or_list"]
            fom_list = data['fom_list']
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1


        fom_sus_list = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            sus = mutOchiai(mutTouple(or_list, fom['kill_list']))
            if line not in fom_sus_list:
                fom_sus_list[line] = []
            fom_sus_list[line].append([sus, fom])

        new_fom_list = []
        for line in fom_sus_list.keys():
            fom_sus_list[line] = sorted(fom_sus_list[line], key=lambda x: x[0], reverse=True)
            neednum = math.ceil(len(fom_sus_list[line])*times)
            for sus, fom in fom_sus_list[line][:neednum]:
                new_fom_list.append(fom)

        data_json[doc]['fom_list'] = new_fom_list
        return data_json

    def samping(self, data_json, times=1):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]

            # 获取所需数据
            or_list = data["or_list"]
            fom_list = data['fom_list']
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        new_fom_list = []
        for fom in random.sample(fom_list, len(fom_list)):
            if len(new_fom_list) >= math.ceil(len(fom_list)*times):
                break
            new_fom_list.append(fom)

        data_json[doc]['fom_list'] = new_fom_list
        return data_json

    def some(self, data_json, times=1):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]

            # 获取所需数据
            or_list = data["or_list"]
            fom_list = data['fom_list']
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        fom_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_dict:
                fom_dict[line] = []
            fom_dict[line].append(fom)

        new_fom_list = []
        for line in fom_dict.keys():
            neednum = math.ceil(len(fom_dict[line])*times)
            new_fom_list += fom_dict[line][:neednum]

        data_json[doc]['fom_list'] = new_fom_list
        return data_json

    def wsome(self, data_json, times=1):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]

            # 获取所需数据
            or_list = data["or_list"]
            fom_list = data['fom_list']
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        fom_sus_list = dict()
        fom_line_list = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            sus = mutOchiai(mutTouple(or_list, fom['kill_list']))
            if line not in fom_sus_list:
                fom_sus_list[line] = []
                fom_line_list[line] = []
            fom_sus_list[line].append(sus+2)
            fom_line_list[line].append(fom)

        tw = sum(map(lambda x: max(fom_sus_list[x]), fom_sus_list.keys()))
        power_line = dict()
        for line in fom_sus_list.keys():
            power_line[line] = math.ceil(max(fom_sus_list[line])/tw*len(fom_sus_list.keys())*times)

        new_fom_list = []
        for line in fom_sus_list.keys():
            # fom_sus_list[line] = sorted(fom_sus_list[line], key=lambda x: x[0], reverse=True)
            # fom_sus_list[line] = sorted(fom_sus_list[line], reverse=True)
            neednum = power_line[line]
            # A = random.sample(fom_line_list[line], neednum)
            new_fom_list += random.sample(fom_line_list[line], min(len(fom_line_list[line]), neednum))
            # try:
            #     new_fom_list += random.sample(fom_line_list[line], min(len(fom_line_list[line]), neednum))
            # except:
            #     continue
            # new_fom_list.append(fom)

        data_json[doc]['fom_list'] = new_fom_list
        return data_json

    def none(self, data_json, times=10):
        return data_json

class Get_sus:
    # 计算last2first方法怀疑度
    def Last2First(self, data_json, times=0.5):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # ------------------------------------------------
        # 生成som组合
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            random_fomloc = random.sample(range(len(fom_list)), len(fom_list))
            while len(random_fomloc) > 1 and sus_data['homnum'] < times*sus_data['fomnum']:
                locfirst = random_fomloc.pop(0)
                fom_first = fom_list[locfirst]
                for i in range(len(random_fomloc)-1, -1, -1):
                    # fom_last = fom_list[i]
                    loclast = random_fomloc[i]
                    fom_last = fom_list[loclast]
                    if fom_first['message'][0][0] == fom_last['message'][0][0]:
                        continue
                    som = sorted([fom_first['message'][0], fom_last['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom = hom_out_list[loc]
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    random_fomloc.pop(i)
                    break
            if times == 1000:
                break

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算differentOperator方法怀疑度
    def DifferentOperator(self, data_json, times=0.5):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            random_fomlist = random.sample(fom_list, len(fom_list))
            while len(random_fomlist) > 1 and sus_data['homnum'] < times*sus_data['fomnum']:
                fom_1 = random_fomlist.pop(0)
                message_1 = fom_1['message'][0]
                for i, fom in enumerate(random_fomlist):
                    message = fom['message'][0]
                    if message_1[0] == message[0] or message_1[3] == message[3]:
                        continue
                    som = sorted([message_1, message], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    random_fomlist.pop(i)
                    hom = hom_out_list[loc]
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, som)] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(som) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    break
            if times == 1000:
                break

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算randomMix方法怀疑度
    def RandomMix(self, data_json, times=0.5):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # ------------------------------------------------
        # 生成som组合
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            random_fom_list = random.sample(fom_list, len(fom_list))
            while sus_data['homnum'] < times*sus_data['fomnum'] and len(random_fom_list) > 0:
                loc1 = random.sample(range(len(random_fom_list)), 1)[0]
                fom1 = random_fom_list.pop(loc1)
                for loc2 in random.sample(range(len(random_fom_list)), len(random_fom_list)):
                    fom2 = random_fom_list[loc2]
                    if fom1['message'][0][0] == fom2['message'][0][0]:
                        continue
                    som_message = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x:x[0])
                    try:
                        loc = hom_list_simple.index(som_message)
                        hom = hom_out_list[loc]
                        random_fom_list.pop(loc2)
                    except:
                        continue
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, som_message)] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(som_message) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    break
            if times == 1000:
                break

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 计算fom的times倍数的 random
    def Random(self, data_json, times):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]

        out_dic = dict()
        kill_dic = dict()
        for hom in random.sample(hom_out_list,  min(fomnum*times, len(hom_out_list))):
            message = hom['message']
            if not MutationRule(message[0], message[1]):
                # 同一行进行了叠加变异
                continue
            out_list = hom['out_list']
            kill_list = hom['kill_list']
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            precisions[Tools().precision_calculate(Fault_Record, message)] += 1
            varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(message) else x[1],
                                enumerate(varietys)))
            for fom in message:
                line = fom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(out_list)
                kill_dic[line].append(kill_list)

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                def charge(ranks, Fault_Record, touple):
                    for line in Fault_Record:
                        if line not in touple['linedata'].keys():
                            return 0
                    if sum(map(lambda x:x[1], ranks)) == len(ranks):
                        return 0
                    print(len(touple['linedata'].keys()), ranks)
                    return 0
                if mbfl_for == "mbfl.mbfl_for.Tarantula" and word == 'max':
                    charge(ranks, Fault_Record, touple)


                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)

                sus_data[word][mbfl_for] = dict()

                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data


    # 计算mbfl方法怀疑度
    def DeltaMbfl(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False


        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'frequency'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        for mbfl_for in mbfl_for_list:
            sus_data[mbfl_for] = dict()
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_list.append([line, max(sus)-sum(sus)/len(sus)])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[mbfl_for]['rank_best'] = rank[0]
            sus_data[mbfl_for]['rank_average'] = rank[1]
            sus_data[mbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[mbfl_for]['exam_best'] = exam[0]
            sus_data[mbfl_for]['exam_average'] = exam[1]
            sus_data[mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 计算sbfl方法怀疑度
    def Sbfl(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            Fault_Record = data["Fault_Record"]
            # fom_list = data['fom_list']
            # fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            sbfl_data = data["sbfl"]
            if "linelen" in data:
                line_len = data['linelen']
            else:
                src_path = data['path']
                try:
                    path_list = src_path.split('/')
                    while len(path_list) > 0:
                        if not path_list[0] == 'access':
                            path_list.pop(0)
                        else:
                            path_list.pop(0)
                            break
                    nowpath = os.getcwd()
                    for part in path_list:
                        nowpath = os.path.join(nowpath, part)
                    line_len = len(util.File().read_line(nowpath))
                except:
                    line_len = 0


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = 0
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False



        # ------------------------------------------------
        # 计算sbfl怀疑度
        touple = sbfl.command.touple(or_list, sbfl_data['cov'])
        for sbfl_for in sbfl_for_list:
            sus_data[sbfl_for] = dict()
            sus_list = []
            for line, touple_list in touple.items():
                sus = eval(sbfl_for)(touple_list)
                sus_list.append([line, sus])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[sbfl_for]['rank_best'] = rank[0]
            sus_data[sbfl_for]['rank_average'] = rank[1]
            sus_data[sbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[sbfl_for]['exam_best'] = exam[0]
            sus_data[sbfl_for]['exam_average'] = exam[1]
            sus_data[sbfl_for]['exam_worst'] = exam[2]
        sus_data['totaltime'] = sbfl_data['time']
        return sus_data

    # 计算sbfl方法怀疑度
    def Mcbfl(self, data_json):
        mcbfl_for_list = [
            ["sbfl.sbfl_for.Tarantula", "mbfl.mbfl_for.Tarantula"],
            ["sbfl.sbfl_for.Op2", "mbfl.mbfl_for.Op2"],
            ["sbfl.sbfl_for.Jaccard", "mbfl.mbfl_for.Jaccard"],
            ["sbfl.sbfl_for.Ochiai", "mbfl.mbfl_for.Ochiai"],
            ["sbfl.sbfl_for.Dstar", "mbfl.mbfl_for.Dstar"],
            ["sbfl.sbfl_for.GP13", "mbfl.mbfl_for.GP13"],
            ["sbfl.sbfl_for.Naish1", "mbfl.mbfl_for.Naish1"],
            ["sbfl.sbfl_for.Barinel", "mbfl.mbfl_for.Barinel"],
        ]


        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False


        # ------------------------------------------------
        # 获取mbfl元组
        mbfl.mbfl_for.type_mbfl = 'max'
        totaltime = 0
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            totaltime += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # 获取sbfl元组
        totaltime += sbfl_data['time']
        touple_sbfl = sbfl.command.touple(or_list, sbfl_data['cov'])

        sus_data['totaltime'] = totaltime

        # 计算怀疑度
        for sbfl_for, mbfl_for in mcbfl_for_list:
            sus_data[mbfl_for] = dict()
            sus_dic = dict()

            # 计算mbfl怀疑度
            for line in touple_mbfl['linedata']:
                touple_list = touple_mbfl['tf'], \
                              touple_mbfl['tp'], \
                              touple_mbfl['linedata'][line]['f'], \
                              touple_mbfl['linedata'][line]['p'], \
                              touple_mbfl['f2p'], \
                              touple_mbfl['p2f'], \
                              touple_mbfl['linedata'][line]['kf'], \
                              touple_mbfl['linedata'][line]['kp'], \
                              touple_mbfl['linedata'][line]['nf'], \
                              touple_mbfl['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_dic[line] = sus + -1

            # 计算sbfl怀疑度
            for line, touple_list in touple_sbfl.items():
                sus = eval(sbfl_for)(touple_list)
                if line in sus_dic:
                    sus_dic[line] += sus + 1
                else:
                    sus_dic[line] = sus + -1

            sus_list = []
            for line, sus in sus_dic.items():
                sus_list.append([line, sus/2])

            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

            rank = [[], [], []]
            for fault in ranks:
                rank[0].append(fault[1])
                rank[1].append(fault[2])
                rank[2].append(fault[3])
            sus_data[mbfl_for]['rank_best'] = rank[0]
            sus_data[mbfl_for]['rank_average'] = rank[1]
            sus_data[mbfl_for]['rank_worst'] = rank[2]

            exam = [[], [], []]
            for ith in range(len(ranks)):
                if line_len > 0:
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)
                else:
                    exam[0].append(rank[0][ith])
                    exam[1].append(rank[1][ith])
                    exam[2].append(rank[2][ith])

            sus_data[mbfl_for]['exam_best'] = exam[0]
            sus_data[mbfl_for]['exam_average'] = exam[1]
            sus_data[mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # sbfl指导生成hmbfl
    def get_sus_SBFLguide_hmbfl(self, data_json, times):
        # print('%s 读取数据' % datetime.datetime.now())
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            # nowpath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        def get_som_out(som_message_self, hom_list_all_self):
            som_dict = {}
            for hom_all in hom_list_all_self:
                if hom_all["message"][0] == som_message_self[0] and hom_all["message"][1] == som_message_self[1]:
                    for fom_message in som_message:
                        som_dict[fom_message[0]] = [hom_all["out_list"], hom_all["kill_list"]]
                    return som_dict
            return False


        # print('%s 计算sbfl怀疑度' % datetime.datetime.now())


        # ------------------------------------------------
        # 计算sbfl怀疑度
        touple = sbfl.command.touple(or_list, sbfl_data['cov'])
        sus_list = []
        for line, touple_list in touple.items():
            sus = sbfl.sbfl_for.GP13(touple_list)
            sus_list.append([line, sus, -1, -1])

        # print('%s 获取组队数据' % datetime.datetime.now())

        # 获取组队数据
        line_pairs_list = Tools().guide_line_pair(sus_list, fomnum*times)
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))



        # print('%s 获取高阶元组' % datetime.datetime.now())


        # 获取元组输入字典
        sfterload_hom = []

        out_dic = dict()
        kill_dic = dict()
        som_num = 0
        # print('%s len %s ' % (datetime.datetime.now(), len(line_pairs_list)))
        for pair in line_pairs_list:
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                # print('%s %s unable' % (datetime.datetime.now(), pair))
                continue
            # print('%s %s able' % (datetime.datetime.now(), pair))

            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))

            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                som_num += 1
                break

            if len(random_som_pair_list) > 0:
                sfterload_hom += random_som_pair_list

        while som_num < len(line_pairs_list) and len(sfterload_hom) > 0:
            som_message = sfterload_hom.pop(0)
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_list_all[loc]
            sus_data['totaltime'] += hom['time']
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])
            som_num += 1

        # print('%s som_num %s' % (datetime.datetime.now(), som_num))

        if som_num/len(line_pairs_list) < 0.8:
            # print('变异体数量不足')
            return False

        # print('%s 计算高阶怀疑度' % datetime.datetime.now())

        # touple = mbfl.command.get_toule(true_out, or_out, out_dic)
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        sus_data['totaltime'] = sbfl_data['time']
        return sus_data

    # mbfl指导生成hmbfl
    def get_sus_MBFLguide_hmbfl(self, data_json, times):
        # print('%s 读取数据' % datetime.datetime.now())
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json


            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            src_path = data['path']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)
            del data

            sus_data = dict()
            sus_data['Fault_Record'] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        try:
            path_list = src_path.split('/')
            while len(path_list) > 0:
                if not path_list[0] == 'access':
                    path_list.pop(0)
                else:
                    path_list.pop(0)
                    break
            # nowpath = os.path.abspath(os.path.join(os.path.dirname("__file__"),os.path.pardir))
            nowpath = os.getcwd()
            for part in path_list:
                nowpath = os.path.join(nowpath, part)
            line_len = len(util.File().read_line(nowpath))
        except:
            line_len = 0

        def get_som_out(som_message_self, hom_list_all_self):
            som_dict = {}
            for hom_all in hom_list_all_self:
                if hom_all["message"][0] == som_message_self[0] and hom_all["message"][1] == som_message_self[1]:
                    # if "out_or" not in hom_all:
                    #     continue
                    # if len(hom_all["out_or"]) == 0:
                    #     continue
                    for fom_message in som_message:
                        som_dict[fom_message[0]] = [hom_all["out_list"], hom_all["kill_list"]]
                    return som_dict
            return False

        # print('%s 计算一阶怀疑度' % datetime.datetime.now())
        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
        sus_list = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list.append([line, sus, -1, -1])

        # print('%s 获取组队数据' % datetime.datetime.now())

        # 获取组队数据
        line_pairs_list = Tools().guide_line_pair(sus_list, fomnum*times)
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # print('%s 获取高阶元组' % datetime.datetime.now())

        # 获取元组输入字典

        sfterload_hom = []

        out_dic = dict()
        kill_dic = dict()
        som_num = 0
        # print('%s len %s ' % (datetime.datetime.now(), len(line_pairs_list)))
        for pair in line_pairs_list:
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue

            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    som_pair_list.append([fom0['message'][0], fom1['message'][0]])

            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_list_all[loc]
                sus_data['totaltime'] += hom['time']
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                som_num += 1
                break

            if len(random_som_pair_list) > 0:
                sfterload_hom += random_som_pair_list


            # for som_message in random.sample(som_pair_list, len(som_pair_list)):
            #     try:
            #         loc = hom_list_simple.index(som_message)
            #     except:
            #         continue
            #     hom = hom_list_all[loc]
            #     sus_data['totaltime'] += hom['time']
            #     for fom_in_hom in hom['message']:
            #         line = fom_in_hom[0]
            #         if line not in out_dic:
            #             out_dic[line] = []
            #             kill_dic[line] = []
            #         out_dic[line].append(hom['out_list'])
            #         kill_dic[line].append(hom['kill_list'])
            #     som_num += 1
            #     break
        # print('%s 补充高阶元组 %s- %s' % (datetime.datetime.now(), som_num , len(line_pairs_list)))
        while som_num < len(line_pairs_list) and len(sfterload_hom) > 0:
            som_message = sfterload_hom.pop(0)
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_list_all[loc]
            sus_data['totaltime'] += hom['time']
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])
            som_num += 1

        if som_num < len(line_pairs_list):
            print('变异体数量不足')
            return False

        # print('%s 计算高阶怀疑度' % datetime.datetime.now())

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        sus_data['totaltime'] += sbfl_data['time']
        return sus_data

    # 计算聚类怀疑度
    def get_sus_numneed_cluster(self, data_json, function, claster_num):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            # or_out = data["or_out"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            hom_list_all = data['hom_list_all']
            # true_out = data["true_out"]
            # sbfl_data = data["sbfl"]
            fomnum = len(fom_list)
            del data

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 聚类生成二阶变异体


        clusterdata = function(points_data, claster_num).ClusterMain()
        if not clusterdata:
            return False
        som_message_list = Cluster_Fom.HomGenerationStrategy().clusterSameClassRandomMix(clusterdata, fom_list)

        out_dic = dict()
        kill_dic = dict()
        for som_message in som_message_list:
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_list_all[loc]
            sus_data['totaltime'] += hom['time']
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])


                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 基于变异体聚类
    def MutCluster(self, data_json, seting=[True, True], times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 聚类生成二阶变异体
        if seting[0]:
            clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        else:
            clusterdata = Cluster_Fom.Kmeans(points_data, 2).ClusterMain()
        sus_data['Desrcibe'] = len(list(clusterdata.keys()))

        if seting[1]:
            som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, times)
        else:
            som_message_list = Tools().clusterSameClassRandomMix(clusterdata, fom_list, times)

        out_dic = dict()
        kill_dic = dict()
        for som_message in som_message_list:
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_out_list[loc]
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
            varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                enumerate(varietys)))
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        alpha_list = []
        for alpha_part_i in range(2):
            try:
                alpha_list.append(sum(alpha_part[alpha_part_i])/len(alpha_part[alpha_part_i])/touple['tf'])
            except:
                alpha_list.append(0)
        alpha = str(alpha_list)
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 基于失败测试用例
    def FailTestCluster(self, data_json, seting, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 生成二阶变异体
        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        # 计算失败测试用例特征序列
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        if seting[0]:
            clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        else:
            clusterdata = Cluster_Fom.Kmeans(points_data, 2).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)
        T = sum(list(map(lambda x: x[1], pairpower)))

        out_dic = dict()
        kill_dic = dict()
        restnum = 0
        for pair, neednum in pairpower:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # mbfl 指导 ErrorDistribution
    def ErrorDistribution(self, data_json, seting=[True], times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0

            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]

        # ------------------------------------------------
        sus_list = []
        if seting[0]:
            # mbfl计算一阶怀疑度
            mbfl.mbfl_for.type_mbfl = 'max'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = fom["out_list"]
                fom_kill_list = fom["kill_list"]
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sus])
        else:
            # sbfl计算一阶怀疑度
            touple = sbfl.command.touple(or_list, sbfl_data['cov'])
            for line, touple_list in touple.items():
                sus = sbfl.sbfl_for.Ochiai(touple_list)
                sus_list.append([line, sus])

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        out_dic = dict()
        kill_dic = dict()
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # mbfl 指导 ErrorDistribution
    def TestAnalysis(self, data_json, seting=[], times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0

            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_sus_list = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            sus = mutOchiai(mutTouple(or_list, fom['kill_list']))
            if line not in fom_sus_list:
                fom_sus_list[line] = []
            fom_sus_list[line].append([sus, fom['message'][0]])
        for line in fom_sus_list.keys():
            fom_sus_list[line] = sorted(fom_sus_list[line], key=lambda x: x[0], reverse=True)

        topNumRead = -1
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            topNumRead += 1
            add_fom_list = []
            exist_fom_list = []
            for line in fom_sus_list.keys():
                for topNum, [sus, fom] in enumerate(fom_sus_list[line]):
                    if topNum < topNumRead:
                        exist_fom_list.append(fom)
                    else:
                        add_fom_list.append(fom)
                        break
            if len(add_fom_list) == 0:
                break
            while len(add_fom_list) > 0 and  sus_data['homnum'] < times*sus_data['fomnum']:
                addFom = add_fom_list.pop(0)
                for existFom in exist_fom_list:
                    if not MutationRule(addFom, existFom):
                        continue
                    som_message = sorted([addFom, existFom], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som_message)
                    except:
                        continue
                    hom = hom_out_list[loc]
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                exist_fom_list.append(addFom)

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # mbfl 指导 ErrorDistribution
    def Samping(self, data_json, seting=[], times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0

            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_sus_list = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            sus = mutOchiai(mutTouple(or_list, fom['kill_list']))
            if line not in fom_sus_list:
                fom_sus_list[line] = []
            fom_sus_list[line].append([sus, fom['message'][0]])
        for line in fom_sus_list:
            fom_sus_list[line] = sorted(fom_sus_list[line], key=lambda x: x[0], reverse=True)

        topNumRead = -1
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum']:
            topNumRead += 1
            add_fom_list = []
            exist_fom_list = []
            for line in fom_sus_list:
                for topNum, [sus, fom] in enumerate(fom_sus_list[line]):
                    if topNum < topNumRead:
                        exist_fom_list.append(fom)
                    else:
                        add_fom_list.append(fom)
                        break
            if len(add_fom_list) == 0:
                break
            while len(add_fom_list) > 0 and  sus_data['homnum'] < times*sus_data['fomnum']:
                addFom = add_fom_list.pop(0)
                for existFom in exist_fom_list:
                    if not MutationRule(addFom, existFom):
                        continue
                    som_message = sorted([addFom, existFom], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som_message)
                    except:
                        continue
                    hom = hom_out_list[loc]
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                exist_fom_list.append(addFom)

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # mbfl 指导 ErrorDistribution
    def SomHom(self, data_json, seting=[], times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0

            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_sus_list = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            sus = mutOchiai(mutTouple(or_list, fom['kill_list']))
            if line not in fom_sus_list:
                fom_sus_list[line] = []
            fom_sus_list[line].append([sus, fom['message'][0]])
        for line in fom_sus_list.keys():
            # fom_sus_list[line] = sorted(fom_sus_list[line], key=lambda x: x[0], reverse=True)
            fom_sus_list[line] = random.sample(fom_sus_list[line], len(fom_sus_list[line]))

        topPercent = 0
        out_dic = dict()
        kill_dic = dict()
        while sus_data['homnum'] < times*sus_data['fomnum'] and topPercent<100:
            topPercent += 1
            add_fom_list = []
            exist_fom_list = []
            for line in fom_sus_list.keys():
                for topNum, [sus, fom] in enumerate(fom_sus_list[line]):
                    if topNum <= math.ceil(len(fom_sus_list[line])*(topPercent-1)):
                        exist_fom_list.append(fom)
                    elif topNum <= math.ceil(len(fom_sus_list[line])*topPercent):
                        add_fom_list.append(fom)
                    else:
                        break
            # if len(add_fom_list) == 0:
            #     break
            while len(add_fom_list) > 0 and  sus_data['homnum'] < times*sus_data['fomnum']:
                addFom = add_fom_list.pop(0)
                for existFom in exist_fom_list:
                    if not MutationRule(addFom, existFom):
                        continue
                    som_message = sorted([addFom, existFom], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som_message)
                    except:
                        continue
                    hom = hom_out_list[loc]
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                exist_fom_list.append(addFom)

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # mbfl 指导 ErrorDistribution
    def WSomHom(self, data_json, seting=[], times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0

            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # 计算单个变异体元组信息
        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        # 计算单个变异体Ochiai怀疑度
        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def getSus(fom_data_list, fom_message):
            for fom, sus in fom_data_list:
                if fom == fom_message:
                    if sus == 0:
                        return 1
                    else:
                        return 0
            return 0

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        linePower = dict()
        tPower = 0
        fom_sus_list = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            sus = mutOchiai(mutTouple(or_list, fom['kill_list']))
            if line not in fom_sus_list:
                fom_sus_list[line] = []
            fom_sus_list[line].append([sus, fom['message'][0]])
        for line in fom_sus_list.keys():
            fom_sus_list[line] = random.sample(fom_sus_list[line], len(fom_sus_list[line]))
            # linePower[line] = max(fom_sus_list[line])
            linePower[line] = max(map(lambda x:x[0], fom_sus_list[line]))
            tPower += linePower[line]
        needFomNum = math.pow(times*1.5*sus_data['fomnum']*2, 0.5)
        for line in linePower.keys():
            if tPower == 0:
                linePower[line] = linePower[line]/len(linePower.keys())*needFomNum
            else:
                linePower[line] = linePower[line]/tPower*needFomNum


        out_dic = dict()
        kill_dic = dict()
        topPercent = -1
        while sus_data['homnum'] < times*sus_data['fomnum']:
            topPercent += 1
            add_fom_list = []
            exist_fom_list = []
            for line in fom_sus_list.keys():
                for topNum, [sus, fom] in enumerate(fom_sus_list[line]):
                    if topNum > linePower[line]:
                        break
                    elif topNum < topPercent:
                        exist_fom_list.append(fom)
                    elif topNum == topPercent:
                        add_fom_list.append(fom)
                    else:
                        continue
            if len(add_fom_list) == 0:
                break
            while len(add_fom_list) > 0 and sus_data['homnum'] < times*sus_data['fomnum']:
                addFom = add_fom_list.pop(0)
                for existFom in exist_fom_list:
                    if not MutationRule(addFom, existFom):
                        continue
                    som_message = sorted([addFom, existFom], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som_message)
                    except:
                        continue
                    hom = hom_out_list[loc]
                    sus_data['totaltime'] += hom['time']
                    sus_data['homnum'] += 1
                    precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                    varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                        enumerate(varietys)))
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                exist_fom_list.append(addFom)

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 非叠加变异random
    def get_sus_Nonsuperposed_Random(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            hom_list_all = data['hom_list_all']
            sbfl_data = data["sbfl"]


            tnum = len(data["testcase"])
            Locp = len(list(set(map(lambda x: x['message'][0][0], fom_list))))
            Mmin = Locp*math.factorial(tnum)/math.factorial(tnum//2)/math.factorial(tnum-tnum//2)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['testcasenum'] = tnum
            sus_data['Locp'] = Locp
            sus_data['Mmin'] = Mmin
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False



        out_dic = dict()
        kill_dic = dict()
        for hom in hom_list_all:
            message = hom['message']
            out_list = hom['out_list']
            kill_list = hom['kill_list']
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            for fom in message:
                line = fom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(out_list)
                kill_dic[line].append(kill_list)
        del hom_list_all

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    exam[0].append(rank[0][ith]/line_len)
                    exam[1].append(rank[1][ith]/line_len)
                    exam[2].append(rank[2][ith]/line_len)

                sus_data[word][mbfl_for] = dict()

                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        return sus_data

    # 根据怀疑度大小重组
    def RegroupBySus(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        # 计算一阶变异体怀疑度
        # [message, sus]
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        def setdic():
            out_dic = dict()
            kill_dic = dict()
            while sus_data['homnum'] < sus_data['fomnum']*times:
                for fom1, sus1 in random.sample(fom_data_list, len(fom_data_list)):
                    for fom2, sus2 in random.sample(fom_data_list, len(fom_data_list)):
                        if not classification(sus1) == classification(sus2):
                            continue
                        if classification(sus1) < 1:
                            continue
                        if not MutationRule(fom1, fom2):
                            continue
                        som_message = sorted([fom1, fom2], key=lambda x:x[0])
                        try:
                            loc = hom_list_simple.index(som_message)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        sus_data['totaltime'] += hom['time']
                        sus_data['homnum'] += 1
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
                        if sus_data['homnum'] >= sus_data['fomnum']*times:
                            return out_dic, kill_dic
                if len(out_dic.keys()) == 0:
                    return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 根据怀疑度大小重组
    def NewRegroupBySus(self, data_json, ratio=1):
        ratio = 0.1
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def DeltaT(or_list, out_list):
            Tpassm = sum(out_list)
            Tfailm = len(out_list)-sum(out_list)
            Tpassp = sum(or_list)
            Tfailp = len(or_list)-sum(or_list)
            DeltaTpassm = (Tpassm-Tpassp)/Tpassp
            DeltaTfailm = (Tfailm-Tfailp)/Tfailp
            DeltaTm = DeltaTpassm+DeltaTfailm
            return DeltaTm

        # 计算一阶变异体怀疑度
        # [message, sus]
        # fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        fom_data_dict = dict()
        for fom in fom_list:
            DeltaTm = DeltaT(or_list, fom["out_list"])
            line = fom["message"][0][0]
            if line not in fom_data_dict:
                fom_data_dict[line] = []
            fom_data_dict[line].append([fom["message"][0], DeltaTm])
        for line in fom_data_dict:
            fom_data_dict[line] = sorted(fom_data_dict[line], key=lambda x: x[1], reverse=True)


        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        def setdic():
            out_dic = dict()
            kill_dic = dict()
            for line1, fom_data1_list in fom_data_dict.items():
                for line2, fom_data2_list in fom_data_dict.items():
                    if line1 >= line2:
                        continue
                    for loc1, [fom1, sus1] in enumerate(fom_data1_list):
                        if loc1 > math.ceil(ratio*len(fom_data1_list)):
                            continue
                        for loc2, [fom2, sus2] in enumerate(fom_data2_list):
                            if loc2 > math.ceil(ratio*len(fom_data2_list)):
                                continue
                            som_message = sorted([fom1, fom2], key=lambda x :x[0])
                            try:
                                loc = hom_list_simple.index(som_message)
                            except:
                                continue
                            hom = hom_out_list[loc]
                            sus_data['totaltime'] += hom['time']
                            sus_data['homnum'] += 1
                            for fom_in_hom in hom['message']:
                                line = fom_in_hom[0]
                                if line not in out_dic:
                                    out_dic[line] = []
                                    kill_dic[line] = []
                                out_dic[line].append(hom['out_list'])
                                kill_dic[line].append(hom['kill_list'])
            return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 基于失败测试用例
    def NSpFTC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        times = times/2

        # 生成二阶变异体
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # --------------------------------------------------------------------------------------------------------------
        # 计算FTC
        # 计算失败测试用例特征序列
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)

        out_dic = dict()
        kill_dic = dict()
        restnum = 0
        for pair, neednum in pairpower:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        # --------------------------------------------------------------------------------------------------------------
        # 计算NS
        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_line_dict.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        guide = 'mbfl'
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0],
                                            (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2
                                            +fomnum*times/2/len(sus_list)],
                                 enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])
        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*math.ceil(fomnum*times/tnum)],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            # random_hom_lines = hom_lines_dic[line]
            random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

            while resnum >= 1 and len(random_hom_lines) > 0:
                # for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                hom = random_hom_lines.pop(0)
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                if len(random_hom_lines) == 0:
                    random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

        # --------------------------------------------------------------------------------------------------------------
        # 开始计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 基于变异体聚类+基于错误分布
    def MC_ED(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # 获取一阶特征
        points_data = list(map(lambda x: x['out_list'], fom_list))

        # 聚类生成二阶变异体
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, 10000)
        # som_message_list = Tools().clusterSameClassRandomMix(clusterdata, fom_list, 10000)

        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # sus_list_mbfl_sorted = sorted(sus_list_mbfl, key=lambda x: x[1], reverse=True)

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        out_dic = dict()
        kill_dic = dict()
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    som_message_list.index(som_message)
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                # if sum(map(lambda x: getSus(fom_data_list, x), hom['message'])) > 0:
                #     continue
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 基于变异体聚类+基于错误分布  各取50%
    def MCpED(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # 获取一阶特征
        points_data = list(map(lambda x: x['out_list'], fom_list))

        # 聚类生成二阶变异体
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, times/2)

        out_dic = dict()
        kill_dic = dict()
        for som_message in som_message_list:
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_out_list[loc]
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
            varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                enumerate(varietys)))
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])

        # ------------------------------------------------
        # 计算一阶怀疑度
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # sus_list_mbfl_sorted = sorted(sus_list_mbfl, key=lambda x: x[1], reverse=True)

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times/2], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 基于变异体聚类+基于失败测试聚类
    def MC_FTC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 聚类生成二阶变异体
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, 10000)
        # som_message_list = Tools().clusterSameClassRandomMix(clusterdata, fom_list, 10000)

        # ------------------------------------------------
        # 计算失败测试用例特征序列
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)
        T = sum(list(map(lambda x: x[1], pairpower)))



        out_dic = dict()
        kill_dic = dict()
        restnum = 0
        for pair, neednum in pairpower:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    som_message_list.index(som_message)
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 基于变异体聚类+基于失败测试聚类
    def MCpFTC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 聚类生成二阶变异体
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, times/2)

        out_dic = dict()
        kill_dic = dict()
        for som_message in som_message_list:
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_out_list[loc]
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])

        # ------------------------------------------------
        # 计算失败测试用例特征序列
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times/2*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times/2*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)

        restnum = 0
        for pair, neednum in pairpower:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 基于变异体聚类
    def NSpMC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        times = times/2
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 生成二阶变异体
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # --------------------------------------------------------------------------------------------------------------
        # 变异体聚类
        # 聚类生成二阶变异体
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        sus_data['Desrcibe'] = len(list(clusterdata.keys()))
        som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, times)

        out_dic = dict()
        kill_dic = dict()
        for som_message in som_message_list:
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_out_list[loc]
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
            varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                enumerate(varietys)))
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])

        # --------------------------------------------------------------------------------------------------------------
        # 计算NS
        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_line_dict.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        guide = 'mbfl'
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0],
                                            (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2
                                            +fomnum*times/2/len(sus_list)],
                                 enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])
        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*math.ceil(fomnum*times/tnum)],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

            while resnum >= 1 and len(random_hom_lines) > 0:
                hom = random_hom_lines.pop(0)
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                if len(random_hom_lines) == 0:
                    random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

        # --------------------------------------------------------------------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 基于变异体聚类
    def NSpED(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]
            fomnum = len(fom_list)

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = len(fom_list)
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        times = times/2

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 生成二阶变异体
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # --------------------------------------------------------------------------------------------------------------
        # 错误分布
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        out_dic = dict()
        kill_dic = dict()
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        # --------------------------------------------------------------------------------------------------------------
        # 计算NS
        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_line_dict.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        guide = 'mbfl'
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0],
                                            (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2
                                            +fomnum*times/2/len(sus_list)],
                                 enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])
        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*math.ceil(fomnum*times/tnum)],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

            while resnum >= 1 and len(random_hom_lines) > 0:
                hom = random_hom_lines.pop(0)
                sus_data['homnum'] += 1
                sus_data['totaltime'] += hom['time']
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                if len(random_hom_lines) == 0:
                    random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

        # --------------------------------------------------------------------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for mbfl_for in mbfl_for_list:
                sus_data[word][mbfl_for] = dict()
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 基于变异体聚类+基于失败测试聚类
    def MC_RS(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]


            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # 获取一阶特征
        points_data = []
        for fom in fom_list:
            points_data.append(fom['out_list'])

        # 聚类生成二阶变异体
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        som_message_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, 10000)
        # som_message_list = Tools().clusterSameClassRandomMix(clusterdata, fom_list, 10000)

        # ------------------------------------------------
        # 计算一阶变异体怀疑度
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        def setdic():
            out_dic = dict()
            kill_dic = dict()
            while sus_data['homnum'] < sus_data['fomnum']*times:
                for fom1, sus1 in random.sample(fom_data_list, len(fom_data_list)):
                    for fom2, sus2 in random.sample(fom_data_list, len(fom_data_list)):
                        if not classification(sus1) == classification(sus2):
                            continue
                        if classification(sus1) > 0:
                            continue
                        if not MutationRule(fom1, fom2):
                            continue
                        som_message = sorted([fom1, fom2], key=lambda x: x[0])
                        try:
                            som_message_list.index(som_message)
                            loc = hom_list_simple.index(som_message)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        sus_data['totaltime'] += hom['time']
                        sus_data['homnum'] += 1
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
                        if sus_data['homnum'] >= sus_data['fomnum']*times:
                            return out_dic, kill_dic
                if len(out_dic.keys()) == 0:
                    return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        # ------------------------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 根据怀疑度大小重组
    def RS_ED(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        # --------------------------------
        # 计算RS高阶变异体
        som_message_list = []
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        for fom1, sus1 in random.sample(fom_data_list, len(fom_data_list)):
            for fom2, sus2 in random.sample(fom_data_list, len(fom_data_list)):
                if not classification(sus1) == classification(sus2):
                    continue
                if classification(sus1) >0:
                    continue
                som_message = sorted([fom1, fom2], key=lambda x: x[0])
                som_message_list.append(som_message)

        # ---------------------------------
        # 计算mbfl指导矩阵
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def setdic():
            restnum = 0
            out_dic = dict()
            kill_dic = dict()
            while sus_data['homnum'] < times*sus_data['fomnum']:
                for pairloc in range(len(pairlist)):
                    pair = pairlist[pairloc][0]
                    restnum += pairlist[pairloc][1]
                    pairlist[pairloc][0] == 0
                    som_pair_list = []
                    if not pairable(pair, fom_line_dict):
                        continue
                    for fom0 in fom_line_dict[pair[0]]:
                        for fom1 in fom_line_dict[pair[1]]:
                            if not MutationRule(fom0['message'][0], fom1['message'][0]):
                                continue
                            som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
                    random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
                    while restnum > 0 and len(random_som_pair_list) > 0:
                        som_message = random_som_pair_list.pop(0)
                        try:
                            som_message_list.index(som_message)
                            loc = hom_list_simple.index(som_message)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        sus_data['totaltime'] += hom['time']
                        sus_data['homnum'] += 1
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
                        restnum -= 1
                    if sus_data['homnum'] >= times*sus_data['fomnum']:
                        return out_dic, kill_dic
                if len(out_dic.keys()) == 0:
                    return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 根据怀疑度大小重组
    def RS_FTC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # --------------------------------
        # 计算RS高阶变异体
        som_message_list = []
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        for fom1, sus1 in random.sample(fom_data_list, len(fom_data_list)):
            for fom2, sus2 in random.sample(fom_data_list, len(fom_data_list)):
                if not classification(sus1) == classification(sus2):
                    continue
                if classification(sus1) > 0:
                    continue
                som_message = sorted([fom1, fom2], key=lambda x: x[0])
                som_message_list.append(som_message)

        # ---------------------------------
        # 计算失败测试用例特征序列
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)
        def setdic():
            out_dic = dict()
            kill_dic = dict()
            restnum = 0
            while sus_data['homnum'] < times*sus_data['fomnum']:
                for pairloc in range(len(pairpower)):
                    # print(pairloc)
                    pair = pairpower[pairloc][0]
                    restnum += pairpower[pairloc][1]
                    pairpower[pairloc][1] = 0
                    som_pair_list = []
                    if not pairable(pair, fom_line_dict):
                        continue
                    for fom0 in fom_line_dict[pair[0]]:
                        for fom1 in fom_line_dict[pair[1]]:
                            if not MutationRule(fom0['message'][0], fom1['message'][0]):
                                continue
                            som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))
                    random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
                    while restnum > 0 and len(random_som_pair_list) > 0:
                        som_message = random_som_pair_list.pop(0)
                        try:
                            som_message_list.index(som_message)
                            loc = hom_list_simple.index(som_message)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        sus_data['totaltime'] += hom['time']
                        sus_data['homnum'] += 1
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
                        restnum -= 1
                    if sus_data['homnum'] >= times*sus_data['fomnum']:
                        return out_dic, kill_dic
                if len(out_dic.keys()) == 0:
                    return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        # -------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 根据怀疑度大小重组
    def MC_RS_ED(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # --------------------------------
        # 计算RS高阶变异体
        rs_som_list = []
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        for fom1, sus1 in random.sample(fom_data_list, len(fom_data_list)):
            for fom2, sus2 in random.sample(fom_data_list, len(fom_data_list)):
                if not classification(sus1) == classification(sus2):
                    continue
                if classification(sus1) > 0:
                    continue
                som_message = sorted([fom1, fom2], key=lambda x: x[0])
                rs_som_list.append(som_message)

        # --------------------------------
        # 计算MC高阶变异体
        points_data = list(map(lambda x: x['out_list'], fom_list))
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        mc_som_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, 10000)
        # mc_som_list = Tools().clusterSameClassRandomMix(clusterdata, fom_list, 10000)

        # ---------------------------------
        # 计算错误分布变异体
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        def setdic():
            restnum = 0
            out_dic = dict()
            kill_dic = dict()
            while sus_data['homnum'] < times*sus_data['fomnum']:
                for pairloc in range(len(pairlist)):
                    pair, neednum = pairlist[pairloc]
                    restnum += neednum
                    pairlist[pairloc][1] = 0
                    som_pair_list = []
                    if not pairable(pair, fom_line_dict):
                        continue
                    for fom0 in fom_line_dict[pair[0]]:
                        for fom1 in fom_line_dict[pair[1]]:
                            if not MutationRule(fom0['message'][0], fom1['message'][0]):
                                continue
                            som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
                    random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
                    while restnum > 0 and len(random_som_pair_list) > 0:
                        som_message = random_som_pair_list.pop(0)
                        if som_message not in rs_som_list and som_message not in mc_som_list:
                            continue
                        try:
                            loc = hom_list_simple.index(som_message)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        sus_data['totaltime'] += hom['time']
                        sus_data['homnum'] += 1
                        # if sum(map(lambda x: getSus(fom_data_list, x), hom['message'])) > 0:
                        #     continue
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
                        restnum -= 1
                        if sus_data['homnum'] >= times*sus_data['fomnum']:
                            return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        # -------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 根据怀疑度大小重组
    def EDpNSpMC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        times = times/3

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # -------------------------------------------------------------------------------------------------
        # MC
        points_data = list(map(lambda x: x['out_list'], fom_list))
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        mc_som_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, times)

        out_dic = dict()
        kill_dic = dict()
        for som_message in mc_som_list:
            try:
                loc = hom_list_simple.index(som_message)
            except:
                continue
            hom = hom_out_list[loc]
            sus_data['totaltime'] += hom['time']
            sus_data['homnum'] += 1
            precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
            varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                enumerate(varietys)))
            for fom_in_hom in hom['message']:
                line = fom_in_hom[0]
                if line not in out_dic:
                    out_dic[line] = []
                    kill_dic[line] = []
                out_dic[line].append(hom['out_list'])
                kill_dic[line].append(hom['kill_list'])

        # ----------------------------------------------------------------------------------------------
        # ED
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        # ----------------------------------------------------------------------------------------------
        # NS
        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_line_dict.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        guide = ''
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0],
                                            (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2
                                            +fomnum*times/2/len(sus_list)],
                                 enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])
        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*math.ceil(fomnum*times/tnum)],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

            while resnum >= 1 and len(random_hom_lines) > 0:
                hom = random_hom_lines.pop(0)
                sus_data['homnum'] += 1
                sus_data['totaltime'] += hom['time']
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                if len(random_hom_lines) == 0:
                    random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

        # -------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 根据怀疑度大小重组
    def EDpNSpFTC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        times = times/3

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # -------------------------------------------------------------------------------------------------
        # FTC
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)

        out_dic = dict()
        kill_dic = dict()
        restnum = 0
        for pair, neednum in pairpower:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        # ----------------------------------------------------------------------------------------------
        # ED
        mbfl.mbfl_for.type_mbfl = 'max'
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        sus_list_mbfl = []
        for line in touple['linedata']:
            touple_list = touple['tf'], \
                          touple['tp'], \
                          touple['linedata'][line]['f'], \
                          touple['linedata'][line]['p'], \
                          touple['f2p'], \
                          touple['p2f'], \
                          touple['linedata'][line]['kf'], \
                          touple['linedata'][line]['kp'], \
                          touple['linedata'][line]['nf'], \
                          touple['linedata'][line]['np'],
            sus = mbfl.mbfl_for.Ochiai(touple_list)
            sus_list_mbfl.append([line, sus])

        # 计算权重
        pairlist = []
        tw = sum(map(lambda x: x[1]+1.1, sus_list_mbfl))
        weight_list = list(map(lambda x: [x[0], x[1]+1.1/tw], sus_list_mbfl))
        tw = 0
        for line_1, weight_1 in weight_list:
            for line_2, weight_2 in weight_list:
                if line_1 >= line_2:
                    continue
                pairlist.append([[line_1, line_2], weight_1*weight_2])
                tw += weight_1*weight_2
        pairlist = sorted(map(lambda x: [x[0], x[1]/tw*fomnum*times], pairlist), key=lambda x: x[1], reverse=True)

        # 生成二阶变异体
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        restnum = 0
        for pair, neednum in pairlist:
            restnum += neednum
            som_pair_list = []
            if not pairable(pair, fom_line_dict):
                continue
            for fom0 in fom_line_dict[pair[0]]:
                for fom1 in fom_line_dict[pair[1]]:
                    if not MutationRule(fom0['message'][0], fom1['message'][0]):
                        continue
                    som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x:x[0]))
            random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
            while restnum > 0 and len(random_som_pair_list) > 0:
                som_message = random_som_pair_list.pop(0)
                try:
                    loc = hom_list_simple.index(som_message)
                except:
                    continue
                hom = hom_out_list[loc]
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                restnum -= 1

        # ----------------------------------------------------------------------------------------------
        # NS
        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_line_dict.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        guide = ''
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0],
                                            (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2
                                            +fomnum*times/2/len(sus_list)],
                                 enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])
        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*math.ceil(fomnum*times/tnum)],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

            while resnum >= 1 and len(random_hom_lines) > 0:
                hom = random_hom_lines.pop(0)
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                if len(random_hom_lines) == 0:
                    random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

        # -------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 根据怀疑度大小重组
    def MC_RS_FTC(self, data_json, times=10000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        classification_list = [0.5]

        def pairable(pair_self, fom_line_dict_self):
            for line in pair_self:
                if line not in fom_line_dict_self:
                    return False
            return True

        def classification(sus):
            for label, classificate in enumerate(classification_list):
                if sus > classificate:
                    return label
            return len(classification_list)

        def mutTouple(or_list, mut_kill):
            kf, kp, nf, np = 0, 0, 0, 0
            for j in range(len(or_list)):
                if or_list[j] == 1:
                    # 原始pass
                    if mut_kill[j] == 1:
                        # 变异体n
                        np += 1
                    else:
                        # 变异体k
                        kp += 1
                else:
                    # 原始fail
                    if mut_kill[j] == 1:
                        # 变异体n
                        kf += 1
                    else:
                        # 变异体k
                        nf += 1
            return [kf, kp, nf, np]

        def mutOchiai(touple):
            akf, akp, anf, anp = touple
            try:
                return akf/math.sqrt((akf+anf)*(akf+akp))
            except:
                return -1

        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))
        fom_line_dict = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if not line in fom_line_dict:
                fom_line_dict[line] = []
            fom_line_dict[line].append(fom)

        # --------------------------------
        # 计算RS高阶变异体
        rs_som_list = []
        fom_data_list = list(map(lambda x: [x["message"][0], mutOchiai(mutTouple(or_list, x["kill_list"]))], fom_list))
        for fom1, sus1 in random.sample(fom_data_list, len(fom_data_list)):
            for fom2, sus2 in random.sample(fom_data_list, len(fom_data_list)):
                if not classification(sus1) == classification(sus2):
                    continue
                if classification(sus1) > 0:
                    continue
                som_message = sorted([fom1, fom2], key=lambda x: x[0])
                rs_som_list.append(som_message)

        # --------------------------------
        # 计算MC高阶变异体
        points_data = list(map(lambda x: x['out_list'], fom_list))
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        mc_som_list = Tools().clusterDifClassRandomMix(clusterdata, fom_list, 10000)
        # mc_som_list = Tools().clusterSameClassRandomMix(clusterdata, fom_list, 10000)


        # ---------------------------------
        # 计算失败测试用例特征序列
        points_data = []
        for i, pf in enumerate(or_list):
            if pf == 1:
                # 跳过通过的测试用例
                continue

            # 失败测试用例对应需要使用的测试用例空间
            testcase_space = list(or_list)
            testcase_space[i] = 1

            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            points_data.append(list(
                map(lambda x: x[0], sorted(list(
                    map(lambda x: [x[0]+1, x[1][0]], enumerate(sus_list_sort))
                ), key=lambda x: x[1]))))

        # 对失败测试用例进行聚类
        clusterdata = Cluster_Fom.MSeer(points_data).ClusterMain()
        RankList = []
        maxlen = 0
        for key, value in clusterdata.items():
            testcase_space = list(or_list)
            failnum = -1
            for i, pf in enumerate(or_list):
                if pf == 0:
                    failnum += 1
                if failnum in value:
                    testcase_space[i] = 1
            # 计算相应测试空间下的rank
            mbfl.mbfl_for.type_mbfl = 'frequency'
            fom_out_dic = dict()
            fom_kill_dic = dict()
            for fom in fom_list:
                fom_message = fom["message"][0]
                fom_out_list = Tools().victor_bite(testcase_space, fom["out_list"])
                fom_kill_list = Tools().victor_bite(testcase_space, fom["kill_list"])
                sus_data['totaltime'] += fom["time"]
                if fom_message[0] not in fom_out_dic:
                    fom_out_dic[fom_message[0]] = []
                    fom_kill_dic[fom_message[0]] = []
                fom_out_dic[fom_message[0]].append(fom_out_list)
                fom_kill_dic[fom_message[0]].append(fom_kill_list)
            touple = mbfl.command.GetTouleList(Tools().victor_bite(testcase_space, or_list), fom_out_dic, fom_kill_dic)
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = mbfl.mbfl_for.Ochiai(touple_list)
                sus_list.append([line, sorted(sus, reverse=True)])
            maxlen = max(maxlen, len(sus_list))
            sus_list = sorted(sus_list, key=lambda x: x[0], reverse=True)
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            sus_list_rank = list(map(lambda x: [x[1][0], x[0]], enumerate(sus_list_sort)))
            RankList.append(sorted(sus_list_rank, key=lambda x: x[0]))

        # 根据聚类结果计算权重数量
        pairpower = []
        if len(RankList) == 1:
            tw = 0
            for line1, power1 in RankList[0]:
                for line2, power2 in RankList[0]:
                    pairpower.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                    tw += (maxlen-power1)*(maxlen-power2)
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        else:
            for i, Rank1 in enumerate(RankList):
                for j, Rank2 in enumerate(RankList):
                    if i >= j:
                        continue
                    Spair = []
                    tw = 0
                    for line1, power1 in Rank1:
                        for line2, power2 in Rank2:
                            Spair.append([sorted([line1, line2]), (maxlen-power1)*(maxlen-power2)])
                            tw += (maxlen-power1)*(maxlen-power2)
                    Spair = list(map(lambda x: [x[0], x[1]/tw], Spair))
                    for pair, power in Spair:
                        loc = -1
                        for loc_i in range(len(pairpower)):
                            if pair == pairpower[i][0]:
                                loc = loc_i
                                break
                        if loc == -1:
                            pairpower.append([pair, power/tw])
                        else:
                            pairpower[loc][1] += power/tw
            tw = sum(list(map(lambda x: x[1], pairpower)))
            pairpower = list(map(lambda x: [x[0], x[1]/tw*times*fomnum], pairpower))
        pairpower = sorted(pairpower, key=lambda x: x[1], reverse=True)

        def setdic():
            out_dic = dict()
            kill_dic = dict()
            restnum = 0
            while sus_data['homnum'] < times*sus_data['fomnum']:
                for pair, neednum in pairpower:
                    restnum += neednum
                    som_pair_list = []
                    if not pairable(pair, fom_line_dict):
                        continue
                    for fom0 in fom_line_dict[pair[0]]:
                        for fom1 in fom_line_dict[pair[1]]:
                            if not MutationRule(fom0['message'][0], fom1['message'][0]):
                                continue
                            som_pair_list.append(sorted([fom0['message'][0], fom1['message'][0]], key=lambda x: x[0]))
                    random_som_pair_list = random.sample(som_pair_list, len(som_pair_list))
                    while restnum > 0 and len(random_som_pair_list) > 0:
                        som_message = random_som_pair_list.pop(0)
                        if not som_message in rs_som_list and som_message not in mc_som_list:
                            continue
                        try:
                            loc = hom_list_simple.index(som_message)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        sus_data['totaltime'] += hom['time']
                        sus_data['homnum'] += 1
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
                        restnum -= 1
                    if sus_data['homnum'] >= times*sus_data['fomnum']:
                        return out_dic, kill_dic
        out_dic, kill_dic = setdic()

        # -------------------------------
        # 计算怀疑度
        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])


                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    def test(self, data_json):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = len(hom_out_list)
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        # print(Fault_Record)
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_lines_dic = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_lines_dic:
                fom_lines_dic[line] = []
            fom_lines_dic[line].append(fom)

        data = []
        for line1, fom1_list in fom_lines_dic.items():
            out_dic = dict()
            kill_dic = dict()
            for line2, fom2_list in fom_lines_dic.items():
                if line1 == line2:
                    continue
                for fom1 in fom1_list:
                    for fom2 in fom2_list:
                        som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                        try:
                            loc = hom_list_simple.index(som)
                        except:
                            continue
                        hom = hom_out_list[loc]
                        for fom_in_hom in hom['message']:
                            line = fom_in_hom[0]
                            if line not in out_dic:
                                out_dic[line] = []
                                kill_dic[line] = []
                            out_dic[line].append(hom['out_list'])
                            kill_dic[line].append(hom['kill_list'])
            touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
            mbfl.mbfl_for.type_mbfl = 'frequency'
            mbfl_for = 'mbfl.mbfl_for.Ochiai'
            sus_list = []
            for line in touple['linedata']:
                touple_list = touple['tf'], \
                              touple['tp'], \
                              touple['linedata'][line]['f'], \
                              touple['linedata'][line]['p'], \
                              touple['f2p'], \
                              touple['p2f'], \
                              touple['linedata'][line]['kf'], \
                              touple['linedata'][line]['kp'], \
                              touple['linedata'][line]['nf'], \
                              touple['linedata'][line]['np'],
                sus = eval(mbfl_for)(touple_list)
                sus_list.append([line, sus])
            sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
            s = list(map(lambda x:x[0], sus_list_sort))
            # print(line1, s)
            data.append([line1, sus_list_sort])
        # print('')
        s = []
        for line1, sus1 in data:
            suslist = []
            for line2, sus2 in data:
                if line1 == line2:
                    continue
                suslist.append(sum(map(lambda x: x[0]if x[1][0] == line1 else 0, enumerate(sus2))))
            sus_ = sum(suslist)/len(suslist)
            variance = sum(map(lambda x: (x/sus_-1)*(x/sus_-1), suslist))
            print(line1, variance)
            s.append([line1, variance])
        sus_list_sort = sorted(s, key=lambda x: x[1], reverse=True)
        ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

        rank = [[], [], []]
        for fault in ranks:
            rank[0].append(fault[1])
            rank[1].append(fault[2])
            rank[2].append(fault[3])
        sus_data[mbfl_for]['rank_best'] = rank[0]
        sus_data[mbfl_for]['rank_average'] = rank[1]
        sus_data[mbfl_for]['rank_worst'] = rank[2]

        exam = [[], [], []]
        for ith in range(len(ranks)):
            if line_len > 0:
                exam[0].append(rank[0][ith]/line_len)
                exam[1].append(rank[1][ith]/line_len)
                exam[2].append(rank[2][ith]/line_len)
            else:
                exam[0].append(rank[0][ith])
                exam[1].append(rank[1][ith])
                exam[2].append(rank[2][ith])

        sus_data[mbfl_for]['exam_best'] = exam[0]
        sus_data[mbfl_for]['exam_average'] = exam[1]
        sus_data[mbfl_for]['exam_worst'] = exam[2]

    def NS(self, data_json, seting=[0], times=1000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            # sus_data['homnum'] = len(hom_out_list)
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        # print(Fault_Record)
        precisions = [0 for i in range(3)]
        varietys = [0 for i in range(13)]
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_lines_dic = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_lines_dic:
                fom_lines_dic[line] = []
            fom_lines_dic[line].append(fom)

        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_lines_dic.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        # ------------------------------------------------
        if not seting[0] == 0:
            if seting[0] == 1:
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif seting[0] == 2:
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2 + fomnum*times/2/len(sus_list)], enumerate(sus_list)))
            # powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times], enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])
        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*math.ceil(fomnum*times/tnum)],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        out_dic = dict()
        kill_dic = dict()
        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            # random_hom_lines = hom_lines_dic[line]
            random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

            while resnum >= 1 and len(random_hom_lines) > 0:
                # for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                hom = random_hom_lines.pop(0)
                sus_data['totaltime'] += hom['time']
                sus_data['homnum'] += 1
                precisions[Tools().precision_calculate(Fault_Record, hom['message'])] += 1
                varietys = list(map(lambda x: x[1]+1 if x[0] in Tools().variety_calculate(hom['message']) else x[1],
                                    enumerate(varietys)))
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                if len(random_hom_lines) == 0:
                    random_hom_lines = random.sample(hom_lines_dic[line], len(hom_lines_dic[line]))

        touple = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)

        alpha_part = [[], []]
        for line in touple['linedata']:
            alpha_part[0] += touple['linedata'][line]['f']
            alpha_part[1] += touple['linedata'][line]['p']
        try:
            alpha = str([
                sum(alpha_part[0])/len(alpha_part[0])/touple['tf'],
                sum(alpha_part[1])/len(alpha_part[1])/touple['tp'],
                ])
        except:
            alpha = str([0, 0])
        sus_data['Desrcibe'] = alpha

        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            mbfl.mbfl_for.type_mbfl = word
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list.append([line, sus])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]

        sus_data['precision'] = precisions
        sus_data['variety'] = varietys
        return sus_data

    # 仅使用高阶的NS-Delta
    def DeltaNS(self, data_json, times=1000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            # sus_data['homnum'] = len(hom_out_list)
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        # ------------------------------------------------
        # 计算一阶怀疑度
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # print(Fault_Record)
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_lines_dic = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_lines_dic:
                fom_lines_dic[line] = []
            fom_lines_dic[line].append(fom)

        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_lines_dic.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        # ------------------------------------------------
        guide = ''
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2 + fomnum*times/2/len(sus_list)], enumerate(sus_list)))
            # powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times], enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])

        else:
            powerlist = list(map(lambda x: [x[0], min(len(x[1])*math.ceil(fomnum*times/tnum), len(x[1]))],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        out_dic = dict()
        kill_dic = dict()
        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            hom_lines = hom_lines_dic[line]
            for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                sus_data['homnum'] += 1


        touple_ns = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        for word in ['max', 'ave', 'frequency', 'none']:
            mbfl.mbfl_for.type_mbfl = 'frequency'
            sus_data[word] = dict()
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list_T = dict()
                for line in touple_mbfl['linedata']:
                    touple_list = touple_mbfl['tf'], \
                                  touple_mbfl['tp'], \
                                  touple_mbfl['linedata'][line]['f'], \
                                  touple_mbfl['linedata'][line]['p'], \
                                  touple_mbfl['f2p'], \
                                  touple_mbfl['p2f'], \
                                  touple_mbfl['linedata'][line]['kf'], \
                                  touple_mbfl['linedata'][line]['kp'], \
                                  touple_mbfl['linedata'][line]['nf'], \
                                  touple_mbfl['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] = sus

                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple_ns['linedata']:
                    touple_list = touple_ns['tf'], \
                                  touple_ns['tp'], \
                                  touple_ns['linedata'][line]['f'], \
                                  touple_ns['linedata'][line]['p'], \
                                  touple_ns['f2p'], \
                                  touple_ns['p2f'], \
                                  touple_ns['linedata'][line]['kf'], \
                                  touple_ns['linedata'][line]['kp'], \
                                  touple_ns['linedata'][line]['nf'], \
                                  touple_ns['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] += sus
                    sus_line_ave = sum(sus_list_T[line])/len(sus_list_T[line])
                    if word == 'max':
                        sus_list.append([line, max(sus)-sus_line_ave])
                    elif word == 'ave':
                        sus_list.append([line, sum(sus)/len(sus)-sus_line_ave])
                    elif word == 'frequency':
                        sus_list.append([line, list(map(lambda x:x-sus_line_ave, sus))])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

    # 使用高阶NS-Delta 和 mbfl-Delta
    def DeltaNsMbfl(self, data_json, times=1000):
        if not data_json:
            return False
        try:
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            del data_json

            # 获取所需数据
            or_list = data["or_list"]
            line_len = data['linelen']
            Fault_Record = data["Fault_Record"]
            fom_list = data['fom_list']
            fomnum = len(fom_list)
            # hom_list_all = data['hom_list_all']
            hom_out_list = data['hom_out_list']
            sbfl_data = data["sbfl"]

            sus_data = dict()
            sus_data["Fault_Record"] = Fault_Record
            sus_data['totaltime'] = 0
            sus_data['fomnum'] = fomnum
            sus_data['homnum'] = 0
            # sus_data['homnum'] = len(hom_out_list)
            sus_data['Desrcibe'] = ''
            del data
        except Exception as e:
            print('数据缺失 %s' % e)
            return False

        def NsRule(message1, meggage2):
            if not message1[0] == meggage2[0]:
                return True
            elif message1[3] == 'delet' or meggage2[3] == 'delet':
                return False
            elif not message1[3] == meggage2[3] or \
                    (message1[5] < meggage2[5] and message1[5]+len(message1[3]) > meggage2[5]) or \
                    (meggage2[5] < message1[5] and meggage2[5]+len(meggage2[3]) > message1[5]):
                return True
            else:
                return False

        def WordNone(sus_list_format):
            frequency_dict = dict()
            for sus in sus_list_format:
                if sus not in frequency_dict:
                    frequency_dict[sus] = 0
                frequency_dict[sus] += 1
            return max(sorted(list(map(lambda x: [frequency_dict[x], x],
                                       list(frequency_dict.keys()))),
                              key=lambda x: x[1], reverse=True))[1]

        # ------------------------------------------------
        # 计算一阶怀疑度
        fom_out_dic = dict()
        fom_kill_dic = dict()
        for fom in fom_list:
            fom_message = fom["message"][0]
            fom_out_list = fom["out_list"]
            fom_kill_list = fom["kill_list"]
            sus_data['totaltime'] += fom["time"]
            if fom_message[0] not in fom_out_dic:
                fom_out_dic[fom_message[0]] = []
                fom_kill_dic[fom_message[0]] = []
            fom_out_dic[fom_message[0]].append(fom_out_list)
            fom_kill_dic[fom_message[0]].append(fom_kill_list)
        touple_mbfl = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)

        # print(Fault_Record)
        hom_list_simple = list(map(lambda x: sorted(x['message'], key=lambda y: y[0]), hom_out_list))

        fom_lines_dic = dict()
        for fom in fom_list:
            line = fom['message'][0][0]
            if line not in fom_lines_dic:
                fom_lines_dic[line] = []
            fom_lines_dic[line].append(fom)

        hom_lines_dic = dict()
        tnum = 0
        for line, fom_list_indic in fom_lines_dic.items():
            hom_lines_dic[line] = []
            for i1, fom1 in enumerate(fom_list_indic):
                for i2, fom2 in enumerate(fom_list_indic):
                    if i1 >= i2:
                        continue
                    if not NsRule(fom1['message'][0], fom2['message'][0]):
                        continue
                    som = sorted([fom1['message'][0], fom2['message'][0]], key=lambda x: x[0])
                    try:
                        loc = hom_list_simple.index(som)
                    except:
                        continue
                    hom_lines_dic[line].append(hom_out_list[loc])
                    tnum += 1

        # ------------------------------------------------
        guide = ''
        if guide == 'sbfl' or guide == 'mbfl':
            if guide == 'mbfl':
                # 计算mbfl怀疑度
                mbfl.mbfl_for.type_mbfl = 'frequency'
                fom_out_dic = dict()
                fom_kill_dic = dict()
                for fom in fom_list:
                    fom_message = fom["message"][0]
                    fom_out_list = fom["out_list"]
                    fom_kill_list = fom["kill_list"]
                    sus_data['totaltime'] += fom["time"]
                    if fom_message[0] not in fom_out_dic:
                        fom_out_dic[fom_message[0]] = []
                        fom_kill_dic[fom_message[0]] = []
                    fom_out_dic[fom_message[0]].append(fom_out_list)
                    fom_kill_dic[fom_message[0]].append(fom_kill_list)
                touple = mbfl.command.GetTouleList(or_list, fom_out_dic, fom_kill_dic)
                sus_list = []
                for line in touple['linedata']:
                    touple_list = touple['tf'], \
                                  touple['tp'], \
                                  touple['linedata'][line]['f'], \
                                  touple['linedata'][line]['p'], \
                                  touple['f2p'], \
                                  touple['p2f'], \
                                  touple['linedata'][line]['kf'], \
                                  touple['linedata'][line]['kp'], \
                                  touple['linedata'][line]['nf'], \
                                  touple['linedata'][line]['np'],
                    sus = mbfl.mbfl_for.Ochiai(touple_list)
                    sus_list.append([line, sus])
            elif guide == 'sbfl':
                # 计算sbfl怀疑度
                touple = sbfl.command.touple(or_list, sbfl_data['cov'])
                sus_list = []
                for line, touple_list in touple.items():
                    sus = sbfl.sbfl_for.GP13(touple_list)
                    sus_list.append([line, sus])

            totalpower = sum(map(lambda x: len(sus_list)-x[0]+1, enumerate(sus_list)))
            powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times/2 + fomnum*times/2/len(sus_list)], enumerate(sus_list)))
            # powerlist = list(map(lambda x: [x[1][0], (len(sus_list)-x[0]+1)/totalpower*fomnum*times], enumerate(sus_list)))
            powerlist_sorted = sorted(powerlist, key=lambda x:x[0])

        else:
            powerlist = list(map(lambda x: [x[0], len(x[1])*fomnum*times/tnum],
                                 hom_lines_dic.items()))
            powerlist_sorted = sorted(powerlist, key=lambda x: x[0])

        out_dic = dict()
        kill_dic = dict()
        resnum = 0
        for line, neednum in powerlist_sorted:
            resnum += neednum
            if line not in hom_lines_dic:
                continue
            hom_lines = hom_lines_dic[line]
            for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                for fom_in_hom in hom['message']:
                    line = fom_in_hom[0]
                    if line not in out_dic:
                        out_dic[line] = []
                        kill_dic[line] = []
                    out_dic[line].append(hom['out_list'])
                    kill_dic[line].append(hom['kill_list'])
                resnum -= 1
                sus_data['homnum'] += 1

        resnum = math.ceil(resnum)
        while resnum > 0:
            for line, neednum in powerlist_sorted:
                if line not in hom_lines_dic:
                    continue
                hom_lines = hom_lines_dic[line]
                for hom in random.sample(hom_lines, min(math.ceil(resnum), len(hom_lines))):
                    for fom_in_hom in hom['message']:
                        line = fom_in_hom[0]
                        if line not in out_dic:
                            out_dic[line] = []
                            kill_dic[line] = []
                        out_dic[line].append(hom['out_list'])
                        kill_dic[line].append(hom['kill_list'])
                    resnum -= 1
                    sus_data['homnum'] += 1
                    if resnum == 0:
                        break
                if resnum == 0:
                    break

        touple_ns = mbfl.command.GetTouleList(or_list, out_dic, kill_dic)
        mbfl.mbfl_for.type_mbfl = 'frequency'
        for word in ['max', 'ave', 'frequency', 'none']:
            sus_data[word] = dict()
            for j, mbfl_for in enumerate(mbfl_for_list):
                sus_list_T = dict()
                for line in touple_mbfl['linedata']:
                    touple_list = touple_mbfl['tf'], \
                                  touple_mbfl['tp'], \
                                  touple_mbfl['linedata'][line]['f'], \
                                  touple_mbfl['linedata'][line]['p'], \
                                  touple_mbfl['f2p'], \
                                  touple_mbfl['p2f'], \
                                  touple_mbfl['linedata'][line]['kf'], \
                                  touple_mbfl['linedata'][line]['kp'], \
                                  touple_mbfl['linedata'][line]['nf'], \
                                  touple_mbfl['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] = sus

                sus_list = []
                sus_data[word][mbfl_for] = dict()
                for line in touple_ns['linedata']:
                    touple_list = touple_ns['tf'], \
                                  touple_ns['tp'], \
                                  touple_ns['linedata'][line]['f'], \
                                  touple_ns['linedata'][line]['p'], \
                                  touple_ns['f2p'], \
                                  touple_ns['p2f'], \
                                  touple_ns['linedata'][line]['kf'], \
                                  touple_ns['linedata'][line]['kp'], \
                                  touple_ns['linedata'][line]['nf'], \
                                  touple_ns['linedata'][line]['np'],
                    sus = eval(mbfl_for)(touple_list)
                    sus_list_T[line] += sus

                    sus_line_ave = sum(sus_list_T[line])/len(sus_list_T[line])
                    if word == 'max':
                        sus_list.append([line, max(sus_list_T[line])-sus_line_ave])
                    elif word == 'ave':
                        sus_list.append([line, sum(sus_list_T[line])/len(sus_list_T[line])-sus_line_ave])
                    elif word == 'frequency':
                        sus_list.append([line, list(map(lambda x:x-sus_line_ave, sus_list_T[line]))])
                    elif word == 'none':
                        sus_list.append([line, WordNone(sus_list_T[line])-sus_line_ave])

                sus_list_sort = sorted(sus_list, key=lambda x: x[1], reverse=True)
                ranks = util.Evaluation().rank(sus_list_sort, Fault_Record, line_len)

                rank = [[], [], []]
                for fault in ranks:
                    rank[0].append(fault[1])
                    rank[1].append(fault[2])
                    rank[2].append(fault[3])
                sus_data[word][mbfl_for]['rank_best'] = rank[0]
                sus_data[word][mbfl_for]['rank_average'] = rank[1]
                sus_data[word][mbfl_for]['rank_worst'] = rank[2]

                exam = [[], [], []]
                for ith in range(len(ranks)):
                    if line_len > 0:
                        exam[0].append(rank[0][ith]/line_len)
                        exam[1].append(rank[1][ith]/line_len)
                        exam[2].append(rank[2][ith]/line_len)
                    else:
                        exam[0].append(rank[0][ith])
                        exam[1].append(rank[1][ith])
                        exam[2].append(rank[2][ith])
                    if rank[0][ith] > rank[1][ith] or rank[1][ith] > rank[2][ith]:
                        print('exam异常', rank[0][ith], rank[1][ith],  rank[2][ith])

                sus_data[word][mbfl_for]['exam_best'] = exam[0]
                sus_data[word][mbfl_for]['exam_average'] = exam[1]
                sus_data[word][mbfl_for]['exam_worst'] = exam[2]
        return sus_data

def CHMBFL():

    def CHMBFL_mbfl_baseline_res():
        baseline_list = [
            ['Mbfl', Get_sus().Mbfl],
            # ['DeltaMbfl', Get_sus().DeltaMbfl],
            # ['Mcbfl', Get_sus().Mcbfl],
            # ['MUSE', Get_sus().Muse],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3

        for method, function in baseline_list:
            # 表格初始化
            wb = openpyxl.Workbook()
            Wss = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
            del wb['Sheet']
            for ws in Wss:
                ws.cell(1, 1, method)
                for j, title in enumerate(ftitle):
                    ws.cell(2, j+1, title)
                for j, mbfl_for in enumerate(mbfl_for_list):
                    ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                    for k in range(maxfaultnum):
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))

            # 获取数据
            for file_i, file in enumerate(os.listdir(data_dirpath)):

                print('%s %s, read:%s-%s' % (datetime.datetime.now(), method, file, file_i))

                # 计算相应方法的怀疑度
                path = os.path.join(data_dirpath, file)
                sus_data_json = function(Tools().json_rules(path))

                sheetlist = [
                    [Wss[0], 'best'],
                    [Wss[1], 'average'],
                    [Wss[2], 'worst'],
                ]
                for sheet, name in sheetlist:
                    sheetdata_list = [
                        str(file_i+1),
                        file[5:-5],
                    ]
                    if not sus_data_json:
                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                        continue
                    sheetdata_list += [
                        len(sus_data_json['Fault_Record']),
                        sus_data_json['totaltime'],
                        sus_data_json['fomnum'],
                        sus_data_json['homnum'],
                        '',
                    ]
                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                        examlist = []
                        ranklist = []
                        for fault_ith in range(maxfaultnum):
                            if fault_ith > len(sus_data_json[mbfl_for]['rank_%s' % name])-1:
                                examlist.append('')
                                ranklist.append('')
                            else:
                                examlist.append(sus_data_json[mbfl_for]['exam_%s' % name][fault_ith])
                                ranklist.append(sus_data_json[mbfl_for]['rank_%s' % name][fault_ith])
                        sheetdata_list += examlist
                        sheetdata_list += ranklist
                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

            docpath = os.path.join(os.getcwd(), 'report', 'CHMBFL', 'susfile', 'baseline', method, 'max')
            if not os.path.exists(docpath):
                os.makedirs(docpath)

            filepath = os.path.join(docpath, '%s_rank.xlsx' % method)
            wb.save(filepath)
            print('存储成功 %s' % filepath)

        return

    def CHMBFL_sbfl_baseline_res():
        # 初始化表格
        functionlist = [
            ['Sbfl', Get_sus().Sbfl],
        ]

        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3

        for method, function in functionlist:
            # 表格初始化
            wb = openpyxl.Workbook()
            Wss = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
            del wb['Sheet']
            for ws in Wss:
                ws.cell(1, 1, method)
                for j, title in enumerate(ftitle):
                    ws.cell(2, j+1, title)
                for j, mbfl_for in enumerate(mbfl_for_list):
                    ws.cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                    for k in range(maxfaultnum):
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                        ws.cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))


            # 获取数据
            # sbfl_dirpath = './report/SBFL'
            for file_i, file in enumerate(os.listdir(data_dirpath)):
                print('%s %s, read:%s-%s' % (datetime.datetime.now(), method, file, file_i))
                # 计算相应方法的怀疑度
                path = os.path.join(data_dirpath, file)
                sus_data_json = function(Tools().json_rules(path))

                sheetlist = [
                    [Wss[0], 'best'],
                    [Wss[1], 'average'],
                    [Wss[2], 'worst'],
                ]
                for sheet, name in sheetlist:
                    sheetdata_list = [
                        str(file_i+1),
                        file[5:-5],
                    ]
                    if not sus_data_json:
                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                        continue
                    sheetdata_list += [
                        len(sus_data_json['Fault_Record']),
                        sus_data_json['totaltime'],
                        '',
                        '',
                        '',
                    ]
                    for sbfl_for_j, sbfl_for in enumerate(sbfl_for_list):
                        examlist = []
                        ranklist = []
                        for fault_ith in range(maxfaultnum):
                            if fault_ith > len(sus_data_json[sbfl_for]['rank_%s' % name])-1:
                                examlist.append('')
                                ranklist.append('')
                            else:
                                examlist.append(sus_data_json[sbfl_for]['exam_%s' % name][fault_ith])
                                ranklist.append(sus_data_json[sbfl_for]['rank_%s' % name][fault_ith])
                        sheetdata_list += examlist
                        sheetdata_list += ranklist
                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)



            # doc_rootpath = os.path.join(os.getcwd(), 'CHMBFL', 'susfile', 'baseline')
            doc_rootpath = os.path.join(os.getcwd(), 'report', 'CHMBFL', 'susfile', 'baseline', method)
            if not os.path.exists(doc_rootpath):
                os.makedirs(doc_rootpath)
            filepath = os.path.join(doc_rootpath, '%s_rank.xlsx' % method)
            wb.save(filepath)
            print('存储成功 %s' % filepath)

        return


    # 需要倍数， 需要重复次数
    def CHMBFL_timesneed_repeatneed_res():
        # 初始化表格
        clasterfunction = [
            ['Random', Get_sus().Random],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3
        for method, function in clasterfunction:
            for times in range(1, 20):
                for repeat in range(1, 20):
                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):
                        print('%s baseline:%s, times:%s repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, times, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)
                        sus_data_json = function(Tools().json_rules(path), times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    '',
                                ]
                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)


                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency'][wb_i]
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/numless', method, str(times), word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%s_%sth_ranke.xlsx' % (method, times, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def CHMBFL_timesable_repeatneed_res():
        # 初始化表格
        clasterfunction = []
        if True:
            clasterfunction += [
                # ['DeltaNS', Get_sus().DeltaNS],
                ['DeltaNsMbfl', Get_sus().DeltaNsMbfl],
            ]
        if False:
            clasterfunction += [
                ['ED.MBFL', Get_sus().ErrorDistribution, [True]],
                ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
                ['FTC', Get_sus().FailTestCluster, [True]],
                ['FTC.kmeans', Get_sus().FailTestCluster, [False]],
                ['MutCluster', Get_sus().MutCluster, [True, True]],
                ['MutCluster.in', Get_sus().MutCluster, [True, False]],
                ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
                ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
                ['NS.RANDOM', Get_sus().NS, [0]],
                ['NS.SBFL', Get_sus().NS, [1]],
                ['NS.MBFL', Get_sus().NS, [2]],
                ['NS.mbfl*FTC', Get_sus().NSpFTC],
                ['NS.mbfl*MC', Get_sus().NSpMC],
                ['NS.mbfl*ED', Get_sus().NSpED],
                ['MC.mseer*ED', Get_sus().MCpED],
                ['ED*NS*MC.mseer', Get_sus().EDpNSpMC],
                ['ED*NS*FTC.mseer', Get_sus().EDpNSpFTC],
            ]


        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',

        ]
        # for i in range(13):
        #     ftitle.append('Operator Type %s' % str(i))
        # ftitle += [
        #     'None Accurate',
        #     'Pare Accurate',
        #     'Accurate',
        # ]
        maxfaultnum = 3
        # times = 11

        # for times in [1, 3, 5, 7, 10, 20, 50, 100]:
        for times in [0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4, 4.5, 5]:
            for parameter in clasterfunction:
                method = parameter[0] + '-%s' % str(times)
                function = parameter[1]
                for repeat in range(1, 40):
                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):

                        print('%s baseline:%s, repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)

                        if len(parameter) == 2:
                            sus_data_json = function(Tools().json_rules(path), times)
                        else:
                            sus_data_json = function(Tools().json_rules(path), parameter[2], times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency', 'none'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    '',
                                ]
                                # varietys = sus_data_json['variety']
                                # sum_varietys = sum(varietys)
                                # for variety in varietys:
                                #     try:
                                #         sheetdata_list.append(variety/sum_varietys)
                                #     except:
                                #         sheetdata_list.append(0)
                                #
                                # precisions = sus_data_json['precision']
                                # sum_precisions = sum(precisions)
                                # for precision in precisions:
                                #     try:
                                #         sheetdata_list.append(precision/sum_precisions)
                                #     except:
                                #         sheetdata_list.append(0)

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency', 'none'][wb_i]
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K/', method, word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def whf_res():
        # 初始化表格
        clasterfunction = []
        if True:
            clasterfunction += [
                ['ED.MBFL', Get_sus().ErrorDistribution, [True]],
                ['ED.SBFL', Get_sus().ErrorDistribution, [False]],
                ['FTC', Get_sus().FailTestCluster, [True]],
                ['FTC.kmeans', Get_sus().FailTestCluster, [False]],
                ['MutCluster', Get_sus().MutCluster, [True, True]],
                ['MutCluster.in', Get_sus().MutCluster, [True, False]],
                ['MutCluster.kmeans', Get_sus().MutCluster, [False, True]],
                ['MutCluster.kmeans.in', Get_sus().MutCluster, [False, False]],
                ['NS.RANDOM', Get_sus().NS, [0]],
                ['NS.SBFL', Get_sus().NS, [1]],
                ['NS.MBFL', Get_sus().NS, [2]],
                ['NS.mbfl^FTC', Get_sus().NSpFTC],
                ['NS.mbfl^MC', Get_sus().NSpMC],
                ['NS.mbfl^ED', Get_sus().NSpED],
                ['MC.mseer^ED', Get_sus().MCpED],
                ['ED^NS^MC.mseer', Get_sus().EDpNSpMC],
                ['ED^NS^FTC.mseer', Get_sus().EDpNSpFTC],
            ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',

        ]
        for i in range(13):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['202.4.157.11', '202.4.157.19', '222.199.230.148', '202.4.130.29'].index(Tools().get_host_ip())
            timeslist = [
                [1, 2, 3],
                [4, 5, 6],
                [7, 8, 9],
                [10, 11],
            ][service]
            timeslist = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        except:
            # return 'IP 错误'
            timeslist = [11]

        for times in timeslist:
            for parameter in clasterfunction:
                method = parameter[0] + '-%s' % str(times)
                function = parameter[1]
                for repeat in range(1, 10):

                    if not Rewrite:
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, 'none')
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, 'none', repeat))
                        if os.path.exists(filepath):
                            continue


                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):

                        print('%s baseline:%s, repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)

                        if len(parameter) == 2:
                            sus_data_json = function(Tools().json_rules(path), times=times)
                        else:
                            sus_data_json = function(Tools().json_rules(path), parameter[2], times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency', 'none'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    sus_data_json['Desrcibe'],
                                ]
                                varietys = sus_data_json['variety']
                                sum_varietys = sum(varietys)
                                for variety in varietys:
                                    try:
                                        sheetdata_list.append(variety/sum_varietys)
                                    except:
                                        sheetdata_list.append(0)

                                precisions = sus_data_json['precision']
                                sum_precisions = sum(precisions)
                                for precision in precisions:
                                    try:
                                        sheetdata_list.append(precision/sum_precisions)
                                    except:
                                        sheetdata_list.append(0)

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency', 'none'][wb_i]
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 给定倍数，需要重复次数
    def yojanTest():
        # 初始化表格
        clasterfunction = []
        reduction_list = [
            # ['testanalysis', Reduction().testanalysis],
            # ['samping', Reduction().samping],
            # ['some', Reduction().some],
            # ['wsome', Reduction().wsome],
            ['none', Reduction().none],
        ]
        clasterfunction += [
            ['DifferentOp', Get_sus().DifferentOperator],
            ['Last2First', Get_sus().Last2First],
            ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],
            ['ErrorDistribution', Get_sus().ErrorDistribution],
            ['MutCluster', Get_sus().MutCluster],
            ['NS_Random', Get_sus().NS],
            ['NS_SBFL', Get_sus().NS, [2]],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(13):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += ['None Accurate', 'Pare Accurate', 'Accurate',]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['202.4.157.19', '222.199.230.148'].index(Tools().get_host_ip())
            timeslist = [
                [1],
                [10],
            ][service]
        except:
            # return 'IP 错误'
            timeslist = [5]

        for times in timeslist:
            for parameter in clasterfunction:
                for reductionname, reductionfunction in reduction_list:
                    method = '%s-%s-%s' % (reductionname, parameter[0], str(times))
                    function = parameter[1]
                    for repeat in range(1, 3):
                        if not Rewrite:
                            docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, 'none')
                            filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, 'none', repeat))
                            if os.path.exists(filepath):
                                continue

                        # 表格初始化
                        Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                        Wss = []
                        for wb in Wbs:
                            ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                            del wb['Sheet']
                            for i in range(3):
                                ws[i].cell(1, 1, method)
                                for j, title in enumerate(ftitle):
                                    ws[i].cell(2, j+1, title)
                                for j, mbfl_for in enumerate(mbfl_for_list):
                                    ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                    for k in range(maxfaultnum):
                                        ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                        ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                            Wss.append(ws)

                        # 获取数据
                        for file_i, file in enumerate(os.listdir(data_dirpath)):

                            print('%s baseline:%s, reduction:%s, repeat:%s read:%s-%s' %
                                  (datetime.datetime.now(), method, reductionname, repeat, file, file_i))
                            # if file_i <= 124:
                            #     continue
                            # 计算相应方法的怀疑度
                            path = os.path.join(data_dirpath, file)
                            data_json = Tools().json_rules(path)
                            reduction_datajson = reductionfunction(data_json, times/10)
                            # print(datetime.datetime.now())

                            try:
                                if len(parameter) == 2:
                                    sus_data_json = function(reduction_datajson, times=1000)
                                else:
                                    sus_data_json = function(reduction_datajson, parameter[2], times=1000)
                            except:
                                sus_data_json = False
                            # print(datetime.datetime.now())

                            for ws_i, ws in enumerate(Wss):
                                word = ['max', 'ave', 'frequency', 'none'][ws_i]
                                sheetlist = [
                                    [ws[0], 'best'],
                                    [ws[1], 'average'],
                                    [ws[2], 'worst'],
                                ]
                                for sheet, name in sheetlist:
                                    sheetdata_list = [
                                        str(file_i+1),
                                        file[5:-5],
                                    ]
                                    if not sus_data_json:
                                        for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                            sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                        continue
                                    sheetdata_list += [
                                        len(sus_data_json['Fault_Record']),
                                        sus_data_json['totaltime'],
                                        sus_data_json['fomnum'],
                                        sus_data_json['homnum'],
                                        sus_data_json['Desrcibe'],
                                    ]
                                    varietys = sus_data_json['variety']
                                    sum_varietys = sum(varietys)
                                    for variety in varietys:
                                        try:
                                            sheetdata_list.append(variety/sum_varietys)
                                        except:
                                            sheetdata_list.append(0)

                                    precisions = sus_data_json['precision']
                                    sum_precisions = sum(precisions)
                                    for precision in precisions:
                                        try:
                                            sheetdata_list.append(precision/sum_precisions)
                                        except:
                                            sheetdata_list.append(0)

                                    for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                        examlist = []
                                        ranklist = []
                                        for fault_ith in range(maxfaultnum):
                                            if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                                examlist.append('')
                                                ranklist.append('')
                                            else:
                                                examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                                ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                        sheetdata_list += examlist
                                        sheetdata_list += ranklist
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                        # 存储文件
                        print('')
                        for wb_i, wb in enumerate(Wbs):
                            word = ['max', 'ave', 'frequency', 'none'][wb_i]
                            docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, word)
                            if not os.path.exists(docpath):
                                os.makedirs(docpath)
                            filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                            wb.save(filepath)
                            print('文件保存 %s ' % filepath)
        return

    # 不需要倍数，需要重复次数
    def CHMBFL_baseline_timesable_repeatneed_res():
        clasterfunction = [
            ['Last2First', Get_sus().Last2First],
            ['DifferentOperator', Get_sus().DifferentOperator],
            ['RandomMix', Get_sus().RandomMix],
            ['Random', Get_sus().Random],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        for i in range(13):
            ftitle.append('Operator Type %s' % str(i))
        ftitle += [
            'None Accurate',
            'Pare Accurate',
            'Accurate',
        ]
        maxfaultnum = 3

        try:
            # service = ['202.4.157.11', '202.4.157.19', '222.199.230.227', '202.4.130.29'].index(Tools().get_host_ip())
            service = ['222.199.230.148', '202.4.130.29'].index(Tools().get_host_ip())
            timeslist = [
                [1, 2, 3, 4, 5],
                [6, 7, 8, 9, 10, 11],
            ][service]
        except:
            # return 'IP 错误'
            timeslist = [11]

        for times in timeslist:
            # for method, function in clasterfunction:
            for parameter in clasterfunction:
                method = parameter[0] + '-%s' % str(times)
                function = parameter[1]
                for repeat in range(1, 10):
                    # 表格初始化
                    Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                    Wss = []
                    for wb in Wbs:
                        ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                        del wb['Sheet']
                        for i in range(3):
                            ws[i].cell(1, 1, method)
                            for j, title in enumerate(ftitle):
                                ws[i].cell(2, j+1, title)
                            for j, mbfl_for in enumerate(mbfl_for_list):
                                ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                                for k in range(maxfaultnum):
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                    ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                        Wss.append(ws)

                    # 获取数据
                    for file_i, file in enumerate(os.listdir(data_dirpath)):
                        print('%s baseline:%s, repeat:%s read:%s-%s' %
                              (datetime.datetime.now(), method, repeat, file, file_i))
                        # 计算相应方法的怀疑度
                        path = os.path.join(data_dirpath, file)
                        sus_data_json = function(Tools().json_rules(path), times)

                        for ws_i, ws in enumerate(Wss):
                            word = ['max', 'ave', 'frequency', 'none'][ws_i]
                            sheetlist = [
                                [ws[0], 'best'],
                                [ws[1], 'average'],
                                [ws[2], 'worst'],
                            ]
                            for sheet, name in sheetlist:
                                sheetdata_list = [
                                    str(file_i+1),
                                    file[5:-5],
                                ]
                                if not sus_data_json:
                                    for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                        sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                    continue
                                sheetdata_list += [
                                    len(sus_data_json['Fault_Record']),
                                    sus_data_json['totaltime'],
                                    sus_data_json['fomnum'],
                                    sus_data_json['homnum'],
                                    sus_data_json['Desrcibe'],
                                ]
                                varietys = sus_data_json['variety']
                                sum_varietys = sum(varietys)
                                for variety in varietys:
                                    try:
                                        sheetdata_list.append(variety/sum_varietys)
                                    except:
                                        sheetdata_list.append(0)

                                precisions = sus_data_json['precision']
                                sum_precisions = sum(precisions)
                                for precision in precisions:
                                    try:
                                        sheetdata_list.append(precision/sum_precisions)
                                    except:
                                        sheetdata_list.append(0)

                                for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                    examlist = []
                                    ranklist = []
                                    for fault_ith in range(maxfaultnum):
                                        if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                            examlist.append('')
                                            ranklist.append('')
                                        else:
                                            examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                            ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                    sheetdata_list += examlist
                                    sheetdata_list += ranklist
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                    # 存储文件
                    print('')
                    for wb_i, wb in enumerate(Wbs):
                        word = ['max', 'ave', 'frequency', 'none'][wb_i]
                        # docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/baseline', method, word)
                        docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/K', str(times), method, word)
                        if not os.path.exists(docpath):
                            os.makedirs(docpath)
                        filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                        wb.save(filepath)
                        print('文件保存 %s ' % filepath)
        return

    # 不需要倍数，需要重复次数
    def CHMBFL_timesless_repeatless_res():
        # 初始化表格
        clasterfunction = [
            # ['DifRank', Get_sus().test],
            ['RegroupBySus-New', Get_sus().NewRegroupBySus],
        ]
        ftitle = [
            'Id',
            'Version',
            'Fault_Record',
            'TimeSpend',
            'Fomnum',
            'Homnum',
            'Desrcibe',
        ]
        maxfaultnum = 3

        for method, function in clasterfunction:
            for repeat in range(1, 20):
                # 表格初始化
                Wbs = [openpyxl.Workbook(), openpyxl.Workbook(), openpyxl.Workbook()]
                Wss = []
                for wb in Wbs:
                    ws = [wb.create_sheet('exam-best'), wb.create_sheet('exam-average'), wb.create_sheet('exam-worst')]
                    del wb['Sheet']
                    for i in range(3):
                        ws[i].cell(1, 1, method)
                        for j, title in enumerate(ftitle):
                            ws[i].cell(2, j+1, title)
                        for j, mbfl_for in enumerate(mbfl_for_list):
                            ws[i].cell(1, j*maxfaultnum*2+len(ftitle)+1, mbfl_for.split('.')[2])
                            for k in range(maxfaultnum):
                                ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+k, 'exam%s' % str(k+1))
                                ws[i].cell(2, j*maxfaultnum*2+len(ftitle)+1+3+k, 'rank%s' % str(k+1))
                    Wss.append(ws)


                # 获取数据
                for file_i, file in enumerate(os.listdir(data_dirpath)):
                    print('%s baseline:%s, repeat:%s read:%s-%s' %
                          (datetime.datetime.now(), method, repeat, file, file_i))
                    # 计算相应方法的怀疑度
                    path = os.path.join(data_dirpath, file)
                    sus_data_json = function(Tools().json_rules(path))

                    for ws_i, ws in enumerate(Wss):
                        word = ['max', 'ave', 'frequency'][ws_i]
                        sheetlist = [
                            [ws[0], 'best'],
                            [ws[1], 'average'],
                            [ws[2], 'worst'],
                        ]
                        for sheet, name in sheetlist:
                            sheetdata_list = [
                                str(file_i+1),
                                file[5:-5],
                            ]
                            if not sus_data_json:
                                for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                    sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)
                                continue
                            sheetdata_list += [
                                len(sus_data_json['Fault_Record']),
                                sus_data_json['totaltime'],
                                sus_data_json['fomnum'],
                                sus_data_json['homnum'],
                                '',
                            ]
                            for mbfl_for_j, mbfl_for in enumerate(mbfl_for_list):
                                examlist = []
                                ranklist = []
                                for fault_ith in range(maxfaultnum):
                                    if fault_ith > len(sus_data_json[word][mbfl_for]['rank_%s' % name])-1:
                                        examlist.append('')
                                        ranklist.append('')
                                    else:
                                        examlist.append(sus_data_json[word][mbfl_for]['exam_%s' % name][fault_ith])
                                        ranklist.append(sus_data_json[word][mbfl_for]['rank_%s' % name][fault_ith])
                                sheetdata_list += examlist
                                sheetdata_list += ranklist
                            for sheetdata_i, sheetdata in enumerate(sheetdata_list):
                                sheet.cell(file_i + 3, sheetdata_i+1, sheetdata)

                # 存储文件
                print('')
                for wb_i, wb in enumerate(Wbs):
                    word = ['max', 'ave', 'frequency'][wb_i]
                    docpath = os.path.join(os.getcwd(), 'report/CHMBFL/susfile/baseline', method, word)
                    if not os.path.exists(docpath):
                        os.makedirs(docpath)
                    filepath = os.path.join(docpath, '%s_%s_%sth_ranke.xlsx' % (method, word, repeat))
                    wb.save(filepath)
                    print('文件保存 %s ' % filepath)
        return

    # CHMBFL_baseline_timesable_repeatneed_res()
    # whf_res()
    # yojanTest()
    # CHMBFL_sbfl_baseline_res()
    CHMBFL_mbfl_baseline_res()

    return

def getVersion():
    filepath = os.path.join('./FoundationFile/1.txt')
    versionlist = []
    with open(filepath, 'r') as f:
        data = f.readlines()
    for line, linedata in enumerate(data):
        if line%2 == 1:
            continue
        v = linedata.split()[3].split(':')[1].split('-')[0]
        versionlist.append(v)
    return versionlist


def compare(defect, true, Faults):
    Operators_dict = dict()
    Operators_dict['error'] = []
    # if not os.path.exists(defect):
    #     print('error', defect)
    #     return
    # if not os.path.exists(true):
    #     print('error', true)
    #     return
    diff_content = list(difflib.Differ().compare(defect, true))
    print(Faults)
    print(''.join(difflib.Differ().compare(defect, true)))
    return


def newdata():
    # 1751 删除？
    # 1605 1684 1713 1724 1732 1735 1750 修改？

    # version_list = getVersion()
    # version_list = ['v1605', 'v1684', 'v1713', 'v1724', 'v1732', 'v1735', 'v1750', 'v1751', ]
    version_list = ['v1735', 'v1751', ]
    for version_name in version_list:
        if 'Feature' not in version_name:
            version_name = 'Fom_%s_Feature.json' % version_name
        json_path = os.path.join('./report/CHMBFL/mutinfo-new', version_name)
        with open(json_path, 'r') as f:
            data_json = json.load(f)
            doc = list(data_json.keys())[0]
            data = data_json[doc]
            Faults = data["Fault_Record"]
        version = version_name.split('_')[1]
        defect_path = os.path.join('./cdata/version', version, 'test_data', './defect_root/source/tcas.c')
        with open(defect_path, 'r') as f:
            defect = f.readlines()
        true_path = os.path.join('./cdata/version', version, 'test_data', './true_root/source/tcas.c')
        with open(true_path, 'r') as f:
            true = f.readlines()
        print('---------------------------------------',
              '\n\n\n\n\n\n\n\n\n\n\n\n\n')
        print(version, 'mbfl')
        data_mbfl = Get_sus().Mbfl(data_json)
        print(version, 'hmbfl')
        data_hmbfl = Get_sus().Random(data_json, 100000)
        compare(defect, true, Faults)
        os.system('cls')
        # with open(json_path) as f_obj:
        #     data_json = json.load(f_obj)


if __name__ == '__main__':
    # CHMBFL()
    newdata()
    # getVersion()

